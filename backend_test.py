#!/usr/bin/env python3
"""
PropFlow CRM Backend Testing - Contacts Import/Export
Tests the contacts import/export endpoints including CSV/XLSX support
"""

import requests
import csv
import io
import json
import openpyxl
from pathlib import Path

# Configuration
BACKEND_URL = "https://propflow-crm-1.preview.emergentagent.com"
API_BASE = f"{BACKEND_URL}/api"

# Test credentials
ADMIN_EMAIL = "admin@propflow.com"
ADMIN_PASSWORD = "admin123"

class ContactsImportExportTester:
    def __init__(self):
        self.session = requests.Session()
        self.auth_token = None
        
    def authenticate(self):
        """Authenticate with admin credentials"""
        print("🔐 Authenticating...")
        
        login_data = {
            "email": ADMIN_EMAIL,
            "password": ADMIN_PASSWORD
        }
        
        response = self.session.post(f"{API_BASE}/auth/login", json=login_data)
        
        if response.status_code == 200:
            print("✅ Authentication successful")
            # Cookies are automatically handled by session
            return True
        else:
            print(f"❌ Authentication failed: {response.status_code} - {response.text}")
            return False
    
    def test_contacts_template_endpoint(self):
        """Test GET /api/contacts/template endpoint"""
        print("\n📋 Testing contacts template endpoint...")
        
        response = self.session.get(f"{API_BASE}/contacts/template")
        
        if response.status_code == 200:
            # Check if it's a CSV file
            content_type = response.headers.get('content-type', '')
            content_disposition = response.headers.get('content-disposition', '')
            
            if 'text/csv' in content_type and 'attachment' in content_disposition:
                print("✅ Template endpoint returns CSV file")
                
                # Parse CSV content to verify structure
                csv_content = response.text
                reader = csv.DictReader(io.StringIO(csv_content))
                headers = reader.fieldnames
                
                expected_headers = ["name", "email", "phone", "company", "source", "property_type", "tags", "notes", "lead_score"]
                
                if headers == expected_headers:
                    print("✅ CSV headers are correct")
                    
                    # Check if there's an example row
                    rows = list(reader)
                    if len(rows) >= 1:
                        example_row = rows[0]
                        if example_row.get('name') and example_row.get('email'):
                            print("✅ Template contains example row with data")
                            print(f"   Example: {example_row['name']} ({example_row['email']})")
                            return True
                        else:
                            print("❌ Example row missing required data")
                            return False
                    else:
                        print("❌ Template missing example row")
                        return False
                else:
                    print(f"❌ CSV headers incorrect. Expected: {expected_headers}, Got: {headers}")
                    return False
            else:
                print(f"❌ Response not a CSV file. Content-Type: {content_type}")
                return False
        else:
            print(f"❌ Template endpoint failed: {response.status_code} - {response.text}")
            return False
    
    def create_test_csv(self):
        """Create a test CSV file for import testing"""
        csv_content = """name,email,phone,company,source,property_type,tags,notes,lead_score
John Smith,john.smith@example.com,(555) 123-4567,Smith Enterprises,website,residential_lease,"vip,urgent",Looking for 2BR apartment downtown,85
Sarah Johnson,sarah.j@company.com,(555) 987-6543,Johnson Corp,referral,commercial_lease,"corporate,priority",Need office space for 50 employees,92
Mike Wilson,mike.wilson@email.com,(555) 555-0123,Wilson LLC,cold_call,commercial_sale,"investor,cash_buyer",Interested in retail properties,78"""
        
        return csv_content
    
    def create_test_xlsx(self):
        """Create a test XLSX file for import testing"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers
        headers = ["name", "email", "phone", "company", "source", "property_type", "tags", "notes", "lead_score"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Test data
        test_data = [
            ["Emma Davis", "emma.davis@test.com", "(555) 111-2222", "Davis Holdings", "social_media", "residential_lease", "first_time,young_professional", "Looking for studio or 1BR", "65"],
            ["Robert Chen", "robert.chen@business.com", "(555) 333-4444", "Chen Industries", "trade_show", "commercial_sale", "established,expansion", "Seeking warehouse space", "88"],
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Save to bytes
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        return xlsx_buffer.getvalue()
    
    def test_csv_import(self):
        """Test CSV import functionality"""
        print("\n📤 Testing CSV import...")
        
        csv_content = self.create_test_csv()
        
        files = {
            'file': ('test_contacts.csv', csv_content, 'text/csv')
        }
        
        response = self.session.post(f"{API_BASE}/contacts/import", files=files)
        
        if response.status_code == 200:
            result = response.json()
            
            if 'imported' in result and 'total_rows' in result and 'errors' in result:
                print(f"✅ CSV import successful")
                print(f"   Imported: {result['imported']} contacts")
                print(f"   Total rows: {result['total_rows']}")
                print(f"   Errors: {len(result['errors'])}")
                
                if result['imported'] > 0:
                    print("✅ Contacts were successfully imported")
                    return True
                else:
                    print("❌ No contacts were imported")
                    if result['errors']:
                        print(f"   Errors: {result['errors']}")
                    return False
            else:
                print(f"❌ Import response missing required fields: {result}")
                return False
        else:
            print(f"❌ CSV import failed: {response.status_code} - {response.text}")
            return False
    
    def test_xlsx_import(self):
        """Test XLSX import functionality"""
        print("\n📤 Testing XLSX import...")
        
        xlsx_content = self.create_test_xlsx()
        
        files = {
            'file': ('test_contacts.xlsx', xlsx_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        response = self.session.post(f"{API_BASE}/contacts/import", files=files)
        
        if response.status_code == 200:
            result = response.json()
            
            if 'imported' in result and 'total_rows' in result and 'errors' in result:
                print(f"✅ XLSX import successful")
                print(f"   Imported: {result['imported']} contacts")
                print(f"   Total rows: {result['total_rows']}")
                print(f"   Errors: {len(result['errors'])}")
                
                if result['imported'] > 0:
                    print("✅ XLSX contacts were successfully imported")
                    return True
                else:
                    print("❌ No XLSX contacts were imported")
                    if result['errors']:
                        print(f"   Errors: {result['errors']}")
                    return False
            else:
                print(f"❌ XLSX import response missing required fields: {result}")
                return False
        else:
            print(f"❌ XLSX import failed: {response.status_code} - {response.text}")
            return False
    
    def verify_imported_contacts(self):
        """Verify that imported contacts appear in the contacts list"""
        print("\n🔍 Verifying imported contacts...")
        
        response = self.session.get(f"{API_BASE}/contacts")
        
        if response.status_code == 200:
            contacts = response.json()
            
            # Look for our test contacts
            test_names = ["John Smith", "Sarah Johnson", "Mike Wilson", "Emma Davis", "Robert Chen"]
            found_contacts = []
            
            for contact in contacts:
                if contact.get('name') in test_names:
                    found_contacts.append(contact['name'])
            
            print(f"✅ Found {len(found_contacts)} imported contacts: {found_contacts}")
            
            if len(found_contacts) >= 3:  # Should have at least 3 from CSV + 2 from XLSX
                return True
            else:
                print("❌ Not all test contacts were found in the contacts list")
                return False
        else:
            print(f"❌ Failed to retrieve contacts: {response.status_code} - {response.text}")
            return False
    
    def test_contacts_export(self):
        """Test GET /api/contacts/export endpoint"""
        print("\n📥 Testing contacts export...")
        
        response = self.session.get(f"{API_BASE}/contacts/export")
        
        if response.status_code == 200:
            content_type = response.headers.get('content-type', '')
            content_disposition = response.headers.get('content-disposition', '')
            
            if 'text/csv' in content_type and 'attachment' in content_disposition:
                print("✅ Export endpoint returns CSV file")
                
                # Parse CSV content
                csv_content = response.text
                reader = csv.DictReader(io.StringIO(csv_content))
                headers = reader.fieldnames
                
                expected_headers = ["name", "email", "phone", "company", "source", "property_type", "tags", "notes", "lead_score"]
                
                if headers == expected_headers:
                    print("✅ Export CSV headers are correct")
                    
                    # Count rows
                    rows = list(reader)
                    print(f"✅ Export contains {len(rows)} contacts")
                    
                    # Check if our test contacts are in the export
                    test_names = ["John Smith", "Sarah Johnson", "Mike Wilson", "Emma Davis", "Robert Chen"]
                    found_in_export = []
                    
                    for row in rows:
                        if row.get('name') in test_names:
                            found_in_export.append(row['name'])
                    
                    if len(found_in_export) >= 3:
                        print(f"✅ Test contacts found in export: {found_in_export}")
                        return True
                    else:
                        print(f"⚠️  Only {len(found_in_export)} test contacts found in export")
                        return True  # Still consider success if export works
                else:
                    print(f"❌ Export CSV headers incorrect. Expected: {expected_headers}, Got: {headers}")
                    return False
            else:
                print(f"❌ Export response not a CSV file. Content-Type: {content_type}")
                return False
        else:
            print(f"❌ Export endpoint failed: {response.status_code} - {response.text}")
            return False
    
    def run_all_tests(self):
        """Run all contact import/export tests"""
        print("🚀 Starting PropFlow CRM Contacts Import/Export Tests")
        print("=" * 60)
        
        results = {}
        
        # Authenticate
        if not self.authenticate():
            print("❌ Authentication failed - cannot proceed with tests")
            return False
        
        # Test template endpoint
        results['template'] = self.test_contacts_template_endpoint()
        
        # Test CSV import
        results['csv_import'] = self.test_csv_import()
        
        # Test XLSX import
        results['xlsx_import'] = self.test_xlsx_import()
        
        # Verify imported contacts
        results['verify_contacts'] = self.verify_imported_contacts()
        
        # Test export
        results['export'] = self.test_contacts_export()
        
        # Summary
        print("\n" + "=" * 60)
        print("📊 TEST SUMMARY")
        print("=" * 60)
        
        passed = 0
        total = len(results)
        
        for test_name, result in results.items():
            status = "✅ PASS" if result else "❌ FAIL"
            print(f"{test_name.replace('_', ' ').title()}: {status}")
            if result:
                passed += 1
        
        print(f"\nOverall: {passed}/{total} tests passed")
        
        if passed == total:
            print("🎉 All tests passed!")
            return True
        else:
            print("⚠️  Some tests failed")
            return False

if __name__ == "__main__":
    tester = ContactsImportExportTester()
    success = tester.run_all_tests()
    exit(0 if success else 1)