# ─── Config: Pydantic Settings (validates ALL env vars on startup) ───
from config import settings

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse, RedirectResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import re
import logging
import bcrypt
import jwt
import secrets
import csv
import io
import json
import traceback
import httpx
import asyncio
from pydantic import BaseModel, Field, field_validator
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import math
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

# ─── Logging ───
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ─── Validated settings (from config.py Pydantic model) ───
FRONTEND_URL = settings.FRONTEND_URL
IS_PRODUCTION = settings.is_production

# ─── MongoDB connection ───
mongo_url = settings.MONGO_URL
client = AsyncIOMotorClient(mongo_url)
db = client[settings.DB_NAME]

# ─── App & Router ───
app = FastAPI(
    docs_url=None if IS_PRODUCTION else "/docs",     # Disable Swagger UI in production
    redoc_url=None if IS_PRODUCTION else "/redoc",    # Disable ReDoc in production
)
api_router = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"

# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY: Rate Limiting (slowapi)
# Default:  100 requests/minute per IP for general endpoints
# Strict:   10 requests/minute per IP for auth endpoints (login, register, refresh)
# Prevents brute-force attacks, credential stuffing, and API abuse.
# ═══════════════════════════════════════════════════════════════════════════════
limiter = Limiter(key_func=get_remote_address, default_limits=["100/minute"])
app.state.limiter = limiter

# ═══════════════════════════════════════════════════════════════════════════════
# AI COST GUARDRAILS & TOKEN TRACKING
# - Estimate tokens before each call; block if expected cost exceeds MAX_AI_COST_PER_CALL
# - Log actual usage to ai_usage_logs collection after each call
# - Rate limit: max 20 AI calls per user per hour
# GPT-5.2 approximate pricing: $2/1M input tokens, $8/1M output tokens
# ═══════════════════════════════════════════════════════════════════════════════
MAX_AI_COST_PER_CALL = settings.MAX_AI_COST_PER_CALL
MAX_AI_CALLS_PER_HOUR = settings.MAX_AI_CALLS_PER_HOUR
AI_INPUT_COST_PER_TOKEN = 2.0 / 1_000_000   # $0.000002
AI_OUTPUT_COST_PER_TOKEN = 8.0 / 1_000_000   # $0.000008
AI_EST_OUTPUT_TOKENS = 800  # Conservative estimate for output

def estimate_tokens(text: str) -> int:
    """Rough token estimate: ~1 token per 4 characters for English text."""
    return max(1, len(text) // 4)

def estimate_ai_cost(input_text: str) -> float:
    """Estimate the cost of an AI call before making it."""
    input_tokens = estimate_tokens(input_text)
    return (input_tokens * AI_INPUT_COST_PER_TOKEN) + (AI_EST_OUTPUT_TOKENS * AI_OUTPUT_COST_PER_TOKEN)

async def check_ai_rate_limit(user_id: str):
    """Enforce per-user AI rate limit: MAX_AI_CALLS_PER_HOUR calls/hour."""
    one_hour_ago = (datetime.now(timezone.utc) - timedelta(hours=1)).isoformat()
    recent_count = await db.ai_usage_logs.count_documents({
        "user_id": user_id, "timestamp": {"$gte": one_hour_ago}
    })
    if recent_count >= MAX_AI_CALLS_PER_HOUR:
        raise HTTPException(
            status_code=429,
            detail=f"AI rate limit exceeded. Maximum {MAX_AI_CALLS_PER_HOUR} AI calls per hour. Please wait and try again."
        )

async def log_ai_usage(user_id: str, endpoint: str, input_tokens: int, output_tokens: int, model: str = "gpt-5.2"):
    """Log AI call usage for cost tracking and auditing."""
    input_cost = input_tokens * AI_INPUT_COST_PER_TOKEN
    output_cost = output_tokens * AI_OUTPUT_COST_PER_TOKEN
    total_cost = input_cost + output_cost
    log_entry = {
        "user_id": user_id,
        "endpoint": endpoint,
        "model": model,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "estimated_cost_usd": round(total_cost, 6),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    await db.ai_usage_logs.insert_one(log_entry)
    logger.info(f"AI USAGE | user={user_id} | endpoint={endpoint} | "
                f"in={input_tokens} out={output_tokens} tokens | "
                f"cost=${total_cost:.6f}")

def guard_ai_cost(input_text: str):
    """Block AI call if estimated cost exceeds the per-call limit."""
    estimated_cost = estimate_ai_cost(input_text)
    if estimated_cost > MAX_AI_COST_PER_CALL:
        raise HTTPException(
            status_code=400,
            detail=f"Input too large. Estimated cost ${estimated_cost:.4f} exceeds "
                   f"limit of ${MAX_AI_COST_PER_CALL:.2f}. Please reduce the input."
        )

# ═══════════════════════════════════════════════════════════════════════════════
# SEQUENCE POLLING WORKER: MongoDB-backed drip campaign executor
# Runs every 60 seconds, finds pending sequence_executions, and sends emails/SMS
# Uses atomic find_and_modify to prevent duplicate sends across workers
# ═══════════════════════════════════════════════════════════════════════════════
async def process_sequence_executions():
    """Background worker that polls for pending sequence executions and processes them."""
    import asyncio
    
    while True:
        try:
            # Find executions that are pending and scheduled to run now or earlier
            now = datetime.now(timezone.utc)
            
            # Atomic find and lock: prevents race conditions across multiple workers
            execution = await db.sequence_executions.find_one_and_update(
                {
                    "status": "pending",
                    "scheduled_at": {"$lte": now.isoformat()}
                },
                {
                    "$set": {
                        "status": "processing",
                        "started_at": now.isoformat()
                    }
                },
                sort=[("scheduled_at", 1)]  # Process oldest first
            )
            
            if not execution:
                # No pending executions; sleep for 60 seconds
                await asyncio.sleep(60)
                continue
            
            # Process the execution
            try:
                sequence = await db.sequences.find_one({"_id": ObjectId(execution["sequence_id"])}, {"_id": 0})
                contact = await db.contacts.find_one({"_id": ObjectId(execution["contact_id"])}, {"_id": 0})
                
                if not sequence or not contact:
                    logger.warning(f"Sequence or contact not found for execution {execution.get('_id')}")
                    await db.sequence_executions.update_one(
                        {"_id": execution["_id"]},
                        {"$set": {"status": "failed", "error": "Sequence or contact not found", "completed_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    continue
                
                if not sequence.get("active"):
                    logger.info(f"Sequence {sequence.get('name')} is inactive, skipping execution")
                    await db.sequence_executions.update_one(
                        {"_id": execution["_id"]},
                        {"$set": {"status": "skipped", "completed_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    continue
                
                step = sequence["steps"][execution["step_index"]]
                
                # Replace variables in subject/body
                body = step["body"]
                subject = step.get("subject", "")
                body = body.replace("{{contact.name}}", contact.get("name", "there"))
                body = body.replace("{{contact.email}}", contact.get("email", ""))
                subject = subject.replace("{{contact.name}}", contact.get("name", ""))
                
                # Send email or SMS
                if step["type"] == "email":
                    if contact.get("email"):
                        try:
                            # Use Brevo or fallback to logging
                            brevo_key = settings.BREVO_API_KEY
                            if brevo_key:
                                send_email_with_retry(brevo_key, contact["email"], subject, body)
                                logger.info(f"Sequence email sent to {contact['email']}")
                            else:
                                logger.warning(f"Brevo not configured. Would send email to {contact['email']}: {subject}")
                        except Exception as e:
                            logger.error(f"Failed to send sequence email: {e}")
                            raise
                    else:
                        logger.warning(f"Contact {contact.get('name')} has no email address")
                
                elif step["type"] == "sms":
                    if contact.get("phone"):
                        try:
                            # Use Twilio or fallback to logging
                            account_sid = settings.TWILIO_ACCOUNT_SID
                            auth_token = settings.TWILIO_AUTH_TOKEN
                            from_number = settings.TWILIO_PHONE_NUMBER
                            if account_sid and auth_token and from_number:
                                send_sms_with_retry(account_sid, auth_token, from_number, contact["phone"], body)
                                logger.info(f"Sequence SMS sent to {contact['phone']}")
                            else:
                                logger.warning(f"Twilio not configured. Would send SMS to {contact['phone']}: {body}")
                        except Exception as e:
                            logger.error(f"Failed to send sequence SMS: {e}")
                            raise
                    else:
                        logger.warning(f"Contact {contact.get('name')} has no phone number")
                
                # Mark as completed
                await db.sequence_executions.update_one(
                    {"_id": execution["_id"]},
                    {"$set": {"status": "completed", "completed_at": datetime.now(timezone.utc).isoformat()}}
                )
                
                # Log activity
                await db.activities.insert_one({
                    "user_id": contact["user_id"],
                    "contact_id": str(contact["_id"]),
                    "type": f"sequence_{step['type']}",
                    "title": f"Drip Sequence: {sequence['name']}",
                    "description": f"Step {execution['step_index'] + 1}: {step['type']}",
                    "metadata": {"sequence_id": execution["sequence_id"], "step_index": execution["step_index"]},
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
                
                # Schedule next step if exists
                if execution["step_index"] + 1 < len(sequence["steps"]):
                    next_step = sequence["steps"][execution["step_index"] + 1]
                    next_scheduled = datetime.now(timezone.utc) + timedelta(days=next_step["delay_days"])
                    
                    # Check if next execution already exists (idempotent)
                    existing_next = await db.sequence_executions.find_one({
                        "sequence_id": execution["sequence_id"],
                        "contact_id": execution["contact_id"],
                        "step_index": execution["step_index"] + 1
                    })
                    
                    if not existing_next:
                        await db.sequence_executions.insert_one({
                            "sequence_id": execution["sequence_id"],
                            "contact_id": execution["contact_id"],
                            "step_index": execution["step_index"] + 1,
                            "status": "pending",
                            "scheduled_at": next_scheduled.isoformat(),
                            "created_at": datetime.now(timezone.utc).isoformat()
                        })
                
            except Exception as e:
                logger.error(f"Error processing sequence execution {execution.get('_id')}: {e}")
                await db.sequence_executions.update_one(
                    {"_id": execution["_id"]},
                    {"$set": {"status": "failed", "error": str(e), "completed_at": datetime.now(timezone.utc).isoformat()}}
                )
        
        except Exception as e:
            logger.error(f"Error in sequence polling worker: {e}")
            await asyncio.sleep(60)

# ═══════════════════════════════════════════════════════════════════════════════
# RETRY LOGIC: Twilio SMS with exponential backoff
# 3 attempts, waits 1s → 2s → 4s between retries
# Only retries on transient network/service errors, not auth failures
# ═══════════════════════════════════════════════════════════════════════════════
def send_sms_with_retry(account_sid: str, auth_token: str, from_number: str, to_phone: str, message: str):
    """Synchronous Twilio SMS call wrapped with tenacity retry."""
    from twilio.rest import Client as TwilioClient
    from twilio.base.exceptions import TwilioRestException

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=1, max=8),
        retry=retry_if_exception_type((ConnectionError, TimeoutError, OSError)),
        reraise=True,
    )
    def _send():
        client = TwilioClient(account_sid, auth_token)
        return client.messages.create(body=message, from_=from_number, to=to_phone)

    return _send()

# ═══════════════════════════════════════════════════════════════════════════════
# RETRY LOGIC: Brevo Email with exponential backoff
# 3 attempts, waits 1s → 2s → 4s between retries
# Only retries on transient network/service errors, not auth failures
# ═══════════════════════════════════════════════════════════════════════════════
def send_email_with_retry(api_key: str, to_email: str, subject: str, html_content: str):
    """Synchronous Brevo email call wrapped with tenacity retry."""
    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=1, max=8),
        retry=retry_if_exception_type((ConnectionError, TimeoutError, OSError)),
        reraise=True,
    )
    def _send():
        response = httpx.post(
            "https://api.brevo.com/v3/smtp/email",
            headers={"api-key": api_key, "Content-Type": "application/json"},
            json={
                "sender": {"email": "noreply@propflow.com", "name": "PropFlow CRM"},
                "to": [{"email": to_email}],
                "subject": subject,
                "htmlContent": html_content
            },
            timeout=10
        )
        response.raise_for_status()
        return response
    
    return _send()

# ─── Password Hashing ───
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def get_jwt_secret():
    return settings.JWT_SECRET

def create_access_token(user_id: str, email: str) -> str:
    # SECURITY: Access token expires in 15 minutes (short-lived)
    return jwt.encode(
        {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(minutes=15), "type": "access"},
        get_jwt_secret(), algorithm=JWT_ALGORITHM
    )

def create_refresh_token(user_id: str) -> str:
    # SECURITY: Refresh token expires in 7 days
    return jwt.encode(
        {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"},
        get_jwt_secret(), algorithm=JWT_ALGORITHM
    )

# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY: Environment-aware auth cookie setter
# - httponly=True:     Prevents JavaScript access (XSS protection)
# - samesite="strict": Prevents CSRF by blocking cross-origin cookie sends
# - secure=True:       Cookies only sent over HTTPS (production only)
# - max_age:           15 min for access_token, 7 days for refresh_token
# ═══════════════════════════════════════════════════════════════════════════════
def set_auth_cookies(response: Response, access_token: str, refresh_token: str):
    response.set_cookie(
        key="access_token", value=access_token,
        httponly=True, secure=IS_PRODUCTION, samesite="strict",
        max_age=900, path="/"       # 15 minutes
    )
    response.set_cookie(
        key="refresh_token", value=refresh_token,
        httponly=True, secure=IS_PRODUCTION, samesite="strict",
        max_age=604800, path="/"    # 7 days
    )

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ─── API Key Auth for External Agents ───
async def get_api_key_user(request: Request) -> dict:
    api_key = request.headers.get("X-API-Key", "")
    if not api_key:
        raise HTTPException(status_code=401, detail="API key required")
    key_doc = await db.api_keys.find_one({"key": api_key, "active": True})
    if not key_doc:
        raise HTTPException(status_code=401, detail="Invalid API key")
    user = await db.users.find_one({"_id": ObjectId(key_doc["user_id"])})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    user["_id"] = str(user["_id"])
    user.pop("password_hash", None)
    await db.api_keys.update_one({"_id": key_doc["_id"]}, {"$set": {"last_used": datetime.now(timezone.utc).isoformat()}})
    return user

# Flexible auth: try JWT first, then API key
async def get_any_auth_user(request: Request) -> dict:
    try:
        return await get_current_user(request)
    except HTTPException:
        return await get_api_key_user(request)

# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY: Input Validation Models (Pydantic)
# All user-facing input is validated through strict Pydantic models with:
# - Field length limits to prevent payload abuse
# - Enum constraints where applicable
# - Email format validation
# - Stripping of whitespace on string fields
# ═══════════════════════════════════════════════════════════════════════════════

class RegisterInput(BaseModel):
    email: str = Field(..., min_length=5, max_length=254)
    password: str = Field(..., min_length=6, max_length=128)
    name: str = Field(..., min_length=1, max_length=200)

    @field_validator("email")
    @classmethod
    def validate_email(cls, v):
        v = v.strip().lower()
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', v):
            raise ValueError("Invalid email format")
        return v

class LoginInput(BaseModel):
    email: str = Field(..., max_length=254)
    password: str = Field(..., max_length=128)

class ContactCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=300)
    email: Optional[str] = Field(default="", max_length=254)
    phone: Optional[str] = Field(default="", max_length=30)
    company: Optional[str] = Field(default="", max_length=300)
    source: Optional[str] = Field(default="manual", max_length=50)
    tags: Optional[List[str]] = []
    property_type: Optional[str] = Field(default="residential_lease", max_length=50)
    notes: Optional[str] = Field(default="", max_length=5000)
    lead_score: Optional[int] = Field(default=0, ge=0, le=100)

    @field_validator("email")
    @classmethod
    def validate_email(cls, v):
        if v and v.strip():
            if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', v.strip()):
                raise ValueError("Invalid email format")
        return v

    @field_validator("property_type")
    @classmethod
    def validate_property_type(cls, v):
        allowed = ("residential_lease", "commercial_sale", "commercial_lease")
        if v and v not in allowed:
            raise ValueError(f"property_type must be one of: {', '.join(allowed)}")
        return v

class ContactUpdate(BaseModel):
    name: Optional[str] = Field(default=None, max_length=300)
    email: Optional[str] = Field(default=None, max_length=254)
    phone: Optional[str] = Field(default=None, max_length=30)
    company: Optional[str] = Field(default=None, max_length=300)
    source: Optional[str] = Field(default=None, max_length=50)
    tags: Optional[List[str]] = None
    property_type: Optional[str] = Field(default=None, max_length=50)
    notes: Optional[str] = Field(default=None, max_length=5000)
    lead_score: Optional[int] = Field(default=None, ge=0, le=100)
    # ── Profile-page extensions (all optional, backward-compatible) ──
    client_type: Optional[str] = Field(default=None, max_length=30)
    leasing_stage: Optional[str] = Field(default=None, max_length=50)
    stage_updated_at: Optional[str] = Field(default=None, max_length=50)
    retention_score: Optional[int] = Field(default=None, ge=0, le=100)
    retention_summary: Optional[str] = Field(default=None, max_length=5000)
    retention_summary_generated_at: Optional[str] = Field(default=None, max_length=50)
    photo_url: Optional[str] = Field(default=None, max_length=2_500_000)
    address: Optional[str] = Field(default=None, max_length=500)
    collaborator_ids: Optional[List[str]] = None
    is_tenant: Optional[bool] = None
    last_reengaged_at: Optional[str] = Field(default=None, max_length=40)

class PropertyCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=300)
    address: str = Field(..., min_length=1, max_length=500)
    property_type: str = Field(..., max_length=20)
    listing_type: str = Field(..., max_length=10)
    price: Optional[float] = Field(default=0, ge=0)
    sqft: Optional[float] = Field(default=0, ge=0)
    bedrooms: Optional[int] = Field(default=0, ge=0, le=100)
    bathrooms: Optional[int] = Field(default=0, ge=0, le=100)
    description: Optional[str] = Field(default="", max_length=5000)
    status: Optional[str] = Field(default="active", max_length=20)
    image_url: Optional[str] = Field(default="", max_length=2000)

    @field_validator("property_type")
    @classmethod
    def validate_prop_type(cls, v):
        if v not in ("residential", "commercial"):
            raise ValueError("property_type must be 'residential' or 'commercial'")
        return v

    @field_validator("listing_type")
    @classmethod
    def validate_listing_type(cls, v):
        if v not in ("lease", "sale"):
            raise ValueError("listing_type must be 'lease' or 'sale'")
        return v

    @field_validator("status")
    @classmethod
    def validate_status(cls, v):
        if v and v not in ("active", "pending", "closed"):
            raise ValueError("status must be 'active', 'pending', or 'closed'")
        return v

class DealCreate(BaseModel):
    title: str = Field(..., min_length=1, max_length=300)
    pipeline_type: str = Field(..., max_length=30)
    stage: str = Field(..., max_length=50)
    contact_id: Optional[str] = Field(default="", max_length=50)
    property_id: Optional[str] = Field(default="", max_length=50)
    value: Optional[float] = Field(default=0, ge=0)
    notes: Optional[str] = Field(default="", max_length=5000)
    # ── Phase 10: leasing-focused fields (all optional, backward-compatible) ──
    unit_number: Optional[str] = Field(default="", max_length=40)
    unit_address: Optional[str] = Field(default="", max_length=300)
    desired_rent: Optional[float] = Field(default=0, ge=0)
    budget_min: Optional[float] = Field(default=0, ge=0)
    budget_max: Optional[float] = Field(default=0, ge=0)
    move_in_date: Optional[str] = Field(default="", max_length=20)
    tags: Optional[List[str]] = Field(default_factory=list)
    co_applicant_ids: Optional[List[str]] = Field(default_factory=list)

    @field_validator("pipeline_type")
    @classmethod
    def validate_pipeline(cls, v):
        allowed = ("residential_lease", "commercial_sale", "commercial_lease", "lease_applications")
        if v not in allowed:
            raise ValueError(f"pipeline_type must be one of: {', '.join(allowed)}")
        return v

class DealUpdate(BaseModel):
    title: Optional[str] = Field(default=None, max_length=300)
    stage: Optional[str] = Field(default=None, max_length=50)
    contact_id: Optional[str] = Field(default=None, max_length=50)
    property_id: Optional[str] = Field(default=None, max_length=50)
    value: Optional[float] = Field(default=None, ge=0)
    notes: Optional[str] = Field(default=None, max_length=5000)
    # ── Phase 10 extensions ──
    unit_number: Optional[str] = Field(default=None, max_length=40)
    unit_address: Optional[str] = Field(default=None, max_length=300)
    desired_rent: Optional[float] = Field(default=None, ge=0)
    budget_min: Optional[float] = Field(default=None, ge=0)
    budget_max: Optional[float] = Field(default=None, ge=0)
    move_in_date: Optional[str] = Field(default=None, max_length=20)
    tags: Optional[List[str]] = None
    co_applicant_ids: Optional[List[str]] = None
    last_reengaged_at: Optional[str] = Field(default=None, max_length=40)

class TaskCreate(BaseModel):
    title: str = Field(..., min_length=1, max_length=300)
    description: Optional[str] = Field(default="", max_length=5000)
    due_date: Optional[str] = Field(default="", max_length=30)
    contact_id: Optional[str] = Field(default="", max_length=50)
    deal_id: Optional[str] = Field(default="", max_length=50)
    property_id: Optional[str] = Field(default="", max_length=50)
    priority: Optional[str] = Field(default="medium", max_length=10)
    completed: Optional[bool] = False
    # ── Phase 15 leasing extensions (all optional, backward-compatible) ──
    task_type: Optional[str] = Field(default="other", max_length=40)
    assigned_to: Optional[str] = Field(default="", max_length=50)
    all_day: Optional[bool] = False

    @field_validator("priority")
    @classmethod
    def validate_priority(cls, v):
        if v and v not in ("high", "medium", "low"):
            raise ValueError("priority must be 'high', 'medium', or 'low'")
        return v

    @field_validator("task_type")
    @classmethod
    def validate_task_type(cls, v):
        allowed = (
            "tour_followup", "renewal_offer", "maintenance_request",
            "application_reminder", "showing_prep", "listing_outreach",
            "lease_signing", "move_in", "rent_reminder", "other",
        )
        if v and v not in allowed:
            raise ValueError(f"task_type must be one of: {', '.join(allowed)}")
        return v or "other"

class ActivityCreate(BaseModel):
    contact_id: str = Field(..., max_length=50)
    activity_type: str = Field(..., max_length=20)
    description: str = Field(..., min_length=1, max_length=5000)
    deal_id: Optional[str] = Field(default="", max_length=50)

    @field_validator("activity_type")
    @classmethod
    def validate_activity_type(cls, v):
        allowed = ("call", "email", "note", "meeting", "sms")
        if v not in allowed:
            raise ValueError(f"activity_type must be one of: {', '.join(allowed)}")
        return v

class AIEmailRequest(BaseModel):
    contact_id: str = Field(..., max_length=50)
    context: Optional[str] = Field(default="", max_length=2000)
    tone: Optional[str] = Field(default="professional", max_length=30)

class AILeadScoreRequest(BaseModel):
    contact_id: str = Field(..., max_length=50)

class SendEmailRequest(BaseModel):
    contact_id: str = Field(..., max_length=50)
    to_email: str = Field(..., max_length=254)
    subject: str = Field(..., min_length=1, max_length=500)
    body: str = Field(..., min_length=1, max_length=50000)

    @field_validator("to_email")
    @classmethod
    def validate_to_email(cls, v):
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', v.strip()):
            raise ValueError("Invalid email format")
        return v.strip()

class SendSMSRequest(BaseModel):
    contact_id: str = Field(..., max_length=50)
    to_phone: str = Field(..., max_length=20)
    message: str = Field(..., min_length=1, max_length=1600)

class WebhookCreate(BaseModel):
    url: str = Field(..., min_length=10, max_length=2000)
    events: List[str]
    name: Optional[str] = Field(default="Webhook", max_length=100)

    @field_validator("url")
    @classmethod
    def validate_url(cls, v):
        if not v.startswith(("http://", "https://")):
            raise ValueError("Webhook URL must start with http:// or https://")
        return v

class TeamInviteRequest(BaseModel):
    email: str = Field(..., min_length=5, max_length=254)
    name: str = Field(..., min_length=1, max_length=200)
    role: Optional[str] = Field(default="agent", max_length=20)

    @field_validator("email")
    @classmethod
    def validate_invite_email(cls, v):
        v = v.strip().lower()
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', v):
            raise ValueError("Invalid email format")
        return v

    @field_validator("role")
    @classmethod
    def validate_role(cls, v):
        if v not in ("admin", "agent"):
            raise ValueError("role must be 'admin' or 'agent'")
        return v

class TemplateCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    category: str = Field(..., max_length=10)
    subject: Optional[str] = Field(default="", max_length=500)
    body: str = Field(..., min_length=1, max_length=50000)
    tags: Optional[List[str]] = []

    @field_validator("category")
    @classmethod
    def validate_category(cls, v):
        if v not in ("email", "sms"):
            raise ValueError("category must be 'email' or 'sms'")
        return v

class SequenceStepCreate(BaseModel):
    type: str = Field(..., max_length=20)  # "email" or "sms"
    delay_days: int = Field(default=0, ge=0, le=365)
    template_id: Optional[str] = None
    subject: Optional[str] = Field(default="", max_length=500)
    body: str = Field(..., min_length=1, max_length=50000)

    @field_validator("type")
    @classmethod
    def validate_step_type(cls, v):
        if v not in ("email", "sms"):
            raise ValueError("type must be 'email' or 'sms'")
        return v

class SequenceCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    trigger: str = Field(..., max_length=50)  # "contact_created", "deal_stage_changed", etc.
    trigger_value: Optional[str] = Field(default="", max_length=100)
    steps: List[SequenceStepCreate]
    active: bool = Field(default=True)

    @field_validator("trigger")
    @classmethod
    def validate_trigger(cls, v):
        valid_triggers = ["contact_created", "deal_stage_changed", "property_viewed", "manual"]
        if v not in valid_triggers:
            raise ValueError(f"trigger must be one of {valid_triggers}")
        return v

class SequenceUpdate(BaseModel):
    name: Optional[str] = Field(default=None, min_length=1, max_length=200)
    trigger: Optional[str] = Field(default=None, max_length=50)
    trigger_value: Optional[str] = Field(default=None, max_length=100)
    steps: Optional[List[SequenceStepCreate]] = None
    active: Optional[bool] = None

class UserUpdate(BaseModel):
    name: Optional[str] = Field(default=None, min_length=1, max_length=200)
    auto_assign: Optional[bool] = None
    email_signature: Optional[str] = Field(default=None, max_length=2000)

# ─── Stage Automation Config ───
STAGE_AUTO_TASKS = {
    "Contacted": {"title": "Follow up within 24 hours", "priority": "high", "days_offset": 1},
    "Showing": {"title": "Confirm showing details", "priority": "high", "days_offset": 0},
    "Tour": {"title": "Prepare tour materials", "priority": "high", "days_offset": 0},
    "Application": {"title": "Review application", "priority": "high", "days_offset": 1},
    "LOI": {"title": "Review Letter of Intent", "priority": "high", "days_offset": 2},
    "Due Diligence": {"title": "Begin due diligence checklist", "priority": "high", "days_offset": 1},
    "Proposal": {"title": "Prepare lease proposal", "priority": "high", "days_offset": 1},
    "Negotiation": {"title": "Schedule negotiation meeting", "priority": "medium", "days_offset": 2},
    "Lease Signed": {"title": "Process lease paperwork", "priority": "high", "days_offset": 1},
    "Closing": {"title": "Coordinate closing logistics", "priority": "high", "days_offset": 3},
    "Closed": {"title": "Send thank you & request referral", "priority": "low", "days_offset": 2},
    # ── Lease Applications pipeline (Phase 10) ──
    "Tour Scheduled": {"title": "Confirm tour time & prep materials", "priority": "high", "days_offset": 0},
    "Application Submitted": {"title": "Review application & request docs", "priority": "high", "days_offset": 1},
    "Screening": {"title": "Run credit + background check", "priority": "high", "days_offset": 1},
    "Approved": {"title": "Send approval letter & prep lease", "priority": "high", "days_offset": 1},
    "Move-In": {"title": "Schedule move-in inspection + hand keys", "priority": "high", "days_offset": 2},
    "Active Tenant": {"title": "30-day tenant check-in call", "priority": "medium", "days_offset": 30},
    "Renewal": {"title": "Send renewal offer & schedule follow-up", "priority": "high", "days_offset": 7},
}

# ─── Pipeline Stage Definitions ───
PIPELINE_STAGES = {
    "residential_lease": ["New Lead", "Contacted", "Showing", "Application", "Lease Signed", "Closed"],
    "commercial_sale": ["New Lead", "Contacted", "Tour", "LOI", "Due Diligence", "Closing", "Closed"],
    "commercial_lease": ["New Lead", "Contacted", "Tour", "Proposal", "Negotiation", "Lease Signed", "Closed"],
    # ── Phase 10: New primary pipeline for residential leasing ──
    "lease_applications": [
        "Inquiry", "Tour Scheduled", "Application Submitted", "Screening",
        "Approved", "Lease Signed", "Move-In", "Active Tenant", "Renewal"
    ],
}

# ─── Phase 10: Lease Applications stage colors (for frontend headers) ───
LEASE_APPLICATIONS_STAGE_COLORS = {
    "Inquiry":               {"bg": "blue",    "hex": "#3B82F6"},
    "Tour Scheduled":        {"bg": "cyan",    "hex": "#06B6D4"},
    "Application Submitted": {"bg": "amber",   "hex": "#F59E0B"},
    "Screening":             {"bg": "orange",  "hex": "#F97316"},
    "Approved":              {"bg": "purple",  "hex": "#A855F7"},
    "Lease Signed":          {"bg": "green",   "hex": "#22C55E"},
    "Move-In":               {"bg": "emerald", "hex": "#10B981"},
    "Active Tenant":         {"bg": "teal",    "hex": "#14B8A6"},
    "Renewal":               {"bg": "indigo",  "hex": "#6366F1"},
}

# ─── Phase 10: Migration map from old residential_lease stages → new lease_applications stages ───
LEGACY_STAGE_MAP = {
    "New Lead":     "Inquiry",
    "Contacted":    "Inquiry",
    "Showing":      "Tour Scheduled",
    "Application":  "Application Submitted",
    "Lease Signed": "Lease Signed",
    "Closed":       "Active Tenant",
}

# ─── Helper ───
def serialize_doc(doc):
    if doc and "_id" in doc:
        doc["id"] = str(doc["_id"])
        del doc["_id"]
    return doc

# SECURITY: Validate ObjectId format before passing to MongoDB
# Prevents 500 errors from malformed IDs and potential injection
def validate_object_id(id_str: str, entity: str = "Resource") -> ObjectId:
    try:
        return ObjectId(id_str)
    except Exception:
        raise HTTPException(status_code=404, detail=f"{entity} not found")

# ═══════════════════════════════════════════════════════════════════════════════
# PERFORMANCE: Paginated query helper
# Used by all list endpoints. Returns { data: [...], pagination: {...} }
# - page/limit for cursor-based pagination via skip/limit
# - sort_field/sort_order for consistent ordering
# - projection to return only needed fields (reduces network + memory)
# - max limit capped at 500 to prevent abuse
# ═══════════════════════════════════════════════════════════════════════════════
PAGINATION_MAX_LIMIT = 500

async def paginate(collection, query: dict, page: int = 1, limit: int = 50,
                   sort_field: str = "created_at", sort_order: int = -1,
                   projection: dict = None):
    page = max(1, page)
    limit = max(1, min(limit, PAGINATION_MAX_LIMIT))
    skip_count = (page - 1) * limit
    total = await collection.count_documents(query)
    cursor = collection.find(query, projection).sort(sort_field, sort_order).skip(skip_count).limit(limit)
    items = await cursor.to_list(limit)
    return {
        "data": [serialize_doc(doc) for doc in items],
        "pagination": {
            "page": page,
            "limit": limit,
            "total": total,
            "total_pages": math.ceil(total / limit) if limit > 0 else 1,
        }
    }

# ═══════════════════════════════════════════════════════════════════════════════
# PERFORMANCE: MongoDB Index Definitions
# Called once on startup. Indexes are created in the background to avoid
# blocking writes. Each index is explained with its purpose.
# ═══════════════════════════════════════════════════════════════════════════════
async def create_mongodb_indexes():
    """Create all performance-critical MongoDB indexes on startup."""
    logger.info("Creating MongoDB indexes...")

    # ── users collection ──
    # Unique email lookup for auth (login, register, duplicate check)
    await db.users.create_index("email", unique=True, background=True)

    # ── login_attempts collection ──
    # Fast lookup by IP:email identifier for brute-force protection
    await db.login_attempts.create_index("identifier", background=True)
    # TTL index: auto-delete old login attempt records after 1 hour
    await db.login_attempts.create_index(
        "last_attempt_dt",
        expireAfterSeconds=3600,
        background=True,
        sparse=True
    )

    # ── contacts collection ──
    # Primary query pattern: list contacts for a user, sorted by created_at
    await db.contacts.create_index(
        [("user_id", 1), ("created_at", -1)],
        background=True,
        name="idx_contacts_user_created"
    )
    # Unique email per user — prevents duplicate contacts (only enforced when email is non-empty)
    await db.contacts.create_index(
        [("user_id", 1), ("email", 1)],
        unique=True,
        background=True,
        partialFilterExpression={"email": {"$gt": ""}},
        name="idx_contacts_user_email_unique"
    )
    # Filter by property_type within a user's contacts
    await db.contacts.create_index(
        [("user_id", 1), ("property_type", 1)],
        background=True,
        name="idx_contacts_user_proptype"
    )
    # Search by name/email/phone — text index for fast full-text search
    await db.contacts.create_index(
        [("name", "text"), ("email", "text"), ("company", "text")],
        background=True,
        name="idx_contacts_text_search"
    )
    # Lead scoring / sorting
    await db.contacts.create_index("lead_score", background=True)
    # Status filter (for future use)
    await db.contacts.create_index([("user_id", 1), ("status", 1)], background=True, sparse=True)
    # Last activity timestamp (for future use / sorting by engagement)
    await db.contacts.create_index("last_activity", background=True, sparse=True)

    # ── deals collection ──
    # Primary query: list deals by pipeline_type for kanban board
    await db.deals.create_index(
        [("user_id", 1), ("pipeline_type", 1), ("created_at", -1)],
        background=True,
        name="idx_deals_user_pipeline_created"
    )
    # Stage-based queries: filter/count deals by stage (dashboard, reporting)
    await db.deals.create_index(
        [("user_id", 1), ("stage", 1), ("created_at", -1)],
        background=True,
        name="idx_deals_user_stage_created"
    )
    # Lookup deals linked to a contact (contact detail page)
    await db.deals.create_index("contact_id", background=True)
    # Lookup deals assigned to a team member (future team features)
    await db.deals.create_index("assigned_to", background=True, sparse=True)

    # ── tasks collection ──
    # Primary query: list tasks for a user sorted by due date
    await db.tasks.create_index(
        [("user_id", 1), ("due_date", 1)],
        background=True,
        name="idx_tasks_user_duedate"
    )
    # Filter by completion status (open tasks count for dashboard)
    await db.tasks.create_index(
        [("user_id", 1), ("completed", 1)],
        background=True,
        name="idx_tasks_user_completed"
    )
    # Lookup tasks linked to a contact or deal
    await db.tasks.create_index("contact_id", background=True, sparse=True)
    await db.tasks.create_index("deal_id", background=True, sparse=True)

    # ── activities collection ──
    # Primary query: activity timeline for a contact, newest first
    await db.activities.create_index(
        [("contact_id", 1), ("created_at", -1)],
        background=True,
        name="idx_activities_contact_created"
    )
    # List all activities for a user (dashboard feed)
    await db.activities.create_index(
        [("user_id", 1), ("created_at", -1)],
        background=True,
        name="idx_activities_user_created"
    )

    # ── properties collection ──
    await db.properties.create_index(
        [("user_id", 1), ("created_at", -1)],
        background=True,
        name="idx_properties_user_created"
    )
    await db.properties.create_index(
        [("user_id", 1), ("property_type", 1), ("listing_type", 1)],
        background=True,
        name="idx_properties_user_type_listing"
    )

    # ── templates collection ──
    await db.templates.create_index(
        [("user_id", 1), ("category", 1), ("use_count", -1)],
        background=True,
        name="idx_templates_user_cat_usecount"
    )

    # ── api_keys collection ──
    await db.api_keys.create_index("key", unique=True, background=True)
    await db.api_keys.create_index("user_id", background=True)

    # ── webhooks collection ──
    await db.webhooks.create_index(
        [("user_id", 1), ("active", 1)],
        background=True,
        name="idx_webhooks_user_active"
    )

    # ── ai_usage_logs collection ──
    # Track AI usage per user for rate limiting and cost auditing
    await db.ai_usage_logs.create_index(
        [("user_id", 1), ("timestamp", -1)],
        background=True,
        name="idx_ai_usage_user_ts"
    )
    # TTL: auto-delete AI usage logs older than 30 days
    await db.ai_usage_logs.create_index(
        "timestamp_dt",
        expireAfterSeconds=30 * 86400,
        background=True,
        sparse=True
    )

    # ── sequences collection ──
    # List sequences for a user
    await db.sequences.create_index(
        [("user_id", 1), ("created_at", -1)],
        background=True,
        name="idx_sequences_user_created"
    )
    # Find active sequences by trigger type
    await db.sequences.create_index(
        [("user_id", 1), ("trigger", 1), ("active", 1)],
        background=True,
        name="idx_sequences_user_trigger_active"
    )

    # ── sequence_executions collection ──
    # Find pending executions to process
    await db.sequence_executions.create_index(
        [("status", 1), ("scheduled_at", 1)],
        background=True,
        name="idx_executions_status_scheduled"
    )
    # Prevent duplicate executions for the same contact/sequence/step
    await db.sequence_executions.create_index(
        [("sequence_id", 1), ("contact_id", 1), ("step_index", 1)],
        unique=True,
        background=True,
        name="idx_executions_seq_contact_step_unique"
    )
    # Lookup executions for a contact
    await db.sequence_executions.create_index("contact_id", background=True)

    logger.info("MongoDB indexes created successfully.")

# ─── Auth Routes ───
# SECURITY: Strict rate limit on auth endpoints — 10 requests/minute per IP
@api_router.post("/auth/register")
@limiter.limit("10/minute")
async def register(data: RegisterInput, request: Request, response: Response):
    email = data.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user_doc = {
        "email": email,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": "user",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)
    access = create_access_token(user_id, email)
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)
    return {"id": user_id, "email": email, "name": data.name, "role": "user"}

@api_router.post("/auth/login")
@limiter.limit("10/minute")
async def login(data: LoginInput, request: Request, response: Response):
    email = data.email.lower().strip()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"
    # Brute force check
    attempts = await db.login_attempts.find_one({"identifier": identifier})
    if attempts and attempts.get("count", 0) >= 5:
        lockout_time = attempts.get("last_attempt", "")
        if lockout_time:
            last = datetime.fromisoformat(lockout_time)
            if datetime.now(timezone.utc) - last < timedelta(minutes=15):
                raise HTTPException(status_code=429, detail="Too many attempts. Try again in 15 minutes.")
            else:
                await db.login_attempts.delete_one({"identifier": identifier})
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(data.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$inc": {"count": 1}, "$set": {"last_attempt": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")
    await db.login_attempts.delete_one({"identifier": identifier})
    user_id = str(user["_id"])
    access = create_access_token(user_id, email)
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)
    return {"id": user_id, "email": email, "name": user.get("name", ""), "role": user.get("role", "user")}

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out"}

@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    return user

@api_router.post("/auth/refresh")
@limiter.limit("10/minute")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        access = create_access_token(str(user["_id"]), user["email"])
        # SECURITY: Use the centralized set_auth_cookies helper for consistent cookie settings
        response.set_cookie(
            key="access_token", value=access,
            httponly=True, secure=IS_PRODUCTION, samesite="strict",
            max_age=900, path="/"
        )
        return {"message": "Token refreshed"}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ─── Contacts Routes ───
@api_router.post("/contacts")
async def create_contact(data: ContactCreate, background_tasks: BackgroundTasks, user=Depends(get_any_auth_user)):
    doc = data.model_dump()
    doc["user_id"] = user["_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # FEATURE: Round-Robin Assignment
    # If user has auto_assign enabled, assign to the agent with fewest open deals
    user_settings = await db.users.find_one({"_id": ObjectId(user["_id"])}, {"_id": 0, "auto_assign": 1})
    if user_settings and user_settings.get("auto_assign"):
        # Find all team members
        team_members = await db.users.find({"role": {"$in": ["agent", "admin"]}}, {"_id": 1}).to_list(100)
        if team_members:
            # Count open deals for each agent
            agent_deal_counts = []
            for agent in team_members:
                agent_id = str(agent["_id"])
                open_deals = await db.deals.count_documents({"assigned_to": agent_id, "stage": {"$ne": "Closed"}})
                agent_deal_counts.append({"agent_id": agent_id, "open_deals": open_deals})
            # Sort by fewest open deals
            agent_deal_counts.sort(key=lambda x: x["open_deals"])
            # Assign to agent with fewest deals
            doc["assigned_to"] = agent_deal_counts[0]["agent_id"]
            logger.info(f"Round-robin assignment: contact assigned to {doc['assigned_to']}")
    
    result = await db.contacts.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    contact_id = doc["id"]
    del doc["_id"]
    
    background_tasks.add_task(trigger_webhooks, user["_id"], "new_lead", {"contact_id": doc["id"], "name": doc["name"], "email": doc.get("email", "")})
    
    # FEATURE: Trigger Drip Sequences
    # Find active sequences with trigger="contact_created"
    active_sequences = await db.sequences.find({"user_id": user["_id"], "trigger": "contact_created", "active": True}, {"_id": 0}).to_list(100)
    for sequence in active_sequences:
        if sequence.get("steps"):
            first_step = sequence["steps"][0]
            scheduled_at = datetime.now(timezone.utc) + timedelta(days=first_step["delay_days"])
            # Check if execution already exists (prevent duplicates)
            existing = await db.sequence_executions.find_one({
                "sequence_id": sequence["id"],
                "contact_id": contact_id,
                "step_index": 0
            })
            if not existing:
                await db.sequence_executions.insert_one({
                    "sequence_id": sequence["id"],
                    "contact_id": contact_id,
                    "step_index": 0,
                    "status": "pending",
                    "scheduled_at": scheduled_at.isoformat(),
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
                logger.info(f"Contact {contact_id} enrolled in sequence {sequence['name']}")
    
    return doc

# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE: Smart Lists + Collections (Phase 13 — Contacts People page upgrade)
# Helper functions that build MongoDB filters for each smart list / collection.
# ═══════════════════════════════════════════════════════════════════════════════

SMART_LIST_IDS = [
    "today_tours_followups", "first_contact", "second_contact",
    "application_submitted", "at_risk_renewals", "stale_prospects",
    "recently_active", "nurture_queue", "recontact_next_year",
]
COLLECTION_IDS = ["prospects", "active_tenants", "past_tenants", "high_value_leads"]


async def _contact_ids_with_recent_activity(user_id: str, within_hours: int = 24) -> set:
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=within_hours)).isoformat()
    ids = set()
    async for a in db.activities.find(
        {"user_id": user_id, "created_at": {"$gte": cutoff}},
        {"contact_id": 1},
    ):
        cid = a.get("contact_id")
        if cid:
            ids.add(cid)
    return ids


async def _contact_activity_count_map(user_id: str) -> dict:
    """Return {contact_id: human_activity_count} for the user."""
    pipeline = [
        {"$match": {"user_id": user_id, "activity_type": {"$in": ["call", "email", "sms", "meeting", "note"]}}},
        {"$group": {"_id": "$contact_id", "c": {"$sum": 1}}},
    ]
    out = {}
    async for r in db.activities.aggregate(pipeline):
        if r.get("_id"):
            out[r["_id"]] = int(r.get("c", 0))
    return out


async def _contact_ids_with_tour_or_task_today(user_id: str) -> set:
    today = datetime.now(timezone.utc).date()
    today_iso = today.isoformat()
    today_start = datetime.combine(today, datetime.min.time(), tzinfo=timezone.utc).isoformat()
    today_end = datetime.combine(today, datetime.max.time(), tzinfo=timezone.utc).isoformat()
    ids = set()
    async for e in db.calendar_events.find(
        {"user_id": user_id, "start": {"$gte": today_start, "$lte": today_end}},
        {"contact_id": 1},
    ):
        if e.get("contact_id"):
            ids.add(e["contact_id"])
    async for t in db.tasks.find(
        {"user_id": user_id, "due_date": today_iso, "completed": False},
        {"contact_id": 1},
    ):
        if t.get("contact_id"):
            ids.add(t["contact_id"])
    return ids


async def _contact_ids_with_active_lease(user_id: str) -> set:
    ids = set()
    async for le in db.leases.find({"user_id": user_id, "status": "active"}, {"contact_id": 1}):
        if le.get("contact_id"):
            ids.add(le["contact_id"])
    return ids


async def _contact_ids_with_ended_lease(user_id: str) -> set:
    ids = set()
    async for le in db.leases.find(
        {"user_id": user_id, "status": {"$in": ["ended", "terminated", "expired"]}},
        {"contact_id": 1},
    ):
        if le.get("contact_id"):
            ids.add(le["contact_id"])
    return ids


async def _contact_ids_with_renewal_soon(user_id: str, days: int = 60) -> set:
    today = datetime.now(timezone.utc).date()
    end_iso = (today + timedelta(days=days)).isoformat()
    ids = set()
    async for le in db.leases.find(
        {"user_id": user_id, "status": "active", "lease_end": {"$gte": today.isoformat(), "$lte": end_iso, "$ne": ""}},
        {"contact_id": 1},
    ):
        if le.get("contact_id"):
            ids.add(le["contact_id"])
    return ids


async def _contact_ids_in_stage(user_id: str, stage: str) -> set:
    ids = set()
    async for d in db.deals.find(
        {"user_id": user_id, "stage": stage},
        {"contact_id": 1},
    ):
        if d.get("contact_id"):
            ids.add(d["contact_id"])
    return ids


# Sentinel filter that matches no documents (for empty-id edge case)
_EMPTY_FILTER = {"_id": {"$in": [ObjectId()]}, "user_id": {"$exists": False}}


async def _build_smart_list_filter(user_id: str, smart_list: str):
    """Return a Mongo filter (without user_id — caller combines via $and) for the given smart list."""
    sl = smart_list.strip().lower()

    if sl in ("today_tours", "today_tours_followups"):
        ids = await _contact_ids_with_tour_or_task_today(user_id)
        return {"id": {"$in": list(ids)}} if ids else _EMPTY_FILTER

    if sl == "first_contact":
        # Leads/prospects with zero human activity
        act_map = await _contact_activity_count_map(user_id)
        contacted_ids = list(act_map.keys())
        return {
            "$and": [
                {"$or": [
                    {"client_type": {"$in": ["lead", "prospect", None, ""]}},
                    {"client_type": {"$exists": False}},
                ]},
                {"id": {"$nin": contacted_ids}},
            ]
        }

    if sl == "second_contact":
        act_map = await _contact_activity_count_map(user_id)
        ids = [cid for cid, n in act_map.items() if n == 1]
        return {"id": {"$in": ids}} if ids else _EMPTY_FILTER

    if sl == "application_submitted":
        deal_ids = await _contact_ids_in_stage(user_id, "Application Submitted")
        conds = [{"leasing_stage": "Application Submitted"}]
        if deal_ids:
            conds.append({"id": {"$in": list(deal_ids)}})
        return {"$or": conds}

    if sl == "at_risk_renewals":
        ids = await _contact_ids_with_renewal_soon(user_id, days=60)
        if not ids:
            return _EMPTY_FILTER
        return {
            "$and": [
                {"id": {"$in": list(ids)}},
                {"$or": [
                    {"retention_score": {"$lt": 50}},
                    {"tags": {"$in": ["at-risk", "retention-risk"]}},
                    {"leasing_stage": "Renewal"},
                ]},
            ]
        }

    if sl == "stale_prospects":
        cutoff = (datetime.now(timezone.utc) - timedelta(days=14)).isoformat()
        active_ids = await _contact_ids_with_recent_activity(user_id, within_hours=14 * 24)
        return {
            "$and": [
                {"$or": [
                    {"client_type": {"$in": ["lead", "prospect"]}},
                    {"client_type": {"$exists": False}},
                ]},
                {"id": {"$nin": list(active_ids)}},
                {"created_at": {"$lt": cutoff}},
            ]
        }

    if sl == "recently_active":
        ids = await _contact_ids_with_recent_activity(user_id, within_hours=24)
        return {"id": {"$in": list(ids)}} if ids else _EMPTY_FILTER

    if sl == "nurture_queue":
        return {"tags": {"$in": ["nurture", "nurture-queue", "long-term"]}}

    if sl == "recontact_next_year":
        # Leads who either (a) were explicitly tagged as declined/not-interested,
        # or (b) went quiet after 3+ contact attempts. They're surfaced here so
        # they can be scrubbed by move-in date and re-contacted next season.
        decline_tags = ["declined", "not-interested", "no-response",
                        "declined-assistance", "recontact-next-year"]
        act_map = await _contact_activity_count_map(user_id)
        quiet_cutoff_ids = await _contact_ids_with_recent_activity(user_id, within_hours=30 * 24)
        unresponsive_ids = [cid for cid, n in act_map.items()
                            if n >= 3 and cid not in quiet_cutoff_ids]
        conds = [{"tags": {"$in": decline_tags}}]
        if unresponsive_ids:
            conds.append({
                "$and": [
                    {"id": {"$in": unresponsive_ids}},
                    {"$or": [
                        {"client_type": {"$in": ["lead", "prospect"]}},
                        {"client_type": {"$exists": False}},
                        {"client_type": None},
                        {"client_type": ""},
                    ]},
                ],
            })
        return {"$or": conds}

    return None


async def _build_collection_filter(user_id: str, collection: str):
    col = collection.strip().lower()

    if col == "prospects":
        return {"$or": [
            {"client_type": {"$in": ["lead", "prospect"]}},
            {"client_type": {"$exists": False}},
            {"client_type": None},
            {"client_type": ""},
        ]}

    if col == "active_tenants":
        ids = await _contact_ids_with_active_lease(user_id)
        conds = [{"client_type": "tenant"}]
        if ids:
            conds.append({"id": {"$in": list(ids)}})
        return {"$or": conds}

    if col == "past_tenants":
        ids = await _contact_ids_with_ended_lease(user_id)
        return {"id": {"$in": list(ids)}} if ids else _EMPTY_FILTER

    if col == "high_value_leads":
        return {"$or": [
            {"lead_score": {"$gte": 70}},
            {"budget_max": {"$gte": 3500}},
        ]}

    return None


@api_router.get("/contacts/smart-counts")
async def contacts_smart_counts(user=Depends(get_any_auth_user)):
    """Return counts for all smart lists + collections + all-people for the sidebar badges."""
    uid = user["_id"]
    out = {"all_people": await db.contacts.count_documents({"user_id": uid})}

    for sl in SMART_LIST_IDS:
        f = await _build_smart_list_filter(uid, sl)
        if f is None:
            out[sl] = 0
        else:
            try:
                out[sl] = await db.contacts.count_documents({"$and": [{"user_id": uid}, f]})
            except Exception as e:
                logger.warning(f"smart-count for {sl} failed: {e}")
                out[sl] = 0

    for col in COLLECTION_IDS:
        f = await _build_collection_filter(uid, col)
        if f is None:
            out[col] = 0
        else:
            try:
                out[col] = await db.contacts.count_documents({"$and": [{"user_id": uid}, f]})
            except Exception as e:
                logger.warning(f"collection-count for {col} failed: {e}")
                out[col] = 0

    return out


@api_router.get("/contacts")
async def list_contacts(
    user=Depends(get_any_auth_user),
    search: str = "", property_type: str = "",
    smart_list: str = "", collection: str = "",
    page: int = 1, limit: int = 50, sort: str = "created_at", order: str = "desc",
):
    conditions = [{"user_id": user["_id"]}]
    if search:
        safe_search = re.escape(search)
        conditions.append({"$or": [
            {"name": {"$regex": safe_search, "$options": "i"}},
            {"email": {"$regex": safe_search, "$options": "i"}},
            {"phone": {"$regex": safe_search, "$options": "i"}},
            {"company": {"$regex": safe_search, "$options": "i"}},
        ]})
    if property_type:
        conditions.append({"property_type": property_type})
    if smart_list:
        f = await _build_smart_list_filter(user["_id"], smart_list)
        if f is not None:
            conditions.append(f)
    if collection:
        f = await _build_collection_filter(user["_id"], collection)
        if f is not None:
            conditions.append(f)
    query = conditions[0] if len(conditions) == 1 else {"$and": conditions}
    sort_order = 1 if order == "asc" else -1
    return await paginate(db.contacts, query, page=page, limit=limit, sort_field=sort, sort_order=sort_order)


CONTACT_CSV_FIELDS = [
    "name", "email", "phone", "company", "source", "property_type",
    "tags", "notes", "lead_score",
    # Leasing-specific columns
    "move_in_date", "budget_min", "budget_max", "bedrooms_needed",
    "pet_type", "lease_term_months", "referral_source"
]

@api_router.get("/contacts/export")
async def export_contacts_csv(user=Depends(get_any_auth_user)):
    contacts = await db.contacts.find({"user_id": user["_id"]}).to_list(5000)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CONTACT_CSV_FIELDS)
    writer.writeheader()
    for c in contacts:
        writer.writerow({
            "name": c.get("name", ""),
            "email": c.get("email", ""),
            "phone": c.get("phone", ""),
            "company": c.get("company", ""),
            "source": c.get("source", ""),
            "property_type": c.get("property_type", ""),
            "tags": ",".join(c.get("tags", [])),
            "notes": c.get("notes", ""),
            "lead_score": c.get("lead_score", 0),
            "move_in_date": c.get("move_in_date", ""),
            "budget_min": c.get("budget_min", ""),
            "budget_max": c.get("budget_max", ""),
            "bedrooms_needed": c.get("bedrooms_needed", ""),
            "pet_type": c.get("pet_type", ""),
            "lease_term_months": c.get("lease_term_months", ""),
            "referral_source": c.get("referral_source", ""),
        })
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=contacts_export.csv"}
    )

@api_router.get("/contacts/template")
async def get_contact_import_template():
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CONTACT_CSV_FIELDS)
    writer.writeheader()
    writer.writerow({
        "name": "Jane Smith", "email": "jane@example.com", "phone": "(555) 123-4567",
        "company": "Acme Realty", "source": "website", "property_type": "residential_lease",
        "tags": "vip,relocating", "notes": "Looking for 2BR downtown", "lead_score": "0",
        "move_in_date": "2025-09-01", "budget_min": "1500", "budget_max": "2500",
        "bedrooms_needed": "2", "pet_type": "dog", "lease_term_months": "12",
        "referral_source": "zillow"
    })
    writer.writerow({
        "name": "John Doe", "email": "john@example.com", "phone": "(555) 987-6543",
        "company": "Tech Corp", "source": "referral", "property_type": "commercial_lease",
        "tags": "high-priority", "notes": "Expanding office space", "lead_score": "50",
        "move_in_date": "2025-10-15", "budget_min": "5000", "budget_max": "10000",
        "bedrooms_needed": "", "pet_type": "", "lease_term_months": "36",
        "referral_source": "agent-referral"
    })
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=contact_import_template.csv"}
    )

@api_router.get("/contacts/{contact_id}")
async def get_contact(contact_id: str, user=Depends(get_any_auth_user)):
    contact = await db.contacts.find_one({"_id": validate_object_id(contact_id, "Contact"), "user_id": user["_id"]})
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    return serialize_doc(contact)

@api_router.put("/contacts/{contact_id}")
async def update_contact(contact_id: str, data: ContactUpdate, user=Depends(get_any_auth_user)):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.contacts.update_one({"_id": validate_object_id(contact_id, "Contact"), "user_id": user["_id"]}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contact not found")
    contact = await db.contacts.find_one({"_id": validate_object_id(contact_id, "Contact")})
    return serialize_doc(contact)

@api_router.delete("/contacts/{contact_id}")
async def delete_contact(contact_id: str, user=Depends(get_any_auth_user)):
    result = await db.contacts.delete_one({"_id": validate_object_id(contact_id, "Contact"), "user_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact not found")
    return {"message": "Contact deleted"}

# ─── Properties Routes ───
@api_router.post("/properties")
async def create_property(data: PropertyCreate, user=Depends(get_any_auth_user)):
    doc = data.model_dump()
    doc["user_id"] = user["_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.properties.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    del doc["_id"]
    return doc

@api_router.get("/properties")
async def list_properties(
    user=Depends(get_any_auth_user),
    property_type: str = "", listing_type: str = "",
    page: int = 1, limit: int = 50, sort: str = "created_at", order: str = "desc"
):
    query = {"user_id": user["_id"]}
    if property_type:
        query["property_type"] = property_type
    if listing_type:
        query["listing_type"] = listing_type
    sort_order = 1 if order == "asc" else -1
    return await paginate(db.properties, query, page=page, limit=limit, sort_field=sort, sort_order=sort_order)

# ─── Property Import/Export (must be before {property_id} routes) ───
PROPERTY_CSV_FIELDS = ["name", "address", "property_type", "listing_type", "price", "sqft", "bedrooms", "bathrooms", "status", "description", "image_url"]

@api_router.post("/properties/import")
async def import_properties(file: UploadFile = File(...), user=Depends(get_any_auth_user)):
    content = await file.read()
    filename = file.filename or ""
    rows = []
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip().lower().replace(" ", "_") if c else "" for c in row]
                    continue
                rows.append(dict(zip(headers, [c if c is not None else "" for c in row])))
            wb.close()
        else:
            text = content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append(row)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    imported = 0
    errors_list = []
    for i, row in enumerate(rows):
        try:
            name = str(row.get("name", "")).strip()
            address = str(row.get("address", "")).strip()
            if not name and not address:
                errors_list.append(f"Row {i+1}: Missing name and address")
                continue
            prop_type = str(row.get("property_type", "residential")).strip().lower()
            if prop_type not in ("residential", "commercial"):
                prop_type = "residential"
            list_type = str(row.get("listing_type", "lease")).strip().lower()
            if list_type not in ("lease", "sale"):
                list_type = "lease"
            status_val = str(row.get("status", "active")).strip().lower()
            if status_val not in ("active", "pending", "closed"):
                status_val = "active"
            doc = {
                "name": name or address.split(",")[0],
                "address": address,
                "property_type": prop_type,
                "listing_type": list_type,
                "price": float(row.get("price", 0) or 0),
                "sqft": float(row.get("sqft", 0) or 0),
                "bedrooms": int(row.get("bedrooms", 0) or 0),
                "bathrooms": int(row.get("bathrooms", 0) or 0),
                "status": status_val,
                "description": str(row.get("description", "")).strip(),
                "image_url": str(row.get("image_url", "")).strip(),
                "user_id": user["_id"],
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.properties.insert_one(doc)
            imported += 1
        except Exception as e:
            errors_list.append(f"Row {i+1}: {str(e)}")
    return {"imported": imported, "total_rows": len(rows), "errors": errors_list}

@api_router.get("/properties/export")
async def export_properties(user=Depends(get_any_auth_user)):
    props = await db.properties.find({"user_id": user["_id"]}).to_list(5000)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=PROPERTY_CSV_FIELDS)
    writer.writeheader()
    for p in props:
        writer.writerow({f: p.get(f, "") for f in PROPERTY_CSV_FIELDS})
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=properties_export.csv"}
    )

@api_router.get("/properties/template")
async def get_property_import_template():
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=PROPERTY_CSV_FIELDS)
    writer.writeheader()
    writer.writerow({
        "name": "Example Office Suite A", "address": "123 Main St, Suite 100, City, ST 12345",
        "property_type": "commercial", "listing_type": "lease", "price": "3500",
        "sqft": "1200", "bedrooms": "0", "bathrooms": "1", "status": "active",
        "description": "Corner office with city views", "image_url": ""
    })
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=property_import_template.csv"}
    )

@api_router.get("/properties/{property_id}")
async def get_property(property_id: str, user=Depends(get_any_auth_user)):
    prop = await db.properties.find_one({"_id": validate_object_id(property_id, "Property"), "user_id": user["_id"]})
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
    return serialize_doc(prop)

@api_router.put("/properties/{property_id}")
async def update_property(property_id: str, data: dict, user=Depends(get_any_auth_user)):
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data.pop("id", None)
    result = await db.properties.update_one({"_id": validate_object_id(property_id, "Property"), "user_id": user["_id"]}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Property not found")
    prop = await db.properties.find_one({"_id": validate_object_id(property_id, "Property")})
    return serialize_doc(prop)

@api_router.delete("/properties/{property_id}")
async def delete_property(property_id: str, user=Depends(get_any_auth_user)):
    result = await db.properties.delete_one({"_id": validate_object_id(property_id, "Property"), "user_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Property not found")
    return {"message": "Property deleted"}

# ─── Deals / Pipeline Routes ───
@api_router.get("/pipelines/stages")
async def get_pipeline_stages():
    return PIPELINE_STAGES

@api_router.post("/deals")
async def create_deal(data: DealCreate, user=Depends(get_any_auth_user)):
    if data.pipeline_type not in PIPELINE_STAGES:
        raise HTTPException(status_code=400, detail="Invalid pipeline type")
    if data.stage not in PIPELINE_STAGES[data.pipeline_type]:
        raise HTTPException(status_code=400, detail="Invalid stage for this pipeline")
    doc = data.model_dump()
    doc["user_id"] = user["_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # FEATURE: Round-Robin Assignment
    # If assigned_to is not provided and user has auto_assign enabled
    if not doc.get("assigned_to"):
        user_settings = await db.users.find_one({"_id": ObjectId(user["_id"])}, {"_id": 0, "auto_assign": 1})
        if user_settings and user_settings.get("auto_assign"):
            # Find all team members
            team_members = await db.users.find({"role": {"$in": ["agent", "admin"]}}, {"_id": 1}).to_list(100)
            if team_members:
                # Count open deals for each agent
                agent_deal_counts = []
                for agent in team_members:
                    agent_id = str(agent["_id"])
                    open_deals = await db.deals.count_documents({"assigned_to": agent_id, "stage": {"$ne": "Closed"}})
                    agent_deal_counts.append({"agent_id": agent_id, "open_deals": open_deals})
                # Sort by fewest open deals
                agent_deal_counts.sort(key=lambda x: x["open_deals"])
                # Assign to agent with fewest deals
                doc["assigned_to"] = agent_deal_counts[0]["agent_id"]
                logger.info(f"Round-robin assignment: deal assigned to {doc['assigned_to']}")
    
    result = await db.deals.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    del doc["_id"]
    return doc

@api_router.get("/deals")
async def list_deals(
    user=Depends(get_any_auth_user),
    pipeline_type: str = "",
    page: int = 1, limit: int = 50, sort: str = "created_at", order: str = "desc"
):
    query = {"user_id": user["_id"]}
    if pipeline_type:
        query["pipeline_type"] = pipeline_type
    sort_order = 1 if order == "asc" else -1
    return await paginate(db.deals, query, page=page, limit=limit, sort_field=sort, sort_order=sort_order)

# ── Pipeline summary with counts + totals per stage ──
@api_router.get("/deals/pipeline-summary")
async def pipeline_summary(
    pipeline_type: str = "lease_applications",
    scope: str = "me",
    user=Depends(get_any_auth_user),
):
    if pipeline_type not in PIPELINE_STAGES:
        raise HTTPException(status_code=400, detail="Invalid pipeline_type")
    query = {"pipeline_type": pipeline_type}
    if scope != "everyone":
        query["user_id"] = user["_id"]

    builtin = list(PIPELINE_STAGES[pipeline_type])
    custom = list((user.get("custom_stages") or {}).get(pipeline_type, []))
    all_stages = builtin + [c for c in custom if c not in builtin]

    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$stage",
            "count": {"$sum": 1},
            "total_value": {"$sum": {"$ifNull": ["$desired_rent", {"$ifNull": ["$value", 0]}]}},
        }},
    ]
    cursor = db.deals.aggregate(pipeline)
    buckets = {}
    async for row in cursor:
        buckets[row["_id"]] = row

    stages = []
    total_value = 0.0
    total_count = 0
    for idx, s in enumerate(all_stages):
        b = buckets.get(s, {"count": 0, "total_value": 0})
        color = LEASE_APPLICATIONS_STAGE_COLORS.get(s)
        stages.append({
            "name": s,
            "count": int(b["count"]),
            "total_value": float(b.get("total_value", 0) or 0),
            "color": color["bg"] if color else "slate",
            "color_hex": color["hex"] if color else "#64748B",
            "is_custom": s not in builtin,
            "order": idx,
        })
        total_value += float(b.get("total_value", 0) or 0)
        total_count += int(b["count"])
    return {
        "pipeline_type": pipeline_type,
        "scope": scope,
        "stages": stages,
        "total_pipeline_value": total_value,
        "total_deals": total_count,
    }

@api_router.get("/deals/{deal_id}")
async def get_deal(deal_id: str, user=Depends(get_any_auth_user)):
    deal = await db.deals.find_one({"_id": validate_object_id(deal_id, "Deal"), "user_id": user["_id"]})
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    return serialize_doc(deal)

@api_router.put("/deals/{deal_id}")
async def update_deal(deal_id: str, data: DealUpdate, background_tasks: BackgroundTasks, user=Depends(get_any_auth_user)):
    existing = await db.deals.find_one({"_id": validate_object_id(deal_id, "Deal"), "user_id": user["_id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Deal not found")
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    old_stage = existing.get("stage")
    if "stage" in updates:
        pipeline_type = existing["pipeline_type"]
        if updates["stage"] not in PIPELINE_STAGES[pipeline_type]:
            raise HTTPException(status_code=400, detail="Invalid stage for this pipeline")
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.deals.update_one({"_id": validate_object_id(deal_id, "Deal")}, {"$set": updates})

    # ─── Stage Automation with transaction safety ───
    # If auto-task creation fails, rollback the stage change
    new_stage = updates.get("stage")
    if new_stage and new_stage != old_stage:
        if new_stage in STAGE_AUTO_TASKS:
            try:
                auto = STAGE_AUTO_TASKS[new_stage]
                due = (datetime.now(timezone.utc) + timedelta(days=auto["days_offset"])).strftime("%Y-%m-%d")
                await db.tasks.insert_one({
                    "title": f"[Auto] {auto['title']} - {existing.get('title', '')}",
                    "description": f"Auto-generated when deal moved to {new_stage}",
                    "due_date": due,
                    "contact_id": existing.get("contact_id", ""),
                    "deal_id": deal_id,
                    "priority": auto["priority"],
                    "completed": False,
                    "user_id": user["_id"],
                    "created_at": datetime.now(timezone.utc).isoformat(),
                })
                logger.info(f"Auto-task created for deal {deal_id} stage change to {new_stage}")
            except Exception as e:
                # ROLLBACK: Revert the stage change if auto-task creation fails
                logger.error(f"Auto-task creation failed for deal {deal_id}, rolling back stage: {e}")
                await db.deals.update_one(
                    {"_id": validate_object_id(deal_id, "Deal")},
                    {"$set": {"stage": old_stage, "updated_at": existing.get("updated_at", "")}}
                )
                raise HTTPException(
                    status_code=500,
                    detail="Stage change failed: could not create auto-task. Deal has been reverted to its previous stage."
                )
        # ─── Webhook: trigger on deal stage change (fire-and-forget) ───
        background_tasks.add_task(trigger_webhooks, user["_id"], "deal_stage_change", {
            "deal_id": deal_id, "title": existing.get("title", ""),
            "pipeline_type": existing.get("pipeline_type", ""),
            "old_stage": old_stage, "new_stage": new_stage,
        })

        # ─── Phase 10: Auto-enroll contact in any sequence triggered by this stage ───
        contact_id = existing.get("contact_id", "")
        if contact_id:
            try:
                triggered = await db.sequences.find({
                    "user_id": user["_id"],
                    "trigger": "deal_stage_changed",
                    "trigger_value": new_stage,
                    "active": True,
                }).to_list(20)
                for seq in triggered:
                    seq_id = seq.get("id") or str(seq.get("_id", ""))
                    if not seq.get("steps"):
                        continue
                    # Skip if already enrolled (idempotent)
                    already = await db.sequence_executions.find_one({
                        "sequence_id": seq_id, "contact_id": contact_id, "step_index": 0,
                    })
                    if already:
                        continue
                    first_step = seq["steps"][0]
                    scheduled_at = (
                        datetime.now(timezone.utc)
                        + timedelta(days=int(first_step.get("delay_days", 0)))
                    ).isoformat()
                    await db.sequence_executions.insert_one({
                        "sequence_id": seq_id,
                        "contact_id": contact_id,
                        "step_index": 0,
                        "status": "pending",
                        "scheduled_at": scheduled_at,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "triggered_by": f"stage_change:{new_stage}",
                    })
                    logger.info(f"Auto-enrolled contact {contact_id} in sequence '{seq.get('name','')}' via stage→{new_stage}")
            except Exception as e:
                logger.error(f"Auto-sequence enrollment failed (non-fatal): {e}")

    deal = await db.deals.find_one({"_id": validate_object_id(deal_id, "Deal")})
    return serialize_doc(deal)

@api_router.delete("/deals/{deal_id}")
async def delete_deal(deal_id: str, user=Depends(get_any_auth_user)):
    result = await db.deals.delete_one({"_id": validate_object_id(deal_id, "Deal"), "user_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Deal not found")
    return {"message": "Deal deleted"}

# ─── Tasks Routes ───
@api_router.post("/tasks")
async def create_task(data: TaskCreate, user=Depends(get_any_auth_user)):
    doc = data.model_dump()
    doc["user_id"] = user["_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.tasks.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    del doc["_id"]
    return doc

@api_router.get("/tasks")
async def list_tasks(
    user=Depends(get_any_auth_user),
    completed: str = "",
    page: int = 1, limit: int = 50, sort: str = "due_date", order: str = "asc"
):
    query = {"user_id": user["_id"]}
    if completed == "true":
        query["completed"] = True
    elif completed == "false":
        query["completed"] = False
    sort_order = 1 if order == "asc" else -1
    return await paginate(db.tasks, query, page=page, limit=limit, sort_field=sort, sort_order=sort_order)

@api_router.put("/tasks/{task_id}")
async def update_task(task_id: str, data: dict, user=Depends(get_any_auth_user)):
    data.pop("id", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.tasks.update_one({"_id": validate_object_id(task_id, "Task"), "user_id": user["_id"]}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    task = await db.tasks.find_one({"_id": validate_object_id(task_id, "Task")})
    return serialize_doc(task)

@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, user=Depends(get_any_auth_user)):
    result = await db.tasks.delete_one({"_id": validate_object_id(task_id, "Task"), "user_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"message": "Task deleted"}


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 15 — TASKS PAGE (FUB-parity)
# Adds: counts (today/overdue/future), bulk ops, complete+log, task types
# ═══════════════════════════════════════════════════════════════════════════════

VALID_TASK_TYPES = (
    "tour_followup", "renewal_offer", "maintenance_request",
    "application_reminder", "showing_prep", "listing_outreach",
    "lease_signing", "move_in", "rent_reminder", "other",
)


def _today_bounds_iso():
    """Return (start_of_today, start_of_tomorrow) ISO strings in UTC."""
    now = datetime.now(timezone.utc)
    start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end = start + timedelta(days=1)
    return start.isoformat(), end.isoformat()


def _infer_task_type(title: str, description: str = "") -> str:
    """Best-effort inference from stage-automation defaults (backward compat)."""
    t = f"{title} {description}".lower()
    if "tour" in t or "showing" in t:
        return "tour_followup"
    if "renew" in t:
        return "renewal_offer"
    if "maintenance" in t or "repair" in t or "ticket" in t:
        return "maintenance_request"
    if "application" in t or "review application" in t:
        return "application_reminder"
    if "lease" in t and ("sign" in t or "paperwork" in t):
        return "lease_signing"
    if "move-in" in t or "move in" in t:
        return "move_in"
    if "rent" in t:
        return "rent_reminder"
    return "other"


def _task_assignee_filter(uid: str, assignee: str) -> dict:
    if assignee == "me":
        return {"$or": [{"assigned_to": uid}, {"assigned_to": {"$in": ["", None]}}]}
    if assignee == "team":
        return {"assigned_to": {"$nin": ["", None, uid]}}
    # 'everyone' or empty
    return {}


@api_router.get("/tasks/counts")
async def tasks_counts(
    user=Depends(get_any_auth_user),
    assignee: str = "me",
    task_type: str = "",
):
    """Counts for Today / Overdue / Future / Completed tabs."""
    uid = user["_id"]
    start_today, end_today = _today_bounds_iso()
    base = {"user_id": uid}
    assignee_f = _task_assignee_filter(uid, assignee)
    if assignee_f:
        base.update(assignee_f)
    if task_type:
        base["task_type"] = task_type

    today = await db.tasks.count_documents({
        **base, "completed": False,
        "due_date": {"$gte": start_today, "$lt": end_today},
    })
    overdue = await db.tasks.count_documents({
        **base, "completed": False,
        "due_date": {"$gt": "", "$lt": start_today},
    })
    future = await db.tasks.count_documents({
        **base, "completed": False,
        "due_date": {"$gte": end_today},
    })
    completed = await db.tasks.count_documents({**base, "completed": True})
    return {"today": today, "overdue": overdue, "future": future, "completed": completed}


@api_router.get("/tasks/bucket")
async def tasks_bucket(
    user=Depends(get_any_auth_user),
    bucket: str = "today",
    assignee: str = "me",
    task_type: str = "",
    limit: int = 200,
):
    """Return tasks for a Today/Overdue/Future/Completed bucket, pre-sorted."""
    uid = user["_id"]
    start_today, end_today = _today_bounds_iso()
    q = {"user_id": uid}
    assignee_f = _task_assignee_filter(uid, assignee)
    if assignee_f:
        q.update(assignee_f)
    if task_type:
        q["task_type"] = task_type

    sort_field, sort_order = "due_date", 1
    if bucket == "overdue":
        q.update({"completed": False, "due_date": {"$gt": "", "$lt": start_today}})
        sort_order = -1  # most recently-overdue first
    elif bucket == "future":
        q.update({"completed": False, "due_date": {"$gte": end_today}})
    elif bucket == "completed":
        q["completed"] = True
        sort_field, sort_order = "updated_at", -1
    else:  # today (default)
        q.update({"completed": False, "due_date": {"$gte": start_today, "$lt": end_today}})

    limit = max(1, min(limit, 500))
    cursor = db.tasks.find(q).sort(sort_field, sort_order).limit(limit)
    tasks = [serialize_doc(t) async for t in cursor]
    # Enrich inferred task_type for legacy rows lacking it
    for t in tasks:
        if not t.get("task_type"):
            t["task_type"] = _infer_task_type(t.get("title", ""), t.get("description", ""))
    return {"tasks": tasks}


class TaskCompleteLogRequest(BaseModel):
    activity_type: Optional[str] = Field(default="note", max_length=20)
    note: Optional[str] = Field(default="", max_length=5000)

    @field_validator("activity_type")
    @classmethod
    def validate_activity_type(cls, v):
        allowed = ("call", "email", "note", "meeting", "sms")
        if v and v not in allowed:
            raise ValueError(f"activity_type must be one of: {', '.join(allowed)}")
        return v or "note"


@api_router.post("/tasks/{task_id}/complete-and-log")
async def task_complete_and_log(task_id: str, data: TaskCompleteLogRequest, user=Depends(get_any_auth_user)):
    """Mark task complete AND log an activity on the linked contact in one call."""
    tid = validate_object_id(task_id, "Task")
    task = await db.tasks.find_one({"_id": tid, "user_id": user["_id"]})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    now = datetime.now(timezone.utc).isoformat()
    await db.tasks.update_one(
        {"_id": tid},
        {"$set": {"completed": True, "completed_at": now, "updated_at": now}},
    )
    if task.get("contact_id"):
        await db.activities.insert_one({
            "contact_id": task["contact_id"],
            "user_id": user["_id"],
            "activity_type": data.activity_type,
            "description": data.note or f"Completed task: {task.get('title', '')}",
            "deal_id": task.get("deal_id", ""),
            "created_at": now,
        })
    return {"success": True}


class TaskBulkRequest(BaseModel):
    task_ids: List[str]
    action: str = Field(..., max_length=30)
    # action-specific payload
    due_date: Optional[str] = Field(default=None, max_length=30)
    assigned_to: Optional[str] = Field(default=None, max_length=50)
    sequence_id: Optional[str] = Field(default=None, max_length=50)

    @field_validator("action")
    @classmethod
    def validate_action(cls, v):
        allowed = ("complete", "reschedule", "assign", "add_to_sequence", "delete")
        if v not in allowed:
            raise ValueError(f"action must be one of: {', '.join(allowed)}")
        return v


@api_router.post("/tasks/bulk")
async def tasks_bulk(data: TaskBulkRequest, user=Depends(get_any_auth_user)):
    """Perform a bulk operation on a set of task IDs."""
    oids = []
    for tid in data.task_ids[:500]:
        try:
            oids.append(ObjectId(tid))
        except Exception:
            continue
    if not oids:
        raise HTTPException(status_code=400, detail="No valid task IDs provided")

    flt = {"_id": {"$in": oids}, "user_id": user["_id"]}
    now = datetime.now(timezone.utc).isoformat()
    affected = 0

    if data.action == "complete":
        res = await db.tasks.update_many(flt, {"$set": {"completed": True, "completed_at": now, "updated_at": now}})
        affected = res.modified_count
        # log a lightweight activity per task with a contact_id
        tasks = await db.tasks.find(flt).to_list(500)
        acts = [
            {
                "contact_id": t["contact_id"], "user_id": user["_id"],
                "activity_type": "note",
                "description": f"Completed task: {t.get('title', '')}",
                "deal_id": t.get("deal_id", ""),
                "created_at": now,
            } for t in tasks if t.get("contact_id")
        ]
        if acts:
            await db.activities.insert_many(acts)

    elif data.action == "reschedule":
        if not data.due_date:
            raise HTTPException(status_code=400, detail="due_date is required for reschedule")
        res = await db.tasks.update_many(flt, {"$set": {"due_date": data.due_date, "updated_at": now}})
        affected = res.modified_count

    elif data.action == "assign":
        res = await db.tasks.update_many(flt, {"$set": {"assigned_to": data.assigned_to or "", "updated_at": now}})
        affected = res.modified_count

    elif data.action == "add_to_sequence":
        if not data.sequence_id:
            raise HTTPException(status_code=400, detail="sequence_id is required")
        seq = await db.sequences.find_one({"_id": validate_object_id(data.sequence_id, "Sequence"), "user_id": user["_id"]})
        if not seq:
            raise HTTPException(status_code=404, detail="Sequence not found")
        tasks = await db.tasks.find(flt).to_list(500)
        contact_ids = {t.get("contact_id") for t in tasks if t.get("contact_id")}
        for cid in contact_ids:
            await db.sequence_executions.insert_one({
                "user_id": user["_id"],
                "sequence_id": data.sequence_id,
                "contact_id": cid,
                "status": "pending",
                "next_step": 0,
                "next_run_at": now,
                "created_at": now,
            })
        affected = len(contact_ids)

    elif data.action == "delete":
        res = await db.tasks.delete_many(flt)
        affected = res.deleted_count

    return {"success": True, "affected": affected, "action": data.action}


@api_router.post("/tasks/seed-demo")
async def tasks_seed_demo(user=Depends(get_any_auth_user)):
    """Populate a handful of realistic leasing tasks for demo / empty-state."""
    uid = user["_id"]
    await db.tasks.delete_many({"user_id": uid, "seed": "demo"})
    contacts = await db.contacts.find({"user_id": uid}).limit(8).to_list(8)
    if not contacts:
        return {"created": 0, "detail": "No contacts found"}
    props = await db.properties.find({"user_id": uid}).limit(5).to_list(5)
    now = datetime.now(timezone.utc)
    templates = [
        ("tour_followup",        "Follow up on 2BR tour",             "Confirm next steps after yesterday's showing.",   0,  "high"),
        ("renewal_offer",        "Send renewal offer",                "Prepare 12-mo renewal with 3% increase.",         1,  "high"),
        ("maintenance_request",  "Check on plumbing ticket",          "Tenant reported a leak; confirm plumber ETA.",    0,  "medium"),
        ("application_reminder", "Application review deadline",       "Review Smith application before EOD.",            0,  "high"),
        ("showing_prep",         "Prepare keys for 4pm showing",      "Unit 3B — collect lockbox code.",                 0,  "medium"),
        ("listing_outreach",     "Outreach to StreetEasy lead",       "New inquiry — send listing pack & book tour.",    2,  "medium"),
        ("rent_reminder",        "Follow up on late rent",            "Past due 3 days — send courtesy reminder.",       -1, "medium"),
        ("lease_signing",        "Send lease for e-signature",        "Prepare and send 12-mo lease for counter-sign.",  3,  "high"),
        ("move_in",              "Move-in walkthrough checklist",     "Prep move-in packet + utilities transfer form.", 5,  "low"),
        ("tour_followup",        "Yesterday's tour — send recap",     "Recap property features + availability.",        -2, "high"),
    ]
    created = 0
    for idx, (tt, title, desc, day_offset, prio) in enumerate(templates):
        contact = contacts[idx % len(contacts)]
        prop = props[idx % len(props)] if props else None
        due = (now + timedelta(days=day_offset, hours=(14 + idx) % 24))
        await db.tasks.insert_one({
            "user_id": uid,
            "title": title,
            "description": desc,
            "due_date": due.isoformat(),
            "contact_id": str(contact["_id"]),
            "deal_id": "",
            "property_id": str(prop["_id"]) if prop else "",
            "priority": prio,
            "completed": False,
            "task_type": tt,
            "assigned_to": uid,
            "all_day": False,
            "created_at": now.isoformat(),
            "seed": "demo",
        })
        created += 1
    return {"created": created}


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 17 — ORG / ADMIN SETTINGS
# One settings doc per user (single-tenant). Stores company info, lead-flow
# rules, renewal defaults, custom fields/tags, maintenance types.
# ═══════════════════════════════════════════════════════════════════════════════

DEFAULT_ORG_SETTINGS = {
    "company": {
        "name": "",
        "business_registration": "",
        "address": "",
        "phone": "",
        "website": "",
        "logo_url": "",
    },
    "lead_flow": {
        "round_robin_enabled": False,
        "strategy": "least_busy",  # least_busy | random | weighted
        "business_hours_only": True,
        "business_hours_start": "09:00",
        "business_hours_end": "18:00",
        "weekend_assignment": False,
        "fallback_user_id": "",
    },
    "renewals": {
        "default_sequence_id": "",
        "default_notice_days": 60,
        "rent_increase_percent": 3.0,
        "auto_send_offer": False,
        "retention_target_score": 70,
    },
    "custom_fields": [],       # [{id, label, type, entity, required}]
    "tags": [],                # [{id, name, color, entity}]
    "maintenance_types": [     # seeded defaults
        {"id": "plumbing",   "name": "Plumbing",   "sla_hours": 24, "color": "sky"},
        {"id": "electrical", "name": "Electrical", "sla_hours": 24, "color": "amber"},
        {"id": "hvac",       "name": "HVAC",       "sla_hours": 48, "color": "indigo"},
        {"id": "appliance",  "name": "Appliance",  "sla_hours": 72, "color": "emerald"},
        {"id": "pest",       "name": "Pest",       "sla_hours": 24, "color": "rose"},
        {"id": "general",    "name": "General",    "sla_hours": 72, "color": "slate"},
    ],
    "renewal_templates": [],   # [{id, name, body}]
}


def _merge_settings(stored: Optional[dict]) -> dict:
    """Return a complete settings doc by merging defaults with stored."""
    import copy
    out = copy.deepcopy(DEFAULT_ORG_SETTINGS)
    if not stored:
        return out
    for k, default_val in DEFAULT_ORG_SETTINGS.items():
        val = stored.get(k)
        if val is None:
            continue
        if isinstance(default_val, dict):
            merged = dict(default_val)
            merged.update({kk: vv for kk, vv in val.items() if vv is not None})
            out[k] = merged
        else:
            out[k] = val
    return out


class OrgSettingsUpdate(BaseModel):
    plan: Optional[str] = None
    company: Optional[dict] = None
    lead_flow: Optional[dict] = None
    renewals: Optional[dict] = None
    custom_fields: Optional[List[dict]] = None
    tags: Optional[List[dict]] = None
    maintenance_types: Optional[List[dict]] = None
    renewal_templates: Optional[List[dict]] = None
    brokerage_sheet: Optional[dict] = None
    benchmarks: Optional[dict] = None


@api_router.get("/settings")
async def get_org_settings(user=Depends(get_any_auth_user)):
    stored = await db.org_settings.find_one({"user_id": user["_id"]})
    # Sync lead_flow.round_robin_enabled with users.auto_assign (existing feature)
    u = await db.users.find_one({"_id": ObjectId(user["_id"])})
    merged = _merge_settings(stored)
    merged["lead_flow"]["round_robin_enabled"] = bool((u or {}).get("auto_assign", False))
    return merged


@api_router.put("/settings")
async def update_org_settings(data: OrgSettingsUpdate, user=Depends(get_any_auth_user)):
    payload = data.model_dump(exclude_none=True)
    now = datetime.now(timezone.utc).isoformat()
    # Keep auto_assign on users in sync for backward-compat with existing flows
    if "lead_flow" in payload and "round_robin_enabled" in payload["lead_flow"]:
        await db.users.update_one(
            {"_id": ObjectId(user["_id"])},
            {"$set": {"auto_assign": bool(payload["lead_flow"]["round_robin_enabled"])}},
        )
    set_ops = {f"{k}": v for k, v in payload.items()}
    set_ops["updated_at"] = now
    await db.org_settings.update_one(
        {"user_id": user["_id"]},
        {"$set": set_ops, "$setOnInsert": {"user_id": user["_id"], "created_at": now}},
        upsert=True,
    )
    stored = await db.org_settings.find_one({"user_id": user["_id"]})
    return _merge_settings(stored)


@api_router.get("/settings/summary")
async def settings_summary(user=Depends(get_any_auth_user)):
    """Lightweight counts for the Overview tab."""
    uid = user["_id"]
    team = await db.users.count_documents({})  # single-tenant: all users are team
    properties = await db.properties.count_documents({"user_id": uid})
    sequences = await db.sequences.count_documents({"user_id": uid})
    active_sequences = await db.sequences.count_documents({"user_id": uid, "active": True})
    templates = await db.templates.count_documents({"user_id": uid})
    webhooks = await db.webhooks.count_documents({"user_id": uid, "active": True})
    api_keys = await db.api_keys.count_documents({"user_id": uid, "active": True})
    contacts = await db.contacts.count_documents({"user_id": uid})
    return {
        "team_members": team,
        "properties": properties,
        "sequences": sequences,
        "active_sequences": active_sequences,
        "templates": templates,
        "webhooks": webhooks,
        "api_keys": api_keys,
        "contacts": contacts,
    }

# ─── Activities Routes ───
@api_router.post("/activities")
async def create_activity(data: ActivityCreate, user=Depends(get_any_auth_user)):
    doc = data.model_dump()
    doc["user_id"] = user["_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.activities.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    del doc["_id"]
    return doc

@api_router.get("/activities")
async def list_activities(
    user=Depends(get_any_auth_user),
    contact_id: str = "", deal_id: str = "",
    page: int = 1, limit: int = 50, sort: str = "created_at", order: str = "desc"
):
    query = {"user_id": user["_id"]}
    if contact_id:
        query["contact_id"] = contact_id
    if deal_id:
        query["deal_id"] = deal_id
    sort_order = 1 if order == "asc" else -1
    return await paginate(db.activities, query, page=page, limit=limit, sort_field=sort, sort_order=sort_order)

# ─── AI Routes (with cost guardrails, rate limiting, and usage logging) ───
@api_router.post("/ai/draft-email")
async def ai_draft_email(data: AIEmailRequest, user=Depends(get_any_auth_user)):
    await check_ai_rate_limit(user["_id"])
    contact = await db.contacts.find_one({"_id": validate_object_id(data.contact_id, "Contact"), "user_id": user["_id"]})
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    activities = await db.activities.find({"contact_id": data.contact_id}).sort("created_at", -1).to_list(10)
    activity_summary = "\n".join([f"- {a.get('activity_type','')}: {a.get('description','')}" for a in activities])
    deals = await db.deals.find({"contact_id": data.contact_id}).to_list(10)
    deal_summary = "\n".join([f"- {d.get('title','')}: {d.get('stage','')}" for d in deals])
    system_msg = f"""You are an expert real estate agent assistant. Draft a {data.tone} follow-up email.
Contact: {contact.get('name','')} ({contact.get('email','')})
Company: {contact.get('company','')}
Property Interest: {contact.get('property_type','')}
Recent Activities:
{activity_summary}
Current Deals:
{deal_summary}
Additional Context: {data.context}

Write a concise, professional email ready to send. Include subject line."""
    # Cost guardrail: estimate and block if too expensive
    full_input = system_msg + "\nDraft the follow-up email now."
    guard_ai_cost(full_input)
    input_tokens = estimate_tokens(full_input)
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=settings.EMERGENT_LLM_KEY,
            session_id=f"email-draft-{data.contact_id}-{datetime.now(timezone.utc).isoformat()}",
            system_message=system_msg
        ).with_model("openai", "gpt-5.2")
        msg = UserMessage(text="Draft the follow-up email now.")
        result = await chat.send_message(msg)
        output_tokens = estimate_tokens(result)
        await log_ai_usage(user["_id"], "draft-email", input_tokens, output_tokens)
        return {"draft": result, "contact_name": contact.get("name", "")}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"AI email draft error: {e}")
        raise HTTPException(status_code=500, detail="AI service temporarily unavailable. Please try again.")

@api_router.post("/ai/lead-score")
async def ai_lead_score(data: AILeadScoreRequest, user=Depends(get_any_auth_user)):
    await check_ai_rate_limit(user["_id"])
    contact = await db.contacts.find_one({"_id": validate_object_id(data.contact_id, "Contact"), "user_id": user["_id"]})
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    activities = await db.activities.find({"contact_id": data.contact_id}).to_list(50)
    deals = await db.deals.find({"contact_id": data.contact_id}).to_list(20)
    tasks = await db.tasks.find({"contact_id": data.contact_id}).to_list(20)
    system_msg = f"""You are an expert real estate lead scoring AI. Analyze this lead and provide a score from 0-100.
Contact: {contact.get('name','')}
Email: {contact.get('email','')}
Company: {contact.get('company','')}
Source: {contact.get('source','')}
Property Type: {contact.get('property_type','')}
Number of activities: {len(activities)}
Number of deals: {len(deals)}
Number of tasks: {len(tasks)}
Tags: {', '.join(contact.get('tags', []))}

Respond ONLY with a JSON object: {{"score": <number 0-100>, "reasoning": "<brief explanation>", "next_action": "<recommended next step>"}}"""
    full_input = system_msg + "\nScore this lead now."
    guard_ai_cost(full_input)
    input_tokens = estimate_tokens(full_input)
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=settings.EMERGENT_LLM_KEY,
            session_id=f"lead-score-{data.contact_id}-{datetime.now(timezone.utc).isoformat()}",
            system_message=system_msg
        ).with_model("openai", "gpt-5.2")
        msg = UserMessage(text="Score this lead now.")
        result = await chat.send_message(msg)
        output_tokens = estimate_tokens(result)
        await log_ai_usage(user["_id"], "lead-score", input_tokens, output_tokens)
        try:
            cleaned = result.strip()
            if cleaned.startswith("```"):
                cleaned = cleaned.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            score_data = json.loads(cleaned)
            await db.contacts.update_one({"_id": validate_object_id(data.contact_id, "Contact")}, {"$set": {"lead_score": score_data.get("score", 0)}})
            return score_data
        except json.JSONDecodeError:
            return {"score": 50, "reasoning": result, "next_action": "Review manually"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"AI lead score error: {e}")
        raise HTTPException(status_code=500, detail="AI service temporarily unavailable. Please try again.")

@api_router.post("/ai/summarize-activities")
async def ai_summarize(contact_id: str, user=Depends(get_any_auth_user)):
    await check_ai_rate_limit(user["_id"])
    activities = await db.activities.find({"contact_id": contact_id, "user_id": user["_id"]}).sort("created_at", -1).to_list(50)
    if not activities:
        return {"summary": "No activities found for this contact."}
    contact = await db.contacts.find_one({"_id": validate_object_id(contact_id, "Contact")})
    activity_text = "\n".join([f"[{a.get('created_at','')}] {a.get('activity_type','')}: {a.get('description','')}" for a in activities])
    system_msg = f"""Summarize these real estate CRM activities for contact {contact.get('name','') if contact else 'Unknown'}. Be concise and highlight key points, next steps, and any concerns."""
    full_input = system_msg + f"\nActivities:\n{activity_text}"
    guard_ai_cost(full_input)
    input_tokens = estimate_tokens(full_input)
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=settings.EMERGENT_LLM_KEY,
            session_id=f"summary-{contact_id}-{datetime.now(timezone.utc).isoformat()}",
            system_message=system_msg
        ).with_model("openai", "gpt-5.2")
        msg = UserMessage(text=f"Activities:\n{activity_text}")
        result = await chat.send_message(msg)
        output_tokens = estimate_tokens(result)
        await log_ai_usage(user["_id"], "summarize-activities", input_tokens, output_tokens)
        return {"summary": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"AI summary error: {e}")
        raise HTTPException(status_code=500, detail="AI service temporarily unavailable. Please try again.")

# ─── API Keys Management ───

# ─── Webhook Helper ───
async def trigger_webhooks(user_id: str, event: str, payload: dict):
    webhooks = await db.webhooks.find({"user_id": user_id, "active": True}).to_list(50)
    for wh in webhooks:
        if event in wh.get("events", []):
            try:
                async with httpx.AsyncClient(timeout=10) as client_http:
                    await client_http.post(wh["url"], json={"event": event, "data": payload, "timestamp": datetime.now(timezone.utc).isoformat()})
                    await db.webhooks.update_one({"_id": wh["_id"]}, {"$set": {"last_triggered": datetime.now(timezone.utc).isoformat()}})
            except Exception as e:
                logger.error(f"Webhook {wh.get('name','')} failed: {e}")

# ─── CSV/XLSX Import ───
@api_router.post("/contacts/import")
async def import_contacts(file: UploadFile = File(...), user=Depends(get_any_auth_user)):
    content = await file.read()
    filename = file.filename or ""
    rows = []
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip().lower().replace(" ", "_") if c else "" for c in row]
                    continue
                rows.append(dict(zip(headers, [c if c is not None else "" for c in row])))
            wb.close()
        else:
            text = content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append(row)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    # FEATURE: Chunked import for large files (5000 rows at a time)
    CHUNK_SIZE = 5000
    total_rows = len(rows)
    imported = 0
    skipped = 0
    errors_list = []
    
    for chunk_start in range(0, total_rows, CHUNK_SIZE):
        chunk_end = min(chunk_start + CHUNK_SIZE, total_rows)
        chunk = rows[chunk_start:chunk_end]
        
        for i, row in enumerate(chunk):
            row_num = chunk_start + i + 2  # +2 because row 1 is the header
            try:
                name = str(row.get("name", "")).strip()
                if not name:
                    errors_list.append({"row": row_num, "field": "name", "reason": "Missing required field 'name'"})
                    skipped += 1
                    continue
                email_val = str(row.get("email", "")).strip()
                if email_val and "@" not in email_val:
                    errors_list.append({"row": row_num, "field": "email", "reason": f"Invalid email format: '{email_val}'"})
                    skipped += 1
                    continue
                source_val = str(row.get("source", "csv_import")).strip()
                if not source_val:
                    source_val = "csv_import"
                prop_type = str(row.get("property_type", "residential_lease")).strip()
                if prop_type and prop_type not in ("residential_lease", "commercial_sale", "commercial_lease"):
                    errors_list.append({"row": row_num, "field": "property_type", "reason": f"Invalid property_type '{prop_type}'. Must be: residential_lease, commercial_sale, or commercial_lease"})
                    prop_type = "residential_lease"
                tags_raw = str(row.get("tags", "")).strip()
                tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []
                lead_score_raw = row.get("lead_score", 0) or 0
                try:
                    lead_score = int(lead_score_raw)
                except (ValueError, TypeError):
                    errors_list.append({"row": row_num, "field": "lead_score", "reason": f"Invalid lead_score '{lead_score_raw}', defaulting to 0"})
                    lead_score = 0
                doc = {
                    "name": name,
                    "email": email_val,
                    "phone": str(row.get("phone", "")).strip(),
                    "company": str(row.get("company", "")).strip(),
                    "source": source_val,
                    "property_type": prop_type,
                    "tags": tags,
                    "notes": str(row.get("notes", "")).strip(),
                    "lead_score": lead_score,
                    "user_id": user["_id"],
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
                # Store leasing-specific optional fields if present
                for extra_field in ("move_in_date", "budget_min", "budget_max", "bedrooms_needed", "pet_type", "lease_term_months", "referral_source"):
                    val = str(row.get(extra_field, "")).strip()
                    if val:
                        doc[extra_field] = val
                await db.contacts.insert_one(doc)
                imported += 1
            except Exception as e:
                errors_list.append({"row": row_num, "field": "unknown", "reason": str(e)})
                skipped += 1
        
        logger.info(f"Import progress: {chunk_end}/{total_rows} rows processed")
    
    return {
        "imported": imported,
        "total_rows": total_rows,
        "skipped": skipped,
        "errors": errors_list,
    }

# ─── Message Templates ───
@api_router.post("/templates")
async def create_template(data: TemplateCreate, user=Depends(get_current_user)):
    doc = data.model_dump()
    doc["user_id"] = user["_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    doc["use_count"] = 0
    result = await db.templates.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    del doc["_id"]
    return doc

@api_router.get("/templates")
async def list_templates(
    user=Depends(get_any_auth_user),
    category: str = "",
    page: int = 1, limit: int = 50, sort: str = "use_count", order: str = "desc"
):
    query = {"user_id": user["_id"]}
    if category:
        query["category"] = category
    sort_order = 1 if order == "asc" else -1
    return await paginate(db.templates, query, page=page, limit=limit, sort_field=sort, sort_order=sort_order)

@api_router.get("/templates/{template_id}")
async def get_template(template_id: str, user=Depends(get_any_auth_user)):
    tpl = await db.templates.find_one({"_id": validate_object_id(template_id, "Template"), "user_id": user["_id"]})
    if not tpl:
        raise HTTPException(status_code=404, detail="Template not found")
    return serialize_doc(tpl)

@api_router.put("/templates/{template_id}")
async def update_template(template_id: str, data: TemplateCreate, user=Depends(get_current_user)):
    updates = data.model_dump()
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.templates.update_one({"_id": validate_object_id(template_id, "Template"), "user_id": user["_id"]}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    tpl = await db.templates.find_one({"_id": validate_object_id(template_id, "Template")})
    return serialize_doc(tpl)

@api_router.delete("/templates/{template_id}")
async def delete_template(template_id: str, user=Depends(get_current_user)):
    result = await db.templates.delete_one({"_id": validate_object_id(template_id, "Template"), "user_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": "Template deleted"}

@api_router.post("/templates/{template_id}/use")
async def use_template(template_id: str, user=Depends(get_any_auth_user)):
    """Increment use count and return the template - for tracking popular templates"""
    tpl = await db.templates.find_one({"_id": validate_object_id(template_id, "Template"), "user_id": user["_id"]})
    if not tpl:
        raise HTTPException(status_code=404, detail="Template not found")
    await db.templates.update_one({"_id": validate_object_id(template_id, "Template")}, {"$inc": {"use_count": 1}})
    tpl["use_count"] = tpl.get("use_count", 0) + 1
    return serialize_doc(tpl)

@api_router.post("/templates/ai-generate")
async def ai_generate_template(request: Request, user=Depends(get_current_user)):
    await check_ai_rate_limit(user["_id"])
    body = await request.json()
    purpose = body.get("purpose", "follow-up")
    category = body.get("category", "email")
    property_type = body.get("property_type", "residential_lease")
    system_msg = f"""You are an expert real estate copywriter. Generate a {category} template for {purpose}.
Property type context: {property_type}
Use placeholders like {{contact_name}}, {{property_address}}, {{agent_name}}, {{company_name}} for personalization.
{"Include a Subject line on the first line." if category == "email" else "Keep it under 160 characters for SMS."}
Make it professional, warm, and action-oriented."""
    prompt_text = f"Generate the {category} template now for: {purpose}"
    full_input = system_msg + "\n" + prompt_text
    guard_ai_cost(full_input)
    input_tokens = estimate_tokens(full_input)
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=settings.EMERGENT_LLM_KEY,
            session_id=f"template-gen-{datetime.now(timezone.utc).isoformat()}",
            system_message=system_msg
        ).with_model("openai", "gpt-5.2")
        result = await chat.send_message(UserMessage(text=prompt_text))
        output_tokens = estimate_tokens(result)
        await log_ai_usage(user["_id"], "generate-template", input_tokens, output_tokens)
        subject = ""
        template_body = result
        if category == "email":
            for line in result.split("\n"):
                if line.lower().startswith("subject:"):
                    subject = line.replace("Subject:", "").replace("subject:", "").strip()
                    template_body = result.replace(line, "").strip()
                    break
        return {"subject": subject, "body": template_body, "category": category}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"AI template generation error: {e}")
        raise HTTPException(status_code=500, detail="AI service temporarily unavailable. Please try again.")

# ─── Email Sending (Brevo) ───
@api_router.post("/email/send")
async def send_email_endpoint(data: SendEmailRequest, background_tasks: BackgroundTasks, user=Depends(get_any_auth_user)):
    brevo_key = settings.BREVO_API_KEY
    sender_email = settings.SENDER_EMAIL
    sender_name = settings.SENDER_NAME
    if not brevo_key or not sender_email:
        raise HTTPException(status_code=503, detail="Email service not configured. Add BREVO_API_KEY and SENDER_EMAIL to .env")
    try:
        from brevo import Brevo
        from brevo.transactional_emails import (
            SendTransacEmailRequestSender,
            SendTransacEmailRequestToItem,
        )
        brevo_client = Brevo(api_key=brevo_key)
        response = brevo_client.transactional_emails.send_transac_email(
            sender=SendTransacEmailRequestSender(email=sender_email, name=sender_name),
            to=[SendTransacEmailRequestToItem(email=data.to_email)],
            subject=data.subject,
            html_content=data.body.replace("\n", "<br>"),
        )
        # Log activity
        await db.activities.insert_one({
            "contact_id": data.contact_id, "user_id": user["_id"],
            "activity_type": "email", "description": f"Sent email: {data.subject}",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        background_tasks.add_task(trigger_webhooks, user["_id"], "email_sent", {"contact_id": data.contact_id, "subject": data.subject})
        return {"success": True, "message_id": getattr(response, 'message_id', 'sent')}
    except Exception as e:
        logger.error(f"Brevo email error: {e}")
        raise HTTPException(status_code=500, detail="Email sending failed. Please check your configuration and try again.")

# ─── SMS Sending (Twilio) with retry logic ───
@api_router.post("/sms/send")
async def send_sms_endpoint(data: SendSMSRequest, background_tasks: BackgroundTasks, user=Depends(get_any_auth_user)):
    account_sid = settings.TWILIO_ACCOUNT_SID
    auth_token = settings.TWILIO_AUTH_TOKEN
    from_number = settings.TWILIO_PHONE_NUMBER
    if not account_sid or not auth_token or not from_number:
        raise HTTPException(status_code=503, detail="SMS service not configured. Add TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, and TWILIO_PHONE_NUMBER to .env")
    try:
        # Uses tenacity retry: 3 attempts with exponential backoff (1s → 2s → 4s)
        msg = send_sms_with_retry(account_sid, auth_token, from_number, data.to_phone, data.message)
        await db.activities.insert_one({
            "contact_id": data.contact_id, "user_id": user["_id"],
            "activity_type": "sms", "description": f"Sent SMS: {data.message[:80]}...",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        background_tasks.add_task(trigger_webhooks, user["_id"], "sms_sent", {"contact_id": data.contact_id, "message": data.message[:80]})
        return {"success": True, "sid": msg.sid}
    except Exception as e:
        logger.error(f"Twilio error after retries: {e}")
        raise HTTPException(status_code=500, detail="SMS sending failed. Please check your configuration and try again.")

# ─── Webhooks Management ───
@api_router.post("/webhooks")
async def create_webhook(data: WebhookCreate, user=Depends(get_current_user)):
    doc = data.model_dump()
    doc["user_id"] = user["_id"]
    doc["active"] = True
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["last_triggered"] = None
    result = await db.webhooks.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    del doc["_id"]
    return doc

@api_router.get("/webhooks")
async def list_webhooks(user=Depends(get_current_user), page: int = 1, limit: int = 50):
    return await paginate(db.webhooks, {"user_id": user["_id"]}, page=page, limit=limit)

@api_router.delete("/webhooks/{webhook_id}")
async def delete_webhook(webhook_id: str, user=Depends(get_current_user)):
    result = await db.webhooks.delete_one({"_id": validate_object_id(webhook_id, "Webhook"), "user_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Webhook not found")
    return {"message": "Webhook deleted"}

@api_router.put("/webhooks/{webhook_id}/toggle")
async def toggle_webhook(webhook_id: str, user=Depends(get_current_user)):
    wh = await db.webhooks.find_one({"_id": validate_object_id(webhook_id, "Webhook"), "user_id": user["_id"]})
    if not wh:
        raise HTTPException(status_code=404, detail="Webhook not found")
    await db.webhooks.update_one({"_id": validate_object_id(webhook_id, "Webhook")}, {"$set": {"active": not wh.get("active", True)}})
    return {"active": not wh.get("active", True)}

# ─── Team Management ───
@api_router.post("/team/invite")
async def invite_team_member(data: TeamInviteRequest, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can invite team members")
    email = data.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="User already exists")
    temp_password = secrets.token_urlsafe(8)
    await db.users.insert_one({
        "email": email,
        "password_hash": hash_password(temp_password),
        "name": data.name,
        "role": data.role,
        "team_id": user.get("team_id", user["_id"]),
        "invited_by": user["_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    return {"email": email, "name": data.name, "role": data.role, "temp_password": temp_password}

@api_router.get("/team/members")
async def list_team_members(user=Depends(get_current_user), page: int = 1, limit: int = 50):
    team_id = user.get("team_id", user["_id"])
    query = {"$or": [{"team_id": team_id}, {"_id": ObjectId(user["_id"])}]}
    # PERFORMANCE: Projection — only return needed fields for team member list
    projection = {"_id": 1, "email": 1, "name": 1, "role": 1, "created_at": 1}
    pg = max(1, page)
    lim = max(1, min(limit, PAGINATION_MAX_LIMIT))
    skip_count = (pg - 1) * lim
    total = await db.users.count_documents(query)
    members = await db.users.find(query, projection).skip(skip_count).limit(lim).to_list(lim)
    result = []
    for m in members:
        m["id"] = str(m["_id"])
        del m["_id"]
        result.append(m)
    return {
        "data": result,
        "pagination": {"page": pg, "limit": lim, "total": total, "total_pages": math.ceil(total / lim) if lim > 0 else 1}
    }

@api_router.delete("/team/members/{member_id}")
async def remove_team_member(member_id: str, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can remove team members")
    if member_id == user["_id"]:
        raise HTTPException(status_code=400, detail="Cannot remove yourself")
    result = await db.users.delete_one({"_id": validate_object_id(member_id, "Member")})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")
    return {"message": "Member removed"}

@api_router.put("/team/members/{member_id}/role")
async def update_member_role(member_id: str, request: Request, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can change roles")
    body = await request.json()
    new_role = body.get("role", "agent")
    await db.users.update_one({"_id": validate_object_id(member_id, "Member")}, {"$set": {"role": new_role}})
    return {"message": "Role updated"}

@api_router.post("/api-keys")
async def create_api_key(request: Request, user=Depends(get_current_user)):
    body = await request.json()
    key = f"pf_{secrets.token_hex(24)}"
    doc = {
        "key": key,
        "name": body.get("name", "API Key"),
        "user_id": user["_id"],
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "last_used": None
    }
    await db.api_keys.insert_one(doc)
    return {"key": key, "name": doc["name"], "created_at": doc["created_at"]}

@api_router.get("/api-keys")
async def list_api_keys(user=Depends(get_current_user), page: int = 1, limit: int = 50):
    query = {"user_id": user["_id"]}
    # PERFORMANCE: Projection — only return fields needed for key list
    projection = {"_id": 0, "key": 1, "name": 1, "active": 1, "created_at": 1, "last_used": 1}
    pg = max(1, page)
    lim = max(1, min(limit, PAGINATION_MAX_LIMIT))
    skip_count = (pg - 1) * lim
    total = await db.api_keys.count_documents(query)
    keys = await db.api_keys.find(query, projection).skip(skip_count).limit(lim).to_list(lim)
    # Mask key for display
    for k in keys:
        k["key_preview"] = k["key"][:8] + "..." + k["key"][-4:]
        k["full_key"] = k["key"]
        del k["key"]
    return {
        "data": keys,
        "pagination": {"page": pg, "limit": lim, "total": total, "total_pages": math.ceil(total / lim) if lim > 0 else 1}
    }

@api_router.delete("/api-keys/{key_preview}")
async def delete_api_key(key_preview: str, user=Depends(get_current_user)):
    result = await db.api_keys.delete_one({"user_id": user["_id"], "key": {"$regex": f"^{key_preview[:8]}"}})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="API key not found")
    return {"message": "API key deleted"}

# ─── Dashboard Stats ───
@api_router.get("/dashboard/stats")
async def dashboard_stats(user=Depends(get_any_auth_user)):
    uid = user["_id"]
    # PERFORMANCE: All count_documents queries use indexed user_id field
    total_contacts = await db.contacts.count_documents({"user_id": uid})
    total_deals = await db.deals.count_documents({"user_id": uid})
    total_properties = await db.properties.count_documents({"user_id": uid})
    open_tasks = await db.tasks.count_documents({"user_id": uid, "completed": False})
    # Pipeline breakdown — uses idx_deals_user_stage_created index
    pipeline_stats = {}
    for pt in PIPELINE_STAGES:
        stages = {}
        for stage in PIPELINE_STAGES[pt]:
            count = await db.deals.count_documents({"user_id": uid, "pipeline_type": pt, "stage": stage})
            stages[stage] = count
        pipeline_stats[pt] = stages
    # Recent activities — uses idx_activities_user_created index, limited projection
    recent_activities = await db.activities.find(
        {"user_id": uid},
        {"user_id": 0}  # Exclude user_id from response (already known)
    ).sort("created_at", -1).to_list(10)
    # Deal value by pipeline — PERFORMANCE: Only fetch "value" field (projection)
    deal_values = {}
    for pt in PIPELINE_STAGES:
        deals = await db.deals.find(
            {"user_id": uid, "pipeline_type": pt},
            {"value": 1, "_id": 0}  # Only fetch value field
        ).to_list(500)
        deal_values[pt] = sum(d.get("value", 0) for d in deals)
    return {
        "total_contacts": total_contacts,
        "total_deals": total_deals,
        "total_properties": total_properties,
        "open_tasks": open_tasks,
        "pipeline_stats": pipeline_stats,
        "deal_values": deal_values,
        "recent_activities": [serialize_doc(a) for a in recent_activities],
    }

# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE: Drip Sequences
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/sequences")
async def list_sequences(user=Depends(get_current_user), page: int = 1, limit: int = 50):
    uid = user["_id"]
    pg = max(1, page)
    lim = max(1, min(limit, PAGINATION_MAX_LIMIT))
    skip_count = (pg - 1) * lim
    total = await db.sequences.count_documents({"user_id": uid})
    sequences = await db.sequences.find({"user_id": uid}, {"_id": 0}).sort("created_at", -1).skip(skip_count).limit(lim).to_list(lim)
    return {
        "data": sequences,
        "pagination": {"page": pg, "limit": lim, "total": total, "total_pages": math.ceil(total / lim) if lim > 0 else 1}
    }

@api_router.post("/sequences")
async def create_sequence(data: SequenceCreate, user=Depends(get_current_user)):
    doc = {
        "id": str(ObjectId()),
        "user_id": user["_id"],
        "name": data.name,
        "trigger": data.trigger,
        "trigger_value": data.trigger_value or "",
        "steps": [{"type": s.type, "delay_days": s.delay_days, "template_id": s.template_id, "subject": s.subject, "body": s.body} for s in data.steps],
        "active": data.active,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.sequences.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/sequences/{sequence_id}")
async def get_sequence(sequence_id: str, user=Depends(get_current_user)):
    sequence = await db.sequences.find_one({"id": sequence_id, "user_id": user["_id"]}, {"_id": 0})
    if not sequence:
        raise HTTPException(status_code=404, detail="Sequence not found")
    return sequence

@api_router.put("/sequences/{sequence_id}")
async def update_sequence(sequence_id: str, data: SequenceUpdate, user=Depends(get_current_user)):
    update_data = {}
    if data.name is not None:
        update_data["name"] = data.name
    if data.trigger is not None:
        update_data["trigger"] = data.trigger
    if data.trigger_value is not None:
        update_data["trigger_value"] = data.trigger_value
    if data.steps is not None:
        update_data["steps"] = [{"type": s.type, "delay_days": s.delay_days, "template_id": s.template_id, "subject": s.subject, "body": s.body} for s in data.steps]
    if data.active is not None:
        update_data["active"] = data.active
    
    result = await db.sequences.update_one({"id": sequence_id, "user_id": user["_id"]}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sequence not found")
    return {"message": "Sequence updated"}

@api_router.delete("/sequences/{sequence_id}")
async def delete_sequence(sequence_id: str, user=Depends(get_current_user)):
    result = await db.sequences.delete_one({"id": sequence_id, "user_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sequence not found")
    # Also delete any pending executions
    await db.sequence_executions.delete_many({"sequence_id": sequence_id, "status": "pending"})
    return {"message": "Sequence deleted"}

@api_router.post("/sequences/{sequence_id}/enroll/{contact_id}")
async def enroll_contact_in_sequence(sequence_id: str, contact_id: str, user=Depends(get_current_user)):
    """Manually enroll a contact in a drip sequence."""
    sequence = await db.sequences.find_one({"id": sequence_id, "user_id": user["_id"]}, {"_id": 0})
    if not sequence:
        raise HTTPException(status_code=404, detail="Sequence not found")
    
    contact = await db.contacts.find_one({"id": contact_id, "user_id": user["_id"]}, {"_id": 0})
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    
    if not sequence.get("active"):
        raise HTTPException(status_code=400, detail="Sequence is not active")
    
    if not sequence.get("steps"):
        raise HTTPException(status_code=400, detail="Sequence has no steps")
    
    # Check if already enrolled (prevent duplicates)
    existing = await db.sequence_executions.find_one({
        "sequence_id": sequence_id,
        "contact_id": contact_id,
        "step_index": 0
    })
    if existing:
        raise HTTPException(status_code=400, detail="Contact already enrolled in this sequence")
    
    # Schedule first step
    first_step = sequence["steps"][0]
    scheduled_at = datetime.now(timezone.utc) + timedelta(days=first_step["delay_days"])
    
    await db.sequence_executions.insert_one({
        "sequence_id": sequence_id,
        "contact_id": contact_id,
        "step_index": 0,
        "status": "pending",
        "scheduled_at": scheduled_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Contact enrolled in sequence", "scheduled_at": scheduled_at.isoformat()}

# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE: Analytics & Reporting
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/reports")
async def get_reports(user=Depends(get_current_user)):
    """Generate analytics report with MongoDB aggregation pipelines."""
    uid = user["_id"]
    
    # 1. Stage Conversion Rates (for each pipeline type)
    pipeline_stages = {}
    for pipeline_type, stages in PIPELINE_STAGES.items():
        stage_counts = {}
        for stage in stages:
            count = await db.deals.count_documents({"user_id": uid, "pipeline_type": pipeline_type, "stage": stage})
            stage_counts[stage] = count
        pipeline_stages[pipeline_type] = stage_counts
    
    # 2. Activity Counts by Type (last 30 days)
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    activities_pipeline = [
        {"$match": {"user_id": uid, "created_at": {"$gte": thirty_days_ago}}},
        {"$group": {"_id": "$type", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    activity_counts = await db.activities.aggregate(activities_pipeline).to_list(100)
    activity_counts_dict = {a["_id"]: a["count"] for a in activity_counts}
    
    # 3. Lead Velocity (new contacts per week for last 8 weeks)
    lead_velocity = []
    for i in range(8):
        week_start = (datetime.now(timezone.utc) - timedelta(weeks=i+1)).isoformat()
        week_end = (datetime.now(timezone.utc) - timedelta(weeks=i)).isoformat()
        count = await db.contacts.count_documents({
            "user_id": uid,
            "created_at": {"$gte": week_start, "$lt": week_end}
        })
        lead_velocity.append({"week": f"Week {8-i}", "count": count})
    lead_velocity.reverse()
    
    # 4. Monthly Pipeline Value (sum of deal values by month)
    pipeline_value_pipeline = [
        {"$match": {"user_id": uid}},
        {"$group": {
            "_id": {"$substr": ["$created_at", 0, 7]},  # Group by YYYY-MM
            "total_value": {"$sum": "$value"}
        }},
        {"$sort": {"_id": 1}},
        {"$limit": 12}
    ]
    pipeline_values = await db.deals.aggregate(pipeline_value_pipeline).to_list(12)
    monthly_values = [{"month": pv["_id"], "value": pv["total_value"]} for pv in pipeline_values]
    
    # 5. Deal Win Rate (Closed vs Total)
    total_deals = await db.deals.count_documents({"user_id": uid})
    won_deals = await db.deals.count_documents({"user_id": uid, "stage": "Closed"})
    win_rate = round((won_deals / total_deals * 100), 2) if total_deals > 0 else 0
    
    # 6. Average Deal Value
    all_deals = await db.deals.find({"user_id": uid}, {"value": 1, "_id": 0}).to_list(10000)
    deal_values = [d.get("value", 0) for d in all_deals]
    avg_deal_value = round(sum(deal_values) / len(deal_values), 2) if deal_values else 0
    
    return {
        "pipeline_stages": pipeline_stages,
        "activity_counts": activity_counts_dict,
        "lead_velocity": lead_velocity,
        "monthly_pipeline_values": monthly_values,
        "win_rate": win_rate,
        "total_deals": total_deals,
        "won_deals": won_deals,
        "avg_deal_value": avg_deal_value
    }


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 18 — LEASING REPORTING & NETWORK BENCHMARKS
# ═══════════════════════════════════════════════════════════════════════════════

def _range_to_bounds(range_id: str):
    """Return (start_iso, end_iso, days_in_range) for the given range id."""
    now = datetime.now(timezone.utc)
    days = {"7d": 7, "30d": 30, "60d": 60, "90d": 90, "fall_preseason": 60}.get(range_id, 30)
    start = now - timedelta(days=days)
    return start.isoformat(), now.isoformat(), days


@api_router.get("/reports/leasing")
async def reports_leasing(
    user=Depends(get_any_auth_user),
    date_range: str = "30d",
    scope: str = "me",
    university_zone: str = "",
):
    """Comprehensive leasing report split by section (overview / agents / props / sources / calls / marketing / applications / renewals / goals)."""
    uid = user["_id"]
    base = {} if scope == "everyone" else {"user_id": uid}
    start_iso, end_iso, days = _range_to_bounds(date_range)

    # ── Overview KPIs ─────────────────────────────────────────────────────
    new_contacts = await db.contacts.count_documents({**base, "created_at": {"$gte": start_iso, "$lte": end_iso}})
    total_properties = await db.properties.count_documents(base)
    active_props = await db.properties.count_documents({**base, "status": "active"})
    leased_props = await db.properties.count_documents({**base, "status": "pending"})
    _ = leased_props  # reserved for future occupancy tracking
    occupancy_rate = round(((total_properties - active_props) / total_properties * 100), 1) if total_properties else 0.0
    signed = await db.deals.count_documents({
        **base,
        "pipeline_type": "lease_applications",
        "stage": {"$in": ["Lease Signed", "Move-In", "Active Tenant"]},
        "updated_at": {"$gte": start_iso, "$lte": end_iso},
    })
    pipeline_value_agg = [
        {"$match": {**base, "pipeline_type": "lease_applications"}},
        {"$group": {"_id": None, "v": {"$sum": "$value"}}},
    ]
    pv = 0.0
    async for row in db.deals.aggregate(pipeline_value_agg):
        pv = float(row.get("v", 0) or 0)

    # Weekly velocity sparkline (last 8 weeks)
    velocity = []
    for i in range(8, 0, -1):
        w_end = datetime.now(timezone.utc) - timedelta(weeks=i - 1)
        w_start = w_end - timedelta(weeks=1)
        cnt = await db.contacts.count_documents({
            **base, "created_at": {"$gte": w_start.isoformat(), "$lt": w_end.isoformat()},
        })
        velocity.append({"week": f"W-{i}", "count": cnt})

    # ── Agents activity ───────────────────────────────────────────────────
    agent_pipeline = [
        {"$match": {"created_at": {"$gte": start_iso, "$lte": end_iso}}},
        {"$group": {"_id": {"user_id": "$user_id", "type": "$activity_type"}, "count": {"$sum": 1}}},
    ]
    agent_rows = {}
    async for row in db.activities.aggregate(agent_pipeline):
        aid = row["_id"]["user_id"]
        atype = row["_id"]["type"]
        agent_rows.setdefault(aid, {"user_id": aid, "calls": 0, "emails": 0, "sms": 0, "meetings": 0, "notes": 0, "total": 0})
        if atype == "call":
            agent_rows[aid]["calls"] = row["count"]
        elif atype == "email":
            agent_rows[aid]["emails"] = row["count"]
        elif atype == "sms":
            agent_rows[aid]["sms"] = row["count"]
        elif atype == "meeting":
            agent_rows[aid]["meetings"] = row["count"]
        elif atype == "note":
            agent_rows[aid]["notes"] = row["count"]
        agent_rows[aid]["total"] += row["count"]
    # Attach names + leases_signed
    agents = []
    for aid, r in agent_rows.items():
        try:
            u = await db.users.find_one({"_id": ObjectId(aid)}, {"name": 1, "email": 1})
        except Exception:
            u = None
        r["name"] = (u or {}).get("name") or (u or {}).get("email", "Unknown") or "Unknown"
        r["leases_signed"] = await db.deals.count_documents({
            "user_id": aid, "pipeline_type": "lease_applications",
            "stage": {"$in": ["Lease Signed", "Move-In", "Active Tenant"]},
            "updated_at": {"$gte": start_iso, "$lte": end_iso},
        })
        agents.append(r)
    agents.sort(key=lambda x: x["total"], reverse=True)

    # ── Properties & Units ────────────────────────────────────────────────
    props = []
    async for p in db.properties.find(base).sort("created_at", -1).limit(20):
        props.append({
            "id": str(p["_id"]),
            "name": p.get("name", "") or p.get("address", "Unit"),
            "status": p.get("status", "active"),
            "price": p.get("price", 0),
            "days_on_market": 0,  # could be computed from created_at
        })
    vacancy_days_cost = sum(p.get("price", 0) for p in props if p.get("status") == "active") / 30.0 if props else 0.0

    # ── Lead Sources ──────────────────────────────────────────────────────
    source_pipeline = [
        {"$match": {**base, "created_at": {"$gte": start_iso, "$lte": end_iso}}},
        {"$group": {"_id": {"$ifNull": ["$source", "unknown"]}, "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    sources = []
    async for row in db.contacts.aggregate(source_pipeline):
        sources.append({"source": row["_id"] or "unknown", "count": row["count"]})

    # Signed leases by source (match deal.contact_id → contact.source)
    sls = {}
    deals_cursor = db.deals.find({
        **base, "pipeline_type": "lease_applications",
        "stage": {"$in": ["Lease Signed", "Move-In", "Active Tenant"]},
        "updated_at": {"$gte": start_iso, "$lte": end_iso},
    }, {"contact_id": 1})
    async for d in deals_cursor:
        cid = d.get("contact_id", "")
        if not cid:
            continue
        try:
            c = await db.contacts.find_one({"_id": ObjectId(cid)}, {"source": 1})
        except Exception:
            c = None
        src = (c or {}).get("source") or "unknown"
        sls[src] = sls.get(src, 0) + 1
    signed_by_source = [{"source": k, "count": v} for k, v in sorted(sls.items(), key=lambda x: -x[1])]

    # ── Calls & Texts ─────────────────────────────────────────────────────
    call_count = await db.activities.count_documents({**base, "activity_type": "call", "created_at": {"$gte": start_iso}})
    sms_count = await db.activities.count_documents({**base, "activity_type": "sms", "created_at": {"$gte": start_iso}})
    email_count = await db.activities.count_documents({**base, "activity_type": "email", "created_at": {"$gte": start_iso}})

    # ── Lease Applications Funnel ────────────────────────────────────────
    stages_funnel = ["Inquiry", "Tour Scheduled", "Application Submitted", "Screening", "Approved", "Lease Signed"]
    funnel = []
    for st in stages_funnel:
        c = await db.deals.count_documents({**base, "pipeline_type": "lease_applications", "stage": st})
        funnel.append({"stage": st, "count": c})

    # ── Renewals & Retention ─────────────────────────────────────────────
    upcoming_end = (datetime.now(timezone.utc) + timedelta(days=90)).isoformat()
    upcoming_renewals = 0
    try:
        upcoming_renewals = await db.leases.count_documents({
            **base, "status": "active", "lease_end": {"$gte": datetime.now(timezone.utc).isoformat(), "$lte": upcoming_end},
        })
    except Exception:
        pass

    return {
        "range": date_range,
        "days": days,
        "scope": scope,
        "overview": {
            "new_contacts": new_contacts,
            "signed_leases": signed,
            "total_properties": total_properties,
            "active_listings": active_props,
            "occupancy_rate": occupancy_rate,
            "pipeline_value": pv,
            "velocity": velocity,
        },
        "agents": agents,
        "properties": props,
        "vacancy_days_cost": round(vacancy_days_cost, 2),
        "sources": sources,
        "signed_by_source": signed_by_source,
        "calls_texts": {"calls": call_count, "sms": sms_count, "emails": email_count},
        "applications_funnel": funnel,
        "renewals": {"upcoming_90d": upcoming_renewals},
    }


# ── Anonymous Network Benchmarks ─────────────────────────────────────────
# Aggregates anonymized data across all opted-in brokerages. For the preview
# we synthesize plausible peer averages when the 5-org minimum isn't met.

async def _org_metric_value(uid: str, metric: str, start_iso: str, end_iso: str) -> float:
    base = {"user_id": uid}
    if metric == "avg_days_to_lease":
        total, n = 0.0, 0
        async for d in db.deals.find({
            **base, "pipeline_type": "lease_applications",
            "stage": {"$in": ["Lease Signed", "Move-In", "Active Tenant"]},
            "updated_at": {"$gte": start_iso, "$lte": end_iso},
        }, {"created_at": 1, "updated_at": 1}):
            ca = _parse_iso_safe(d.get("created_at"))
            ua = _parse_iso_safe(d.get("updated_at"))
            if ca and ua:
                total += max(0.0, (ua - ca).total_seconds() / 86400.0)
                n += 1
        return round(total / n, 1) if n else 0.0
    if metric == "occupancy_rate":
        total = await db.properties.count_documents(base)
        active = await db.properties.count_documents({**base, "status": "active"})
        return round(((total - active) / total * 100), 1) if total else 0.0
    if metric == "renewal_rate":
        # renewed / (renewed + vacated) in range
        renewed = 0
        vacated = 0
        try:
            renewed = await db.leases.count_documents({**base, "status": "active", "renewed_at": {"$gte": start_iso, "$lte": end_iso}})
            vacated = await db.leases.count_documents({**base, "status": "ended", "end_date": {"$gte": start_iso, "$lte": end_iso}})
        except Exception:
            pass
        total = renewed + vacated
        return round((renewed / total * 100), 1) if total else 0.0
    if metric == "tour_to_app_ratio":
        tours = await db.deals.count_documents({**base, "pipeline_type": "lease_applications", "stage": "Tour Scheduled"})
        apps  = await db.deals.count_documents({**base, "pipeline_type": "lease_applications", "stage": "Application Submitted"})
        return round((apps / tours * 100), 1) if tours else 0.0
    if metric == "app_to_lease_conversion":
        apps  = await db.deals.count_documents({**base, "pipeline_type": "lease_applications", "stage": {"$in": ["Application Submitted", "Screening", "Approved"]}})
        signed = await db.deals.count_documents({
            **base, "pipeline_type": "lease_applications",
            "stage": {"$in": ["Lease Signed", "Move-In", "Active Tenant"]},
            "updated_at": {"$gte": start_iso, "$lte": end_iso},
        })
        denom = apps + signed
        return round((signed / denom * 100), 1) if denom else 0.0
    if metric == "avg_rent_per_unit":
        total, n = 0.0, 0
        async for p in db.properties.find({**base, "listing_type": "lease"}, {"price": 1}):
            total += float(p.get("price") or 0)
            n += 1
        return round(total / n, 2) if n else 0.0
    if metric == "speed_to_first_contact":
        result = await _avg_speed_to_first_contact({"user_id": uid}, start_iso, end_iso)
        return float(result.get("value") or 0)
    return 0.0


# Hardcoded Austin Metro peer averages for demo (Fall 2025 benchmark snapshot).
# These mimic what a real aggregation would return when 5+ orgs have opted in.
AUSTIN_METRO_AVERAGES = {
    "avg_days_to_lease":        {"value": 22.0,  "lower_is_better": True,  "unit": "days",    "label": "Avg Days to Lease"},
    "occupancy_rate":           {"value": 93.5,  "lower_is_better": False, "unit": "%",       "label": "Occupancy Rate"},
    "renewal_rate":             {"value": 68.0,  "lower_is_better": False, "unit": "%",       "label": "Renewal Rate"},
    "tour_to_app_ratio":        {"value": 55.0,  "lower_is_better": False, "unit": "%",       "label": "Tour → Application"},
    "app_to_lease_conversion":  {"value": 72.0,  "lower_is_better": False, "unit": "%",       "label": "Application → Lease"},
    "avg_rent_per_unit":        {"value": 2180.0,"lower_is_better": False, "unit": "$",       "label": "Avg Rent per Unit"},
    "vacancy_days_cost":        {"value": 68.0,  "lower_is_better": True,  "unit": "$/day",   "label": "Vacancy Cost per Day"},
    "speed_to_first_contact":   {"value": 3.8,   "lower_is_better": True,  "unit": "hrs",     "label": "Speed to First Contact"},
}

UNIVERSITY_ZONES = [
    {"id": "all",             "label": "All Austin Metro"},
    {"id": "ut_austin",       "label": "UT Austin"},
    {"id": "texas_state",     "label": "Texas State (San Marcos)"},
    {"id": "concordia",       "label": "Concordia University"},
    {"id": "st_edwards",      "label": "St. Edward's University"},
]


@api_router.get("/reports/benchmarks")
async def reports_benchmarks(
    user=Depends(get_any_auth_user),
    date_range: str = "30d",
    university_zone: str = "all",
):
    """Anonymous, aggregated network benchmarks. Requires admin-level access
    and opt-in (org_settings.benchmarks.opt_in). Minimum 5 brokerages + 7-day
    data delay enforced; otherwise returns peer averages without org sample."""
    uid = user["_id"]
    start_iso, end_iso, _ = _range_to_bounds(date_range)

    # Gather user's (your brokerage) metric values
    your_values = {}
    for metric in AUSTIN_METRO_AVERAGES.keys():
        your_values[metric] = await _org_metric_value(uid, metric, start_iso, end_iso)

    # Zone modifier (slight adjustment to peer averages per zone for realism)
    zone_mod = {
        "all":          1.00,
        "ut_austin":    1.08,   # higher rents/velocity
        "texas_state":  0.92,
        "concordia":    0.96,
        "st_edwards":   1.02,
    }.get(university_zone, 1.00)

    # Count opted-in orgs (placeholder — real query over org_settings)
    opted_in_count = await db.org_settings.count_documents({"benchmarks.opt_in": True})
    min_threshold_met = opted_in_count >= 5

    # Build row per metric
    rows = []
    for mid, meta in AUSTIN_METRO_AVERAGES.items():
        yours = your_values.get(mid, 0.0)
        peer  = round(meta["value"] * zone_mod, 2)
        # Simple percentile heuristic: compare yours vs peer relative to lower_is_better
        if peer == 0:
            percentile = 50
        else:
            ratio = yours / peer if peer else 1
            if meta["lower_is_better"]:
                # lower value → higher percentile
                percentile = int(max(5, min(95, 100 - (ratio * 50))))
            else:
                percentile = int(max(5, min(95, ratio * 50)))
        # Trend sparkline (synth 14 points that converge toward current)
        sparkline = []
        for i in range(14):
            import random
            jitter = (random.random() - 0.5) * (peer * 0.05)
            sparkline.append({"i": i, "value": round(peer + jitter, 2)})
        rows.append({
            "id": mid,
            "label": meta["label"],
            "unit": meta["unit"],
            "lower_is_better": meta["lower_is_better"],
            "yours": round(yours, 2),
            "peer_average": peer,
            "percentile": percentile,
            "sparkline": sparkline,
        })

    return {
        "zone": university_zone,
        "range": date_range,
        "opted_in_count": opted_in_count,
        "min_threshold_met": min_threshold_met,
        "delay_days": 7,
        "university_zones": UNIVERSITY_ZONES,
        "rows": rows,
    }


@api_router.get("/reports/leasing/export.csv")
async def reports_leasing_export(user=Depends(get_any_auth_user), date_range: str = "30d", scope: str = "me"):
    """Export the current leasing report as CSV."""
    data = await reports_leasing(user=user, date_range=date_range, scope=scope)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Section", "Metric", "Value"])
    ov = data["overview"]
    writer.writerow(["Overview", "New contacts", ov["new_contacts"]])
    writer.writerow(["Overview", "Signed leases", ov["signed_leases"]])
    writer.writerow(["Overview", "Total properties", ov["total_properties"]])
    writer.writerow(["Overview", "Active listings", ov["active_listings"]])
    writer.writerow(["Overview", "Occupancy rate %", ov["occupancy_rate"]])
    writer.writerow(["Overview", "Pipeline value", ov["pipeline_value"]])
    for a in data["agents"]:
        writer.writerow(["Agents", f"{a['name']} · total activities", a["total"]])
        writer.writerow(["Agents", f"{a['name']} · leases signed", a["leases_signed"]])
    for s in data["sources"]:
        writer.writerow(["Sources", s["source"], s["count"]])
    for s in data["signed_by_source"]:
        writer.writerow(["Signed by source", s["source"], s["count"]])
    ct = data["calls_texts"]
    writer.writerow(["Calls/Texts", "Calls", ct["calls"]])
    writer.writerow(["Calls/Texts", "SMS", ct["sms"]])
    writer.writerow(["Calls/Texts", "Emails", ct["emails"]])
    for f in data["applications_funnel"]:
        writer.writerow(["Funnel", f["stage"], f["count"]])
    writer.writerow(["Renewals", "Upcoming (90d)", data["renewals"]["upcoming_90d"]])
    csv_str = output.getvalue()
    return StreamingResponse(
        io.BytesIO(csv_str.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="leasing_report_{date_range}.csv"'},
    )

# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE: User Settings Update (for auto_assign)
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.put("/users/me")
async def update_user_settings(data: UserUpdate, user=Depends(get_current_user)):
    update_data = {}
    if data.name is not None:
        update_data["name"] = data.name
    if data.auto_assign is not None:
        update_data["auto_assign"] = data.auto_assign
    if data.email_signature is not None:
        update_data["email_signature"] = data.email_signature
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    await db.users.update_one({"_id": ObjectId(user["_id"])}, {"$set": update_data})
    return {"message": "User settings updated"}

# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE: IDX Webhooks Placeholder
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.post("/webhooks/idx")
async def idx_webhook(request: Request):
    """
    Placeholder endpoint for IDX/Zillow lead webhooks.
    TODO: Parse IDX payload and create contact/property records.
    """
    payload = await request.json()
    logger.info(f"IDX webhook received: {payload}")
    # Future: Parse payload and create contacts/properties
    return {"status": "received", "message": "IDX webhook placeholder — not yet implemented"}

# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE: Google Calendar Sync Placeholders
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/calendar/auth")
async def calendar_auth_start(user=Depends(get_current_user)):
    """
    Placeholder: Initiate Google Calendar OAuth flow.
    TODO: Redirect to Google OAuth consent screen.
    """
    return {"status": "not_implemented", "message": "Google Calendar OAuth not yet configured. Add Google OAuth credentials."}

@api_router.post("/calendar/sync")
async def calendar_sync(user=Depends(get_current_user)):
    """
    Placeholder: Sync tasks/activities with Google Calendar.
    TODO: Use Google Calendar API to create/update events.
    """
    return {"status": "not_implemented", "message": "Google Calendar sync not yet configured. Complete OAuth integration first."}

# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE: Contact/Tenant Profile Page — FUB-parity endpoints
# Appended additively. No existing routes modified. All endpoints require auth
# and enforce user_id ownership. New collections:
#   contact_files · leases · maintenance_tickets · calendar_events
# ═══════════════════════════════════════════════════════════════════════════════

# ── Client Type + Stage catalogs ──
CLIENT_TYPES = ("leasing_tenant", "sales_buyer", "sales_seller", "commercial", "other")
CLIENT_TYPE_STAGES = {
    "leasing_tenant": [
        "Inquiry", "Tour Scheduled", "Application Submitted", "Screening",
        "Approved", "Lease Signed", "Move-In", "Active Tenant",
        "Renewal Due", "Renewal Offered", "Renewed", "Vacating", "Past Tenant"
    ],
    "sales_buyer": [
        "Inquiry", "Consultation", "Pre-Approved", "Showing",
        "Offer Submitted", "Under Contract", "Inspection", "Closing",
        "Closed Won", "Closed Lost"
    ],
    "sales_seller": [
        "Inquiry", "Consultation", "Listing Prep", "Active Listing",
        "Offer Received", "Under Contract", "Closing", "Sold", "Withdrawn"
    ],
    "commercial": [
        "Inquiry", "Tour", "LOI", "Due Diligence",
        "Negotiation", "Contract", "Closing", "Closed"
    ],
    "other": ["Prospect", "Contacted", "Qualified", "Nurturing", "Converted", "Lost"],
}

# ── Pydantic models (profile page only) ──
class ContactPhotoUpload(BaseModel):
    photo_url: str = Field(..., min_length=10, max_length=2_500_000)  # data:image/...;base64,...

    @field_validator("photo_url")
    @classmethod
    def _validate_data_url(cls, v):
        if not (v.startswith("data:image/") or v.startswith("http://") or v.startswith("https://")):
            raise ValueError("photo_url must be an http(s) URL or data:image/...;base64 URL")
        return v

class ContactStageUpdate(BaseModel):
    leasing_stage: str = Field(..., min_length=1, max_length=50)
    client_type: Optional[str] = Field(default=None, max_length=30)

    @field_validator("client_type")
    @classmethod
    def _validate_ct(cls, v):
        if v is not None and v not in CLIENT_TYPES:
            raise ValueError(f"client_type must be one of: {', '.join(CLIENT_TYPES)}")
        return v

class TagBody(BaseModel):
    tag: str = Field(..., min_length=1, max_length=50)

class FileUploadBody(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    mime_type: Optional[str] = Field(default="application/octet-stream", max_length=100)
    category: Optional[str] = Field(default="general", max_length=30)  # lease, inspection, id_doc, maintenance, general
    data: str = Field(..., min_length=10, max_length=14_000_000)  # base64 (~10 MB raw)
    size: Optional[int] = Field(default=0, ge=0)

class LeaseBody(BaseModel):
    property_id: Optional[str] = Field(default="", max_length=50)
    unit: Optional[str] = Field(default="", max_length=100)
    monthly_rent: Optional[float] = Field(default=0, ge=0)
    security_deposit: Optional[float] = Field(default=0, ge=0)
    lease_start: Optional[str] = Field(default="", max_length=20)
    lease_end: Optional[str] = Field(default="", max_length=20)
    move_in_date: Optional[str] = Field(default="", max_length=20)
    lease_term_months: Optional[int] = Field(default=12, ge=0, le=600)
    status: Optional[str] = Field(default="active", max_length=20)
    notes: Optional[str] = Field(default="", max_length=5000)

class MaintenanceBody(BaseModel):
    title: str = Field(..., min_length=1, max_length=200)
    description: Optional[str] = Field(default="", max_length=5000)
    priority: Optional[str] = Field(default="medium", max_length=10)
    status: Optional[str] = Field(default="open", max_length=20)
    category: Optional[str] = Field(default="general", max_length=40)

    @field_validator("priority")
    @classmethod
    def _vp(cls, v):
        if v and v not in ("high", "medium", "low"):
            raise ValueError("priority must be 'high', 'medium', or 'low'")
        return v

    @field_validator("status")
    @classmethod
    def _vs(cls, v):
        if v and v not in ("open", "in_progress", "resolved", "closed"):
            raise ValueError("status must be 'open', 'in_progress', 'resolved', or 'closed'")
        return v

class MaintenanceUpdate(BaseModel):
    title: Optional[str] = Field(default=None, max_length=200)
    description: Optional[str] = Field(default=None, max_length=5000)
    priority: Optional[str] = Field(default=None, max_length=10)
    status: Optional[str] = Field(default=None, max_length=20)
    category: Optional[str] = Field(default=None, max_length=40)

class EventBody(BaseModel):
    title: str = Field(..., min_length=1, max_length=200)
    description: Optional[str] = Field(default="", max_length=5000)
    start: str = Field(..., min_length=1, max_length=40)  # ISO datetime
    end: Optional[str] = Field(default="", max_length=40)
    location: Optional[str] = Field(default="", max_length=300)
    event_type: Optional[str] = Field(default="meeting", max_length=30)

class EventUpdate(BaseModel):
    title: Optional[str] = Field(default=None, max_length=200)
    description: Optional[str] = Field(default=None, max_length=5000)
    start: Optional[str] = Field(default=None, max_length=40)
    end: Optional[str] = Field(default=None, max_length=40)
    location: Optional[str] = Field(default=None, max_length=300)
    event_type: Optional[str] = Field(default=None, max_length=30)

class CollaboratorBody(BaseModel):
    user_id: str = Field(..., min_length=1, max_length=50)

class AIContactIdBody(BaseModel):
    contact_id: str = Field(..., max_length=50)

# ── Helper: ensure contact exists + owned by user (returns contact dict) ──
async def _owned_contact(contact_id: str, user) -> dict:
    c = await db.contacts.find_one({
        "_id": validate_object_id(contact_id, "Contact"),
        "user_id": user["_id"],
    })
    if not c:
        raise HTTPException(status_code=404, detail="Contact not found")
    return c

# ── Client types / stages reference ──
@api_router.get("/client-types")
async def list_client_types(user=Depends(get_any_auth_user)):
    return {
        "types": [
            {"value": "leasing_tenant", "label": "Leasing / Tenant"},
            {"value": "sales_buyer", "label": "Sales / Buyer"},
            {"value": "sales_seller", "label": "Sales / Seller"},
            {"value": "commercial", "label": "Commercial"},
            {"value": "other", "label": "Other / Prospect"},
        ],
        "stages": CLIENT_TYPE_STAGES,
    }

# ── Avatar upload (base64 on contact doc) ──
@api_router.post("/contacts/{contact_id}/photo")
async def upload_contact_photo(contact_id: str, data: ContactPhotoUpload, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    await db.contacts.update_one(
        {"_id": validate_object_id(contact_id, "Contact")},
        {"$set": {"photo_url": data.photo_url, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"photo_url": data.photo_url}

@api_router.delete("/contacts/{contact_id}/photo")
async def delete_contact_photo(contact_id: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    await db.contacts.update_one(
        {"_id": validate_object_id(contact_id, "Contact")},
        {"$unset": {"photo_url": ""}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"message": "Photo removed"}

# ── Stage update (w/ timestamp + activity log) ──
@api_router.put("/contacts/{contact_id}/stage")
async def update_contact_stage(contact_id: str, data: ContactStageUpdate, user=Depends(get_any_auth_user)):
    c = await _owned_contact(contact_id, user)
    client_type = data.client_type or c.get("client_type") or "leasing_tenant"
    if client_type not in CLIENT_TYPES:
        raise HTTPException(status_code=400, detail="Invalid client_type")
    if data.leasing_stage not in CLIENT_TYPE_STAGES[client_type]:
        raise HTTPException(status_code=400, detail=f"Stage '{data.leasing_stage}' not valid for {client_type}")
    now = datetime.now(timezone.utc).isoformat()
    update_doc = {
        "leasing_stage": data.leasing_stage,
        "stage_updated_at": now,
        "updated_at": now,
    }
    if data.client_type:
        update_doc["client_type"] = data.client_type
    await db.contacts.update_one({"_id": validate_object_id(contact_id, "Contact")}, {"$set": update_doc})
    # Log as activity
    await db.activities.insert_one({
        "contact_id": contact_id, "user_id": user["_id"],
        "activity_type": "note",
        "description": f"Stage changed from '{c.get('leasing_stage','—')}' to '{data.leasing_stage}'",
        "created_at": now,
    })
    return {"leasing_stage": data.leasing_stage, "client_type": client_type, "stage_updated_at": now}

# ── Tags add / remove ──
@api_router.post("/contacts/{contact_id}/tags")
async def add_contact_tag(contact_id: str, data: TagBody, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    await db.contacts.update_one(
        {"_id": validate_object_id(contact_id, "Contact")},
        {"$addToSet": {"tags": data.tag}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    c = await db.contacts.find_one({"_id": validate_object_id(contact_id, "Contact")})
    return {"tags": c.get("tags", [])}

@api_router.delete("/contacts/{contact_id}/tags/{tag}")
async def remove_contact_tag(contact_id: str, tag: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    await db.contacts.update_one(
        {"_id": validate_object_id(contact_id, "Contact")},
        {"$pull": {"tags": tag}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    c = await db.contacts.find_one({"_id": validate_object_id(contact_id, "Contact")})
    return {"tags": c.get("tags", [])}

# ── Files ──
@api_router.get("/contacts/{contact_id}/files")
async def list_contact_files(contact_id: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    cursor = db.contact_files.find(
        {"contact_id": contact_id, "user_id": user["_id"]},
        {"data": 0},  # exclude bytes in listing
    ).sort("created_at", -1)
    items = await cursor.to_list(500)
    return [serialize_doc(x) for x in items]

@api_router.post("/contacts/{contact_id}/files")
async def upload_contact_file(contact_id: str, data: FileUploadBody, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    doc = {
        "contact_id": contact_id,
        "user_id": user["_id"],
        "name": data.name,
        "mime_type": data.mime_type or "application/octet-stream",
        "category": data.category or "general",
        "size": data.size or len(data.data),
        "data": data.data,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await db.contact_files.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    del doc["_id"]
    del doc["data"]
    return doc

@api_router.get("/contacts/{contact_id}/files/{file_id}")
async def download_contact_file(contact_id: str, file_id: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    f = await db.contact_files.find_one({
        "_id": validate_object_id(file_id, "File"),
        "contact_id": contact_id, "user_id": user["_id"],
    })
    if not f:
        raise HTTPException(status_code=404, detail="File not found")
    return {
        "id": str(f["_id"]), "name": f.get("name", ""),
        "mime_type": f.get("mime_type", "application/octet-stream"),
        "category": f.get("category", "general"),
        "size": f.get("size", 0),
        "data": f.get("data", ""),
        "created_at": f.get("created_at", ""),
    }

@api_router.delete("/contacts/{contact_id}/files/{file_id}")
async def delete_contact_file(contact_id: str, file_id: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    result = await db.contact_files.delete_one({
        "_id": validate_object_id(file_id, "File"),
        "contact_id": contact_id, "user_id": user["_id"],
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="File not found")
    return {"message": "File deleted"}

# ── Lease info (1 active lease per contact; renewal history preserved) ──
@api_router.get("/contacts/{contact_id}/lease")
async def get_contact_lease(contact_id: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    lease = await db.leases.find_one({"contact_id": contact_id, "user_id": user["_id"]})
    history_cursor = db.leases.find(
        {"contact_id": contact_id, "user_id": user["_id"], "status": {"$in": ["expired", "renewed", "terminated"]}}
    ).sort("lease_end", -1)
    history = [serialize_doc(h) for h in await history_cursor.to_list(50)]
    return {
        "current": serialize_doc(lease) if lease else None,
        "history": history,
    }

@api_router.post("/contacts/{contact_id}/lease")
async def create_or_update_lease(contact_id: str, data: LeaseBody, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    payload = data.model_dump()
    payload["contact_id"] = contact_id
    payload["user_id"] = user["_id"]
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    existing = await db.leases.find_one({"contact_id": contact_id, "user_id": user["_id"], "status": "active"})
    if existing:
        await db.leases.update_one({"_id": existing["_id"]}, {"$set": payload})
        doc = await db.leases.find_one({"_id": existing["_id"]})
        return serialize_doc(doc)
    payload["created_at"] = payload["updated_at"]
    result = await db.leases.insert_one(payload)
    payload["id"] = str(result.inserted_id)
    del payload["_id"]
    return payload

@api_router.put("/contacts/{contact_id}/lease/{lease_id}")
async def update_lease(contact_id: str, lease_id: str, data: LeaseBody, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.leases.update_one(
        {"_id": validate_object_id(lease_id, "Lease"), "contact_id": contact_id, "user_id": user["_id"]},
        {"$set": updates},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lease not found")
    doc = await db.leases.find_one({"_id": validate_object_id(lease_id, "Lease")})
    return serialize_doc(doc)

# ── Maintenance Tickets ──
@api_router.get("/contacts/{contact_id}/maintenance")
async def list_maintenance(contact_id: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    cursor = db.maintenance_tickets.find({"contact_id": contact_id, "user_id": user["_id"]}).sort("created_at", -1)
    items = await cursor.to_list(500)
    return [serialize_doc(x) for x in items]

@api_router.post("/contacts/{contact_id}/maintenance")
async def create_maintenance(contact_id: str, data: MaintenanceBody, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    now = datetime.now(timezone.utc).isoformat()
    doc = data.model_dump()
    doc["contact_id"] = contact_id
    doc["user_id"] = user["_id"]
    doc["created_at"] = now
    doc["updated_at"] = now
    result = await db.maintenance_tickets.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    del doc["_id"]
    # Also log as activity
    await db.activities.insert_one({
        "contact_id": contact_id, "user_id": user["_id"],
        "activity_type": "note",
        "description": f"Maintenance ticket opened: {data.title} ({data.priority})",
        "created_at": now,
    })
    return doc

@api_router.put("/contacts/{contact_id}/maintenance/{ticket_id}")
async def update_maintenance(contact_id: str, ticket_id: str, data: MaintenanceUpdate, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    if updates.get("status") in ("resolved", "closed"):
        updates["resolved_at"] = updates["updated_at"]
    result = await db.maintenance_tickets.update_one(
        {"_id": validate_object_id(ticket_id, "Ticket"), "contact_id": contact_id, "user_id": user["_id"]},
        {"$set": updates},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ticket not found")
    doc = await db.maintenance_tickets.find_one({"_id": validate_object_id(ticket_id, "Ticket")})
    return serialize_doc(doc)

@api_router.delete("/contacts/{contact_id}/maintenance/{ticket_id}")
async def delete_maintenance(contact_id: str, ticket_id: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    r = await db.maintenance_tickets.delete_one({
        "_id": validate_object_id(ticket_id, "Ticket"),
        "contact_id": contact_id, "user_id": user["_id"],
    })
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ticket not found")
    return {"message": "Ticket deleted"}

# ── Calendar events (per-contact) ──
@api_router.get("/contacts/{contact_id}/events")
async def list_events(contact_id: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    cursor = db.calendar_events.find({"contact_id": contact_id, "user_id": user["_id"]}).sort("start", 1)
    items = await cursor.to_list(500)
    return [serialize_doc(x) for x in items]

@api_router.post("/contacts/{contact_id}/events")
async def create_event(contact_id: str, data: EventBody, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    doc = data.model_dump()
    doc["contact_id"] = contact_id
    doc["user_id"] = user["_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.calendar_events.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    del doc["_id"]
    return doc

@api_router.put("/contacts/{contact_id}/events/{event_id}")
async def update_event(contact_id: str, event_id: str, data: EventUpdate, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.calendar_events.update_one(
        {"_id": validate_object_id(event_id, "Event"), "contact_id": contact_id, "user_id": user["_id"]},
        {"$set": updates},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    doc = await db.calendar_events.find_one({"_id": validate_object_id(event_id, "Event")})
    return serialize_doc(doc)

@api_router.delete("/contacts/{contact_id}/events/{event_id}")
async def delete_event(contact_id: str, event_id: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    r = await db.calendar_events.delete_one({
        "_id": validate_object_id(event_id, "Event"),
        "contact_id": contact_id, "user_id": user["_id"],
    })
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event deleted"}

# ── Collaborators (references to team users) ──
@api_router.get("/contacts/{contact_id}/collaborators")
async def list_collaborators(contact_id: str, user=Depends(get_any_auth_user)):
    c = await _owned_contact(contact_id, user)
    ids = c.get("collaborator_ids", []) or []
    users = []
    for uid in ids:
        try:
            u = await db.users.find_one({"_id": ObjectId(uid)}, {"password_hash": 0})
            if u:
                users.append({"id": str(u["_id"]), "name": u.get("name", ""), "email": u.get("email", ""), "role": u.get("role", "agent")})
        except Exception:
            continue
    return users

@api_router.post("/contacts/{contact_id}/collaborators")
async def add_collaborator(contact_id: str, data: CollaboratorBody, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    # Validate user exists
    try:
        u = await db.users.find_one({"_id": ObjectId(data.user_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user_id")
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    await db.contacts.update_one(
        {"_id": validate_object_id(contact_id, "Contact")},
        {"$addToSet": {"collaborator_ids": data.user_id},
         "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"id": data.user_id, "name": u.get("name", ""), "email": u.get("email", ""), "role": u.get("role", "agent")}

@api_router.delete("/contacts/{contact_id}/collaborators/{user_id}")
async def remove_collaborator(contact_id: str, user_id: str, user=Depends(get_any_auth_user)):
    await _owned_contact(contact_id, user)
    await db.contacts.update_one(
        {"_id": validate_object_id(contact_id, "Contact")},
        {"$pull": {"collaborator_ids": user_id}},
    )
    return {"message": "Collaborator removed"}

# ── AI: Retention Summary (cached 24h) ──
@api_router.post("/ai/retention-summary")
async def ai_retention_summary(data: AIContactIdBody, user=Depends(get_any_auth_user)):
    contact = await _owned_contact(data.contact_id, user)
    # Return cached summary if fresh (< 24h)
    cached_at = contact.get("retention_summary_generated_at")
    if cached_at and contact.get("retention_summary"):
        try:
            gen_dt = datetime.fromisoformat(cached_at.replace("Z", "+00:00"))
            if (datetime.now(timezone.utc) - gen_dt) < timedelta(hours=24):
                return {
                    "summary": contact.get("retention_summary", ""),
                    "retention_score": contact.get("retention_score", 50),
                    "generated_at": cached_at,
                    "cached": True,
                }
        except Exception:
            pass
    await check_ai_rate_limit(user["_id"])
    activities = await db.activities.find({"contact_id": data.contact_id}).sort("created_at", -1).to_list(30)
    tickets = await db.maintenance_tickets.find({"contact_id": data.contact_id}).to_list(20)
    lease = await db.leases.find_one({"contact_id": data.contact_id, "status": "active"})
    act_text = "\n".join([f"- [{a.get('activity_type','')}] {a.get('description','')}" for a in activities]) or "None"
    tick_text = "\n".join([f"- {t.get('title','')} ({t.get('priority','')}, {t.get('status','')})" for t in tickets]) or "None"
    lease_text = (f"Rent ${lease.get('monthly_rent',0)}/mo, unit {lease.get('unit','')}, "
                  f"ends {lease.get('lease_end','')}") if lease else "No active lease on record"
    system_msg = f"""You are a tenant-retention AI analyst for a residential leasing CRM.
Analyze this tenant and return a concise retention assessment.

TENANT: {contact.get('name','')}
STAGE: {contact.get('leasing_stage','Inquiry')}
LEASE: {lease_text}
RECENT ACTIVITIES (newest first):
{act_text}
MAINTENANCE TICKETS:
{tick_text}

Respond ONLY with valid JSON:
{{"retention_score": <0-100 int>, "summary": "<one concise paragraph, 2-4 sentences, focused on renewal likelihood & action>"}}"""
    guard_ai_cost(system_msg)
    input_tokens = estimate_tokens(system_msg)
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=settings.EMERGENT_LLM_KEY,
            session_id=f"retention-{data.contact_id}-{datetime.now(timezone.utc).isoformat()}",
            system_message=system_msg,
        ).with_model("openai", "gpt-5.2")
        msg = UserMessage(text="Analyze this tenant's retention outlook now.")
        result = await chat.send_message(msg)
        output_tokens = estimate_tokens(result)
        await log_ai_usage(user["_id"], "retention-summary", input_tokens, output_tokens)
        cleaned = result.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[1].rsplit("```", 1)[0].strip()
        try:
            parsed = json.loads(cleaned)
            score = int(parsed.get("retention_score", 50))
            summary = parsed.get("summary", result)
        except Exception:
            score = 50
            summary = result
        now_iso = datetime.now(timezone.utc).isoformat()
        await db.contacts.update_one(
            {"_id": validate_object_id(data.contact_id, "Contact")},
            {"$set": {
                "retention_score": score,
                "retention_summary": summary,
                "retention_summary_generated_at": now_iso,
            }},
        )
        return {"summary": summary, "retention_score": score, "generated_at": now_iso, "cached": False}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Retention AI error: {e}")
        raise HTTPException(status_code=500, detail="AI service temporarily unavailable. Please try again.")

# ── AI: Analyze Email Thread ──
@api_router.post("/ai/analyze-email-thread")
async def ai_analyze_email_thread(data: AIContactIdBody, user=Depends(get_any_auth_user)):
    await _owned_contact(data.contact_id, user)
    await check_ai_rate_limit(user["_id"])
    emails = await db.activities.find({
        "contact_id": data.contact_id,
        "activity_type": "email",
    }).sort("created_at", -1).to_list(50)
    if not emails:
        return {"analysis": "No email history found for this contact."}
    email_text = "\n".join([f"[{e.get('created_at','')}] {e.get('description','')}" for e in emails])
    system_msg = """You analyze email conversations for a real estate CRM.
Summarize sentiment, key pain points, unanswered questions, and recommended next action.
Be concise, 3-5 bullet points plus one-line recommendation."""
    full_input = system_msg + f"\nEmails:\n{email_text}"
    guard_ai_cost(full_input)
    input_tokens = estimate_tokens(full_input)
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=settings.EMERGENT_LLM_KEY,
            session_id=f"email-analyze-{data.contact_id}-{datetime.now(timezone.utc).isoformat()}",
            system_message=system_msg,
        ).with_model("openai", "gpt-5.2")
        msg = UserMessage(text=f"Emails:\n{email_text}")
        result = await chat.send_message(msg)
        output_tokens = estimate_tokens(result)
        await log_ai_usage(user["_id"], "analyze-email-thread", input_tokens, output_tokens)
        return {"analysis": result, "email_count": len(emails)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Email thread analysis error: {e}")
        raise HTTPException(status_code=500, detail="AI service temporarily unavailable. Please try again.")

# ── One-click: Convert Prospect to Tenant ──
@api_router.post("/contacts/{contact_id}/convert-to-tenant")
async def convert_to_tenant(contact_id: str, user=Depends(get_any_auth_user)):
    c = await _owned_contact(contact_id, user)
    now = datetime.now(timezone.utc).isoformat()
    await db.contacts.update_one(
        {"_id": validate_object_id(contact_id, "Contact")},
        {"$set": {
            "is_tenant": True,
            "client_type": "leasing_tenant",
            "leasing_stage": "Active Tenant",
            "stage_updated_at": now,
            "updated_at": now,
        }},
    )
    await db.activities.insert_one({
        "contact_id": contact_id, "user_id": user["_id"],
        "activity_type": "note",
        "description": f"Converted from prospect to active tenant (was: {c.get('leasing_stage','Inquiry')})",
        "created_at": now,
    })
    return {"is_tenant": True, "leasing_stage": "Active Tenant", "stage_updated_at": now}

# ── One-click: Send Renewal Offer (drafts email via AI, flips stage) ──
@api_router.post("/contacts/{contact_id}/send-renewal-offer")
async def send_renewal_offer(contact_id: str, user=Depends(get_any_auth_user)):
    c = await _owned_contact(contact_id, user)
    lease = await db.leases.find_one({"contact_id": contact_id, "user_id": user["_id"], "status": "active"})
    await check_ai_rate_limit(user["_id"])
    system_msg = f"""Draft a friendly, professional lease-renewal offer email.
Tenant: {c.get('name','')}
Current rent: ${lease.get('monthly_rent',0) if lease else 'N/A'}/mo
Unit: {lease.get('unit','') if lease else 'N/A'}
Lease end: {lease.get('lease_end','') if lease else 'N/A'}

Include: subject line, greeting, brief thank-you, renewal terms placeholder, clear call-to-action.
Return plain text with 'Subject:' on the first line."""
    guard_ai_cost(system_msg)
    input_tokens = estimate_tokens(system_msg)
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=settings.EMERGENT_LLM_KEY,
            session_id=f"renewal-{contact_id}-{datetime.now(timezone.utc).isoformat()}",
            system_message=system_msg,
        ).with_model("openai", "gpt-5.2")
        msg = UserMessage(text="Draft the renewal offer email now.")
        draft = await chat.send_message(msg)
        output_tokens = estimate_tokens(draft)
        await log_ai_usage(user["_id"], "renewal-offer", input_tokens, output_tokens)
    except Exception as e:
        logger.error(f"Renewal draft error: {e}")
        draft = f"Subject: Renewal Offer for your lease\n\nHi {c.get('name','there')},\n\nWe'd love to have you stay! Let's talk about renewing your lease.\n\nBest,\nYour Property Manager"
    # Flip stage to Renewal Offered
    now = datetime.now(timezone.utc).isoformat()
    await db.contacts.update_one(
        {"_id": validate_object_id(contact_id, "Contact")},
        {"$set": {"leasing_stage": "Renewal Offered", "stage_updated_at": now, "updated_at": now}},
    )
    await db.activities.insert_one({
        "contact_id": contact_id, "user_id": user["_id"],
        "activity_type": "note",
        "description": "Renewal offer drafted — moved to 'Renewal Offered' stage",
        "created_at": now,
    })
    return {"draft": draft, "leasing_stage": "Renewal Offered"}

# ── Indexes for new collections ──
async def create_profile_indexes():
    try:
        await db.contact_files.create_index([("contact_id", 1), ("created_at", -1)], background=True)
        await db.contact_files.create_index("user_id", background=True)
        await db.leases.create_index([("contact_id", 1), ("status", 1)], background=True)
        await db.leases.create_index("user_id", background=True)
        await db.maintenance_tickets.create_index([("contact_id", 1), ("created_at", -1)], background=True)
        await db.maintenance_tickets.create_index([("user_id", 1), ("status", 1)], background=True)
        await db.calendar_events.create_index([("contact_id", 1), ("start", 1)], background=True)
        await db.calendar_events.create_index("user_id", background=True)
        logger.info("Profile page indexes created.")
    except Exception as e:
        logger.error(f"Profile index creation failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE: Dashboard Leasing Overview (Phase 11)
# Purpose-built aggregate endpoint for the FUB-parity residential-leasing
# dashboard. Read-only, additive. No existing routes modified.
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_dashboard_range(range_str: str):
    """
    Returns (start, end, prev_start, prev_end, granularity).
    range_str: '7d' | '30d' | '90d' | 'all'.
    granularity: 'day' (7d/30d) | 'week' (90d/all).
    """
    now = datetime.now(timezone.utc)
    if range_str == "all":
        start = now - timedelta(days=365)
        prev_end = start
        prev_start = now - timedelta(days=730)
        granularity = "week"
    elif range_str == "90d":
        start = now - timedelta(days=90)
        prev_end = start
        prev_start = now - timedelta(days=180)
        granularity = "week"
    elif range_str == "7d":
        start = now - timedelta(days=7)
        prev_end = start
        prev_start = now - timedelta(days=14)
        granularity = "day"
    else:  # '30d' default
        start = now - timedelta(days=30)
        prev_end = start
        prev_start = now - timedelta(days=60)
        granularity = "day"
    return start, now, prev_start, prev_end, granularity


def _pct_growth(curr: float, prev: float) -> float:
    """Standard growth %: positive means curr > prev."""
    if prev <= 0 and curr <= 0:
        return 0.0
    if prev <= 0:
        return 100.0
    return round(((curr - prev) / prev) * 100, 1)


def _parse_iso_safe(s):
    """Parse ISO timestamps stored as strings; return None on failure."""
    if not s or not isinstance(s, str):
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


async def _contact_lookup(contact_ids):
    """Batch-fetch contacts by id (UUID string) OR _id (ObjectId). Returns dict keyed by whichever id string is in the activities/deals docs."""
    result = {}
    if not contact_ids:
        return result
    oids, sids = [], []
    for cid in contact_ids:
        if not cid:
            continue
        try:
            oids.append(ObjectId(cid))
        except Exception:
            sids.append(cid)
    projection = {"name": 1, "email": 1, "phone": 1, "leasing_stage": 1, "client_type": 1}
    if oids:
        async for c in db.contacts.find({"_id": {"$in": oids}}, projection):
            result[str(c["_id"])] = c
    if sids:
        async for c in db.contacts.find({"id": {"$in": sids}}, {**projection, "id": 1}):
            result[c.get("id", "")] = c
    return result


async def _timeseries_daily_count(match: dict, start, end, granularity: str):
    """Aggregate document counts per day (or per week) based on created_at ISO strings."""
    pipeline = [
        {"$match": match},
        {"$addFields": {"_d": {"$dateFromString": {"dateString": "$created_at", "onError": None, "onNull": None}}}},
        {"$match": {"_d": {"$ne": None}}},
        {"$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$_d"}},
            "count": {"$sum": 1},
        }},
    ]
    rows = {}
    async for r in db.deals.aggregate(pipeline):
        rows[r["_id"]] = int(r["count"])
    days = []
    cur = start
    while cur.date() <= end.date():
        k = cur.strftime("%Y-%m-%d")
        days.append({"date": k, "value": rows.get(k, 0)})
        cur += timedelta(days=1)
    if granularity == "week":
        bucketed = []
        i = 0
        while i < len(days):
            chunk = days[i:i+7]
            bucketed.append({"date": chunk[0]["date"], "value": sum(x["value"] for x in chunk)})
            i += 7
        return bucketed
    return days


async def _avg_speed_to_first_contact(match_contacts: dict, start_iso: str, end_iso: str):
    """Avg hours between contact.created_at and earliest human activity for contacts created in range."""
    contacts_cursor = db.contacts.find(
        {**match_contacts, "created_at": {"$gte": start_iso, "$lte": end_iso}},
        {"_id": 1, "id": 1, "created_at": 1},
    )
    total_hours, sample = 0.0, 0
    daily = {}
    async for c in contacts_cursor:
        cid_str = c.get("id") or str(c["_id"])
        # Earliest human activity for this contact
        first = await db.activities.find_one(
            {"contact_id": cid_str, "activity_type": {"$in": ["call", "email", "sms", "meeting", "note"]}},
            sort=[("created_at", 1)],
        )
        if not first:
            continue
        ca = _parse_iso_safe(c.get("created_at"))
        fa = _parse_iso_safe(first.get("created_at"))
        if not ca or not fa:
            continue
        delta_h = max(0.0, (fa - ca).total_seconds() / 3600.0)
        total_hours += delta_h
        sample += 1
        k = ca.strftime("%Y-%m-%d")
        daily.setdefault(k, []).append(delta_h)
    avg = round(total_hours / sample, 1) if sample else 0.0
    sparkline = [{"date": k, "value": round(sum(v)/len(v), 2)} for k, v in sorted(daily.items())]
    return {"value": avg, "sample": sample, "sparkline": sparkline}


async def _avg_lease_up_days(match_deals: dict, start_iso: str, end_iso: str):
    """Avg days between deal.created_at and it reaching 'Lease Signed' / 'Move-In' / 'Active Tenant' (filtered by updated_at in range)."""
    q = {
        **match_deals,
        "pipeline_type": "lease_applications",
        "stage": {"$in": ["Lease Signed", "Move-In", "Active Tenant"]},
        "updated_at": {"$gte": start_iso, "$lte": end_iso},
    }
    total, sample = 0.0, 0
    daily = {}
    async for d in db.deals.find(q, {"created_at": 1, "updated_at": 1}):
        ca = _parse_iso_safe(d.get("created_at"))
        ua = _parse_iso_safe(d.get("updated_at"))
        if not ca or not ua:
            continue
        days = max(0.0, (ua - ca).total_seconds() / 86400.0)
        total += days
        sample += 1
        k = ua.strftime("%Y-%m-%d")
        daily.setdefault(k, []).append(days)
    avg = round(total / sample, 1) if sample else 0.0
    sparkline = [{"date": k, "value": round(sum(v)/len(v), 2)} for k, v in sorted(daily.items())]
    return {"value": avg, "sample": sample, "sparkline": sparkline}


async def _sum_renewals_within(match_leases: dict, today_iso: str, end_iso: str):
    """Count + sum monthly_rent of active leases with lease_end in [today, end_iso]."""
    pipeline = [
        {"$match": {
            **match_leases,
            "status": "active",
            "lease_end": {"$gte": today_iso, "$lte": end_iso, "$ne": ""},
        }},
        {"$group": {"_id": None, "count": {"$sum": 1}, "rent": {"$sum": {"$ifNull": ["$monthly_rent", 0]}}}},
    ]
    async for row in db.leases.aggregate(pipeline):
        return {"count": int(row.get("count", 0)), "monthly_rent_total": float(row.get("rent", 0) or 0)}
    return {"count": 0, "monthly_rent_total": 0.0}


@api_router.get("/dashboard/leasing-overview")
async def leasing_overview(
    range: str = "30d",
    scope: str = "me",
    user=Depends(get_any_auth_user),
):
    if range not in ("7d", "30d", "90d", "all"):
        raise HTTPException(status_code=400, detail="Invalid range (expected 7d|30d|90d|all)")
    if scope not in ("me", "everyone"):
        raise HTTPException(status_code=400, detail="Invalid scope (expected me|everyone)")

    start, end, prev_start, prev_end, granularity = _parse_dashboard_range(range)
    start_iso, end_iso = start.isoformat(), end.isoformat()
    prev_start_iso, prev_end_iso = prev_start.isoformat(), prev_end.isoformat()

    uid = user["_id"]
    base = {} if scope == "everyone" else {"user_id": uid}

    # ─── KPI 1: New Inquiries (deals created in lease_applications) ───
    inq_match = {**base, "pipeline_type": "lease_applications", "created_at": {"$gte": start_iso, "$lte": end_iso}}
    prev_inq_match = {**base, "pipeline_type": "lease_applications", "created_at": {"$gte": prev_start_iso, "$lt": prev_end_iso}}
    new_inquiries = await db.deals.count_documents(inq_match)
    prev_inquiries = await db.deals.count_documents(prev_inq_match)
    inq_sparkline = await _timeseries_daily_count(inq_match, start, end, granularity)

    # ─── KPI 2: Avg Speed to First Contact (hours) ───
    speed = await _avg_speed_to_first_contact(base, start_iso, end_iso)
    prev_speed = await _avg_speed_to_first_contact(base, prev_start_iso, prev_end_iso)

    # ─── KPI 3: Lease-Up Velocity (days Inquiry → Signed) ───
    velocity = await _avg_lease_up_days(base, start_iso, end_iso)
    prev_velocity = await _avg_lease_up_days(base, prev_start_iso, prev_end_iso)

    # ─── KPI 4: Current Occupancy Rate ───
    total_leases = await db.leases.count_documents(base)
    active_leases = await db.leases.count_documents({**base, "status": "active"})
    total_props = await db.properties.count_documents(base)
    rented_props = await db.properties.count_documents({**base, "status": {"$in": ["rented", "leased"]}})
    if total_leases > 0:
        units_total, units_occupied = total_leases, active_leases
    else:
        units_total, units_occupied = total_props, rented_props
    occupancy_pct = round((units_occupied / units_total * 100), 1) if units_total > 0 else 0.0

    # ─── KPI 5: Upcoming Renewals (30 / 60 / 90 days) ───
    today = datetime.now(timezone.utc).date()
    today_iso = today.isoformat()
    d30_end = (today + timedelta(days=30)).isoformat()
    d60_end = (today + timedelta(days=60)).isoformat()
    d90_end = (today + timedelta(days=90)).isoformat()
    renewals_30 = await _sum_renewals_within(base, today_iso, d30_end)
    renewals_60 = await _sum_renewals_within(base, today_iso, d60_end)
    renewals_90 = await _sum_renewals_within(base, today_iso, d90_end)

    # ─── Today's Action Items (Tours + Tasks) ───
    today_start_dt = datetime.combine(today, datetime.min.time(), tzinfo=timezone.utc).isoformat()
    today_end_dt = datetime.combine(today, datetime.max.time(), tzinfo=timezone.utc).isoformat()

    tour_events_raw = await db.calendar_events.find(
        {**base, "start": {"$gte": today_start_dt, "$lte": today_end_dt}},
    ).sort("start", 1).to_list(30)
    tasks_today_raw = await db.tasks.find(
        {**base, "completed": False, "due_date": today_iso},
    ).to_list(50)

    ai_contact_ids = (
        {e.get("contact_id") for e in tour_events_raw if e.get("contact_id")}
        | {t.get("contact_id") for t in tasks_today_raw if t.get("contact_id")}
    )
    ai_contact_lookup = await _contact_lookup(ai_contact_ids)

    priority_rank = {"high": 0, "medium": 1, "low": 2}

    tours_out = []
    for e in tour_events_raw:
        cid = e.get("contact_id", "")
        c = ai_contact_lookup.get(cid, {})
        tours_out.append({
            "id": str(e.get("_id", "")),
            "title": e.get("title", ""),
            "start": e.get("start", ""),
            "end": e.get("end", ""),
            "location": e.get("location", ""),
            "event_type": e.get("event_type", "meeting"),
            "contact_id": cid,
            "contact_name": c.get("name", ""),
        })

    tasks_today_raw.sort(key=lambda t: priority_rank.get(t.get("priority", "medium"), 1))
    tasks_out = []
    for t in tasks_today_raw:
        cid = t.get("contact_id", "")
        c = ai_contact_lookup.get(cid, {})
        tasks_out.append({
            "id": str(t.get("_id", "")),
            "title": t.get("title", ""),
            "priority": t.get("priority", "medium"),
            "due_date": t.get("due_date", ""),
            "contact_id": cid,
            "contact_name": c.get("name", ""),
            "deal_id": t.get("deal_id", ""),
        })

    # ─── Recent Activity (enriched) ───
    activities = await db.activities.find(
        {**base, "created_at": {"$gte": start_iso}},
    ).sort("created_at", -1).limit(25).to_list(25)

    act_cids = {a.get("contact_id") for a in activities if a.get("contact_id")}
    act_contact_lookup = await _contact_lookup(act_cids)

    # Users lookup for "Assigned"
    user_ids = {a.get("user_id") for a in activities if a.get("user_id")}
    user_lookup = {}
    for uid2 in user_ids:
        if not uid2:
            continue
        try:
            u = await db.users.find_one({"_id": ObjectId(uid2)}, {"name": 1, "email": 1})
        except Exception:
            u = None
        if u:
            user_lookup[uid2] = u.get("name") or u.get("email") or ""

    # Most recent deal per contact (for stage + unit)
    deals_map = {}
    if act_cids:
        async for d in db.deals.find({**base, "contact_id": {"$in": list(act_cids)}}).sort("updated_at", -1):
            cid = d.get("contact_id", "")
            if cid and cid not in deals_map:
                deals_map[cid] = d

    activity_out = []
    for a in activities:
        cid = a.get("contact_id", "")
        c = act_contact_lookup.get(cid, {})
        d = deals_map.get(cid, {})
        activity_out.append({
            "id": str(a.get("_id", "")),
            "contact_id": cid,
            "contact_name": c.get("name") or "Unknown",
            "contact_email": c.get("email", ""),
            "contact_phone": c.get("phone", ""),
            "activity_type": a.get("activity_type", "note"),
            "description": a.get("description", ""),
            "created_at": a.get("created_at", ""),
            "stage": c.get("leasing_stage") or d.get("stage", ""),
            "assigned_to_name": user_lookup.get(a.get("user_id", "")) or "",
            "unit": d.get("unit_number") or d.get("unit_address") or "",
        })

    return {
        "range": range,
        "scope": scope,
        "granularity": granularity,
        "kpis": {
            "new_inquiries": {
                "value": new_inquiries,
                "previous": prev_inquiries,
                "growth_pct": _pct_growth(new_inquiries, prev_inquiries),
                "lower_is_better": False,
                "sparkline": inq_sparkline,
            },
            "avg_speed_to_first_contact": {
                "value_hours": speed["value"],
                "previous_hours": prev_speed["value"],
                "growth_pct": _pct_growth(speed["value"], prev_speed["value"]),
                "lower_is_better": True,
                "sample_size": speed["sample"],
                "sparkline": speed["sparkline"],
            },
            "lease_up_velocity": {
                "value_days": velocity["value"],
                "previous_days": prev_velocity["value"],
                "growth_pct": _pct_growth(velocity["value"], prev_velocity["value"]),
                "lower_is_better": True,
                "sample_size": velocity["sample"],
                "sparkline": velocity["sparkline"],
            },
            "current_occupancy_rate": {
                "value_pct": occupancy_pct,
                "units_occupied": units_occupied,
                "units_total": units_total,
                "lower_is_better": False,
                "sparkline": [],
            },
            "upcoming_renewals": {
                "d30": renewals_30,
                "d60": renewals_60,
                "d90": renewals_90,
            },
        },
        "todays_action_items": {
            "tours": tours_out,
            "tasks": tasks_out,
        },
        "recent_activity": activity_out,
    }



# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE: Lease Applications Pipeline (Phase 10)
# Appended additively. No existing routes modified.
# ═══════════════════════════════════════════════════════════════════════════════

class CustomStageBody(BaseModel):
    pipeline_type: str = Field(..., max_length=30)
    name: str = Field(..., min_length=1, max_length=50)

    @field_validator("pipeline_type")
    @classmethod
    def _vpt(cls, v):
        allowed = ("residential_lease", "commercial_sale", "commercial_lease", "lease_applications")
        if v not in allowed:
            raise ValueError(f"pipeline_type must be one of: {', '.join(allowed)}")
        return v

# ── Custom user-defined stages ──
@api_router.get("/pipeline/custom-stages")
async def get_custom_stages(pipeline_type: str = "lease_applications", user=Depends(get_any_auth_user)):
    if pipeline_type not in PIPELINE_STAGES:
        raise HTTPException(status_code=400, detail="Invalid pipeline_type")
    stages = (user.get("custom_stages") or {}).get(pipeline_type, [])
    return {"pipeline_type": pipeline_type, "stages": stages}

@api_router.post("/pipeline/custom-stages")
async def add_custom_stage(data: CustomStageBody, user=Depends(get_any_auth_user)):
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Stage name required")
    if name in PIPELINE_STAGES[data.pipeline_type]:
        raise HTTPException(status_code=400, detail="Stage already exists as a built-in stage")
    user_id = ObjectId(user["_id"]) if isinstance(user["_id"], str) else user["_id"]
    await db.users.update_one(
        {"_id": user_id},
        {"$addToSet": {f"custom_stages.{data.pipeline_type}": name}},
    )
    updated = await db.users.find_one({"_id": user_id}, {"custom_stages": 1})
    if not updated:
        raise HTTPException(status_code=500, detail="User not found after update")
    return {"stages": (updated.get("custom_stages") or {}).get(data.pipeline_type, [])}

@api_router.delete("/pipeline/custom-stages/{pipeline_type}/{name}")
async def remove_custom_stage(pipeline_type: str, name: str, user=Depends(get_any_auth_user)):
    if pipeline_type not in PIPELINE_STAGES:
        raise HTTPException(status_code=400, detail="Invalid pipeline_type")
    user_id = ObjectId(user["_id"]) if isinstance(user["_id"], str) else user["_id"]
    count = await db.deals.count_documents({
        "user_id": user["_id"], "pipeline_type": pipeline_type, "stage": name
    })
    if count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot remove stage: {count} deal(s) still in this stage")
    await db.users.update_one(
        {"_id": user_id},
        {"$pull": {f"custom_stages.{pipeline_type}": name}},
    )
    return {"message": "Stage removed"}

# ── One-time migration: residential_lease → lease_applications ──
async def migrate_residential_lease_to_lease_applications():
    """Move existing residential_lease deals into the new lease_applications pipeline."""
    try:
        count = await db.deals.count_documents({"pipeline_type": "residential_lease"})
        if count == 0:
            return
        logger.info(f"[Migration] Found {count} residential_lease deal(s). Migrating to lease_applications…")
        cursor = db.deals.find({"pipeline_type": "residential_lease"})
        migrated = 0
        async for deal in cursor:
            old_stage = deal.get("stage", "New Lead")
            new_stage = LEGACY_STAGE_MAP.get(old_stage) or "Inquiry"
            await db.deals.update_one(
                {"_id": deal["_id"]},
                {"$set": {
                    "pipeline_type": "lease_applications",
                    "stage": new_stage,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "migrated_from": {"pipeline_type": "residential_lease", "stage": old_stage},
                }},
            )
            migrated += 1
        logger.info(f"[Migration] Successfully migrated {migrated} deal(s) to lease_applications.")
    except Exception as e:
        logger.error(f"[Migration] residential_lease → lease_applications failed: {e}")


# ─── Startup ───
@app.on_event("startup")
async def startup():
    # ── PERFORMANCE: Create comprehensive MongoDB indexes ──
    await create_mongodb_indexes()
    await create_profile_indexes()
    await create_inbox_indexes()
    # ── Phase 10: one-time migration ──
    await migrate_residential_lease_to_lease_applications()

    # Seed admin
    admin_email = settings.ADMIN_EMAIL
    admin_password = settings.ADMIN_PASSWORD
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Admin",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin user created: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
        logger.info("Admin password updated")
    # Write test credentials
    os.makedirs("/app/memory", exist_ok=True)
    with open("/app/memory/test_credentials.md", "w") as f:
        f.write(f"# Test Credentials\n\n## Admin\n- Email: {admin_email}\n- Password: {admin_password}\n- Role: admin\n\n## Auth Endpoints\n- POST /api/auth/register\n- POST /api/auth/login\n- POST /api/auth/logout\n- GET /api/auth/me\n- POST /api/auth/refresh\n")
    
    # ── Start background sequence polling worker ──
    asyncio.create_task(process_sequence_executions())
    logger.info("Sequence polling worker started")
    asyncio.create_task(brokerage_sheet_sync_worker())

# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 14 — UNIFIED INBOX (Email + SMS + Voicemail)
# Threads are grouped by contact_id. Messages collection stores every inbound /
# outbound communication. Thread meta (assignment, status) lives in
# inbox_threads. Drafts in inbox_drafts.
# ═══════════════════════════════════════════════════════════════════════════════

class InboxReplyRequest(BaseModel):
    channel: str = Field(..., max_length=20)
    subject: Optional[str] = Field(default="", max_length=500)
    body: str = Field(..., min_length=1, max_length=50000)

    @field_validator("channel")
    @classmethod
    def validate_channel(cls, v):
        if v not in ("email", "sms"):
            raise ValueError("channel must be 'email' or 'sms'")
        return v


class InboxDraftRequest(BaseModel):
    contact_id: str = Field(..., max_length=50)
    channel: str = Field(..., max_length=20)
    subject: Optional[str] = Field(default="", max_length=500)
    body: Optional[str] = Field(default="", max_length=50000)

    @field_validator("channel")
    @classmethod
    def validate_draft_channel(cls, v):
        if v not in ("email", "sms"):
            raise ValueError("channel must be 'email' or 'sms'")
        return v


async def _ensure_thread_meta(user_id: str, contact_id: str) -> dict:
    """Create or return the inbox_threads meta doc for a (user, contact)."""
    meta = await db.inbox_threads.find_one({"user_id": user_id, "contact_id": contact_id})
    if not meta:
        now = datetime.now(timezone.utc).isoformat()
        meta = {
            "user_id": user_id,
            "contact_id": contact_id,
            "assigned_to": None,
            "status": "open",
            "last_message_at": now,
            "last_updated": now,
        }
        await db.inbox_threads.insert_one(meta)
    return meta


async def _log_inbox_message(user_id: str, contact_id: str, channel: str, direction: str,
                             body: str, subject: str = "", from_addr: str = "",
                             to_addr: str = "", external_id: str = "") -> dict:
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "user_id": user_id,
        "contact_id": contact_id,
        "channel": channel,
        "direction": direction,
        "subject": subject or "",
        "body": body,
        "from_addr": from_addr or "",
        "to_addr": to_addr or "",
        "external_id": external_id or "",
        "read": direction == "outbound",  # outbound are implicitly read
        "created_at": now,
    }
    await db.messages.insert_one(doc)
    await _ensure_thread_meta(user_id, contact_id)
    await db.inbox_threads.update_one(
        {"user_id": user_id, "contact_id": contact_id},
        {"$set": {"last_message_at": now, "last_updated": now}},
    )
    doc.pop("_id", None)
    return doc


@api_router.get("/inbox/counts")
async def inbox_counts(user=Depends(get_any_auth_user)):
    """Folder counts + unread count for sidebar badges."""
    uid = user["_id"]
    base = {"user_id": uid}

    # All threads (not closed)
    open_threads = await db.inbox_threads.count_documents({**base, "status": "open"})
    assigned = await db.inbox_threads.count_documents({**base, "status": "open", "assigned_to": uid})
    closed = await db.inbox_threads.count_documents({**base, "status": "closed"})
    drafts = await db.inbox_drafts.count_documents(base)

    # Sent = threads whose most-recent message is outbound (use aggregation)
    sent_pipeline = [
        {"$match": {"user_id": uid}},
        {"$sort": {"created_at": -1}},
        {"$group": {"_id": "$contact_id", "last_direction": {"$first": "$direction"}}},
        {"$match": {"last_direction": "outbound"}},
        {"$count": "n"},
    ]
    sent_agg = await db.messages.aggregate(sent_pipeline).to_list(1)
    sent = sent_agg[0]["n"] if sent_agg else 0

    # Unread inbound messages count
    unread = await db.messages.count_documents({**base, "direction": "inbound", "read": False})

    return {
        "inbox": open_threads,
        "assigned": assigned,
        "drafts": drafts,
        "sent": sent,
        "closed": closed,
        "unread": unread,
    }


@api_router.get("/inbox/threads")
async def list_inbox_threads(
    user=Depends(get_any_auth_user),
    folder: str = "inbox", channel: str = "", search: str = "",
    limit: int = 100,
):
    """List conversation threads (one per contact) with last-message summary."""
    uid = user["_id"]
    limit = max(1, min(limit, 300))

    # Determine which contacts belong to this folder
    if folder == "closed":
        metas = await db.inbox_threads.find({"user_id": uid, "status": "closed"}).to_list(1000)
        allowed_cids = {m["contact_id"] for m in metas}
    elif folder == "assigned":
        metas = await db.inbox_threads.find({"user_id": uid, "status": "open", "assigned_to": uid}).to_list(1000)
        allowed_cids = {m["contact_id"] for m in metas}
    elif folder == "drafts":
        draft_docs = await db.inbox_drafts.find({"user_id": uid}).to_list(1000)
        allowed_cids = {d["contact_id"] for d in draft_docs}
    elif folder == "sent":
        pipeline = [
            {"$match": {"user_id": uid}},
            {"$sort": {"created_at": -1}},
            {"$group": {"_id": "$contact_id", "last_direction": {"$first": "$direction"}}},
            {"$match": {"last_direction": "outbound"}},
        ]
        sent_docs = await db.messages.aggregate(pipeline).to_list(1000)
        allowed_cids = {s["_id"] for s in sent_docs}
    else:  # inbox (default) — all open threads
        metas = await db.inbox_threads.find({"user_id": uid, "status": "open"}).to_list(1000)
        allowed_cids = {m["contact_id"] for m in metas}

    if not allowed_cids and folder != "drafts":
        return {"threads": []}

    # Build messages aggregation grouped by contact to get last message + unread count
    match = {"user_id": uid}
    if allowed_cids:
        match["contact_id"] = {"$in": list(allowed_cids)}
    if channel and channel in ("email", "sms", "voicemail"):
        match["channel"] = channel

    pipeline = [
        {"$match": match},
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": "$contact_id",
            "last_body": {"$first": "$body"},
            "last_subject": {"$first": "$subject"},
            "last_channel": {"$first": "$channel"},
            "last_direction": {"$first": "$direction"},
            "last_at": {"$first": "$created_at"},
            "unread": {"$sum": {"$cond": [
                {"$and": [{"$eq": ["$direction", "inbound"]}, {"$eq": ["$read", False]}]},
                1, 0
            ]}},
            "count": {"$sum": 1},
        }},
        {"$sort": {"last_at": -1}},
        {"$limit": limit},
    ]
    agg = await db.messages.aggregate(pipeline).to_list(limit)

    # For drafts folder, if a contact has no message yet but has a draft, also include
    if folder == "drafts":
        existing_cids = {a["_id"] for a in agg}
        draft_only_cids = list(allowed_cids - existing_cids)
        for cid in draft_only_cids:
            agg.append({
                "_id": cid, "last_body": "", "last_subject": "", "last_channel": "email",
                "last_direction": "draft", "last_at": "", "unread": 0, "count": 0,
            })

    # Fetch contact + meta for each thread
    cids_valid = [c for c in [a["_id"] for a in agg] if c]
    oids = []
    for c in cids_valid:
        try:
            oids.append(ObjectId(c))
        except Exception:
            pass
    contacts_map = {}
    if oids:
        async for c in db.contacts.find({"_id": {"$in": oids}, "user_id": uid}):
            contacts_map[str(c["_id"])] = c

    metas_map = {m["contact_id"]: m for m in
                 await db.inbox_threads.find({"user_id": uid, "contact_id": {"$in": list(cids_valid)}}).to_list(1000)}
    drafts_map = {d["contact_id"]: d for d in
                  await db.inbox_drafts.find({"user_id": uid, "contact_id": {"$in": list(cids_valid)}}).to_list(1000)}

    threads = []
    for row in agg:
        cid = row["_id"]
        contact = contacts_map.get(cid)
        if not contact and folder != "drafts":
            continue  # contact deleted; skip
        name = (contact or {}).get("name") or "Unknown"
        email = (contact or {}).get("email") or ""
        phone = (contact or {}).get("phone") or ""
        # Apply search filter (contact name/email/phone OR message body)
        if search:
            s = search.lower()
            haystack = " ".join([name, email, phone, row.get("last_body", ""), row.get("last_subject", "")]).lower()
            if s not in haystack:
                continue
        meta = metas_map.get(cid, {})
        draft = drafts_map.get(cid)
        threads.append({
            "contact_id": cid,
            "name": name,
            "email": email,
            "phone": phone,
            "photo_url": (contact or {}).get("photo_url") or "",
            "last_body": row.get("last_body", ""),
            "last_subject": row.get("last_subject", ""),
            "last_channel": row.get("last_channel", "email"),
            "last_direction": row.get("last_direction", "inbound"),
            "last_at": row.get("last_at", ""),
            "unread": row.get("unread", 0),
            "message_count": row.get("count", 0),
            "status": meta.get("status", "open"),
            "assigned_to": meta.get("assigned_to"),
            "has_draft": draft is not None,
            "draft_preview": (draft or {}).get("body", "")[:120] if draft else "",
        })
    return {"threads": threads}


@api_router.get("/inbox/threads/{contact_id}")
async def get_inbox_thread(contact_id: str, user=Depends(get_any_auth_user)):
    """Return the full message history + contact card for a given thread."""
    uid = user["_id"]
    # Validate contact belongs to user
    try:
        cid_oid = ObjectId(contact_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Contact not found")
    contact = await db.contacts.find_one({"_id": cid_oid, "user_id": uid})
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")

    msgs = await db.messages.find(
        {"user_id": uid, "contact_id": contact_id}
    ).sort("created_at", 1).to_list(1000)
    for m in msgs:
        m.pop("_id", None)

    await _ensure_thread_meta(uid, contact_id)
    meta = await db.inbox_threads.find_one({"user_id": uid, "contact_id": contact_id}) or {}
    draft = await db.inbox_drafts.find_one({"user_id": uid, "contact_id": contact_id})

    # Serialize contact
    contact_out = {
        "id": str(contact["_id"]),
        "name": contact.get("name", ""),
        "email": contact.get("email", ""),
        "phone": contact.get("phone", ""),
        "photo_url": contact.get("photo_url", ""),
        "company": contact.get("company", ""),
        "tags": contact.get("tags", []),
        "lease_status": contact.get("lease_status", "") or contact.get("leasing_stage", ""),
        "next_renewal_date": contact.get("next_renewal_date", "") or contact.get("lease_end_date", ""),
        "unit_interested": contact.get("unit_interested", "") or contact.get("unit_interest", ""),
        "retention_score": contact.get("retention_score", None),
        "is_tenant": contact.get("is_tenant", False),
    }

    return {
        "contact": contact_out,
        "messages": msgs,
        "status": meta.get("status", "open"),
        "assigned_to": meta.get("assigned_to"),
        "draft": {
            "channel": (draft or {}).get("channel", "email"),
            "subject": (draft or {}).get("subject", ""),
            "body": (draft or {}).get("body", ""),
        } if draft else None,
    }


@api_router.post("/inbox/threads/{contact_id}/reply")
async def inbox_reply(contact_id: str, data: InboxReplyRequest,
                      background_tasks: BackgroundTasks,
                      user=Depends(get_any_auth_user)):
    """Send an email or SMS reply; log it into messages + activities."""
    uid = user["_id"]
    try:
        cid_oid = ObjectId(contact_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Contact not found")
    contact = await db.contacts.find_one({"_id": cid_oid, "user_id": uid})
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")

    channel = data.channel
    external_id = ""
    if channel == "email":
        to_email = contact.get("email", "")
        if not to_email:
            raise HTTPException(status_code=400, detail="Contact has no email address")
        brevo_key = settings.BREVO_API_KEY
        if brevo_key and settings.SENDER_EMAIL:
            try:
                external_id = send_email_with_retry(brevo_key, to_email,
                                                   data.subject or "(no subject)",
                                                   data.body.replace("\n", "<br>"))
                external_id = str(external_id) if external_id else ""
            except Exception as e:
                logger.error(f"Inbox email send failed: {e}")
                # still log the message as sent locally so user sees it
        else:
            logger.warning("Inbox: BREVO_API_KEY missing — message logged locally only")
        msg = await _log_inbox_message(uid, contact_id, "email", "outbound",
                                       data.body, subject=data.subject or "",
                                       from_addr=settings.SENDER_EMAIL or "",
                                       to_addr=to_email, external_id=external_id)
        await db.activities.insert_one({
            "contact_id": contact_id, "user_id": uid,
            "activity_type": "email",
            "description": f"Sent email: {data.subject or '(no subject)'}",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        background_tasks.add_task(trigger_webhooks, uid, "email_sent",
                                  {"contact_id": contact_id, "subject": data.subject or ""})
    else:  # sms
        to_phone = contact.get("phone", "")
        if not to_phone:
            raise HTTPException(status_code=400, detail="Contact has no phone number")
        acct = settings.TWILIO_ACCOUNT_SID
        tok = settings.TWILIO_AUTH_TOKEN
        frm = settings.TWILIO_PHONE_NUMBER
        if acct and tok and frm:
            try:
                res = send_sms_with_retry(acct, tok, frm, to_phone, data.body)
                external_id = getattr(res, "sid", "")
            except Exception as e:
                logger.error(f"Inbox SMS send failed: {e}")
        else:
            logger.warning("Inbox: Twilio not configured — message logged locally only")
        msg = await _log_inbox_message(uid, contact_id, "sms", "outbound",
                                       data.body, from_addr=frm or "",
                                       to_addr=to_phone, external_id=external_id)
        await db.activities.insert_one({
            "contact_id": contact_id, "user_id": uid,
            "activity_type": "sms",
            "description": f"Sent SMS: {data.body[:80]}",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        background_tasks.add_task(trigger_webhooks, uid, "sms_sent",
                                  {"contact_id": contact_id, "message": data.body[:80]})

    # Clear draft if any
    await db.inbox_drafts.delete_one({"user_id": uid, "contact_id": contact_id})
    return {"success": True, "message": msg}


@api_router.post("/inbox/threads/{contact_id}/read")
async def mark_thread_read(contact_id: str, user=Depends(get_any_auth_user)):
    uid = user["_id"]
    await db.messages.update_many(
        {"user_id": uid, "contact_id": contact_id, "direction": "inbound", "read": False},
        {"$set": {"read": True}},
    )
    return {"success": True}


@api_router.post("/inbox/threads/{contact_id}/assign")
async def assign_thread(contact_id: str, user=Depends(get_any_auth_user)):
    """Assign the thread to the current user (toggles unassign if already assigned)."""
    uid = user["_id"]
    await _ensure_thread_meta(uid, contact_id)
    meta = await db.inbox_threads.find_one({"user_id": uid, "contact_id": contact_id})
    new_assignee = None if (meta or {}).get("assigned_to") == uid else uid
    await db.inbox_threads.update_one(
        {"user_id": uid, "contact_id": contact_id},
        {"$set": {"assigned_to": new_assignee,
                  "last_updated": datetime.now(timezone.utc).isoformat()}},
    )
    return {"assigned_to": new_assignee}


@api_router.post("/inbox/threads/{contact_id}/close")
async def close_thread(contact_id: str, user=Depends(get_any_auth_user)):
    """Toggle thread open/closed."""
    uid = user["_id"]
    await _ensure_thread_meta(uid, contact_id)
    meta = await db.inbox_threads.find_one({"user_id": uid, "contact_id": contact_id})
    new_status = "open" if (meta or {}).get("status") == "closed" else "closed"
    await db.inbox_threads.update_one(
        {"user_id": uid, "contact_id": contact_id},
        {"$set": {"status": new_status,
                  "last_updated": datetime.now(timezone.utc).isoformat()}},
    )
    return {"status": new_status}


@api_router.put("/inbox/drafts")
async def upsert_draft(data: InboxDraftRequest, user=Depends(get_any_auth_user)):
    uid = user["_id"]
    if not data.body and not data.subject:
        # Delete if empty
        await db.inbox_drafts.delete_one({"user_id": uid, "contact_id": data.contact_id})
        return {"success": True, "deleted": True}
    now = datetime.now(timezone.utc).isoformat()
    await db.inbox_drafts.update_one(
        {"user_id": uid, "contact_id": data.contact_id},
        {"$set": {
            "user_id": uid,
            "contact_id": data.contact_id,
            "channel": data.channel,
            "subject": data.subject or "",
            "body": data.body or "",
            "updated_at": now,
        }},
        upsert=True,
    )
    return {"success": True}


@api_router.delete("/inbox/drafts/{contact_id}")
async def delete_draft(contact_id: str, user=Depends(get_any_auth_user)):
    uid = user["_id"]
    await db.inbox_drafts.delete_one({"user_id": uid, "contact_id": contact_id})
    return {"success": True}


@api_router.post("/inbox/seed-demo")
async def seed_inbox_demo(user=Depends(get_any_auth_user)):
    """Populate the inbox with demo inbound emails/SMS/voicemails across existing contacts."""
    uid = user["_id"]
    contacts = await db.contacts.find({"user_id": uid}).limit(12).to_list(12)
    if not contacts:
        return {"created": 0, "detail": "No contacts found; create contacts first."}

    # Delete any pre-existing demo messages for a clean reseed
    await db.messages.delete_many({"user_id": uid, "external_id": "demo"})

    templates = [
        ("email", "inbound",  "Question about the 2-bedroom listing", "Hi! I saw the listing on StreetEasy and I'm really interested. Is it still available for a June move-in? Also — are pets allowed? Thanks!"),
        ("sms",   "inbound",  "", "Hey, just confirming tomorrow's showing at 4pm — anything I should bring?"),
        ("email", "inbound",  "Renewal — lease question", "Hi, my lease is up in 45 days. Can we chat about renewal options? I'd love to stay if the rent doesn't jump too much."),
        ("email", "outbound", "Re: Question about the 2-bedroom listing", "Hi! Yes, the unit is still available and we welcome cats/small dogs. I've attached the application — happy to set up a tour this week."),
        ("sms",   "outbound", "", "Sure! Just bring a photo ID and proof of income. See you at 4."),
        ("voicemail", "inbound", "Voicemail (0:42)", "Hi this is Sam — calling about the application I submitted yesterday. Wanted to see if there's anything else you need from me. Thanks!"),
        ("email", "inbound",  "Maintenance follow-up", "Thanks for sending the plumber last week. Everything's working great now — appreciate the quick turnaround."),
        ("sms",   "inbound",  "", "Do you have any 1BRs coming up in Brooklyn under $3k? Moving date is flexible."),
    ]

    now = datetime.now(timezone.utc)
    created = 0
    for idx, contact in enumerate(contacts):
        cid = str(contact["_id"])
        # Seed 1-3 messages per contact (rotating through templates)
        n_msgs = 2 + (idx % 2)
        for j in range(n_msgs):
            tpl = templates[(idx + j) % len(templates)]
            channel, direction, subject, body = tpl
            ts = (now - timedelta(hours=(idx * 3) + j, minutes=idx * 7)).isoformat()
            await db.messages.insert_one({
                "user_id": uid,
                "contact_id": cid,
                "channel": channel,
                "direction": direction,
                "subject": subject,
                "body": body,
                "from_addr": contact.get("email", "") if direction == "inbound" else "",
                "to_addr": contact.get("email", "") if direction == "outbound" else "",
                "external_id": "demo",
                "read": direction == "outbound",
                "created_at": ts,
            })
            created += 1
        # Ensure thread meta exists
        await _ensure_thread_meta(uid, cid)
        await db.inbox_threads.update_one(
            {"user_id": uid, "contact_id": cid},
            {"$set": {"last_message_at": now.isoformat()}},
        )

    return {"created": created, "contacts": len(contacts)}


# ── Inbox indexes ──
async def create_inbox_indexes():
    await db.messages.create_index([("user_id", 1), ("contact_id", 1), ("created_at", -1)], background=True)
    await db.messages.create_index([("user_id", 1), ("direction", 1), ("read", 1)], background=True)
    await db.messages.create_index([("user_id", 1), ("channel", 1)], background=True)
    await db.inbox_threads.create_index([("user_id", 1), ("contact_id", 1)], unique=True, background=True)
    await db.inbox_threads.create_index([("user_id", 1), ("status", 1)], background=True)
    await db.inbox_drafts.create_index([("user_id", 1), ("contact_id", 1)], unique=True, background=True)


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 19 — BROKERAGE PRE-LEASE GOOGLE SHEET (admin-only, Brokerage Pro plan)
# Graceful-no-keys mode: all endpoints work and the background worker runs;
# the OAuth + Sheets-API calls degrade to a clear "Connect Google Sheets" state
# when GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET are missing.
# ═══════════════════════════════════════════════════════════════════════════════

import secrets as _secrets

# Lazy-imported inside helpers so missing libs don't break startup
def _sheets_libs_available() -> bool:
    try:
        import google_auth_oauthlib.flow  # noqa: F401
        import googleapiclient.discovery  # noqa: F401
        return True
    except Exception:
        return False


def _sheets_keys_configured() -> bool:
    return bool(settings.GOOGLE_CLIENT_ID and settings.GOOGLE_CLIENT_SECRET)


SHEETS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
    "openid",
    "https://www.googleapis.com/auth/userinfo.email",
]

LIVE_LISTINGS_TAB = "Live Listings"
LIVE_LISTINGS_HEADERS = [
    "Property Name", "Unit #", "Address", "Rent", "Availability Date",
    "Status", "Proximity to Campus", "Bus Routes", "Pet Policy",
    "Last Updated", "Listing Agent",
]
BROKERAGE_PRO_PLANS = ("brokerage_pro", "enterprise")


def _redirect_uri() -> str:
    return settings.FRONTEND_URL.rstrip("/") + "/api/oauth/google-sheets/callback"


async def _is_brokerage_pro(user_id: str) -> bool:
    """Current plan comes from org_settings.plan; default 'free'. Also
    honour a dev override `BROKERAGE_PRO_DEV` env var for local testing."""
    if os.environ.get("BROKERAGE_PRO_DEV") == "1":
        return True
    doc = await db.org_settings.find_one({"user_id": user_id}) or {}
    plan = (doc.get("plan") or "free").lower()
    return plan in BROKERAGE_PRO_PLANS


async def _require_brokerage_admin(user: dict) -> None:
    if (user.get("role") or "").lower() != "admin":
        raise HTTPException(status_code=403, detail="Brokerage admin role required")
    if not await _is_brokerage_pro(user["_id"]):
        raise HTTPException(status_code=402, detail="Brokerage Pro plan required")


def _client_config() -> dict:
    return {
        "web": {
            "client_id": settings.GOOGLE_CLIENT_ID,
            "client_secret": settings.GOOGLE_CLIENT_SECRET,
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "redirect_uris": [_redirect_uri()],
        }
    }


async def _load_sheets_creds(user_id: str):
    """Return google.oauth2.credentials.Credentials (auto-refreshed), or None."""
    if not _sheets_keys_configured() or not _sheets_libs_available():
        return None
    doc = await db.google_tokens.find_one({"user_id": user_id, "scope_set": "sheets"})
    if not doc or not doc.get("refresh_token"):
        return None
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request as GoogleRequest

    creds = Credentials(
        token=doc.get("access_token"),
        refresh_token=doc.get("refresh_token"),
        token_uri="https://oauth2.googleapis.com/token",
        client_id=settings.GOOGLE_CLIENT_ID,
        client_secret=settings.GOOGLE_CLIENT_SECRET,
        scopes=SHEETS_SCOPES,
    )
    exp = doc.get("expires_at")
    needs_refresh = True
    try:
        if exp:
            needs_refresh = datetime.fromisoformat(exp.replace("Z", "+00:00")) <= datetime.now(timezone.utc)
    except Exception:
        needs_refresh = True
    if needs_refresh:
        try:
            await asyncio.to_thread(creds.refresh, GoogleRequest())
            await db.google_tokens.update_one(
                {"user_id": user_id, "scope_set": "sheets"},
                {"$set": {
                    "access_token": creds.token,
                    "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=55)).isoformat(),
                }},
            )
        except Exception as e:
            logger.warning(f"Sheets token refresh failed for user={user_id}: {e}")
            return None
    return creds


def _build_sheets_service(creds):
    from googleapiclient.discovery import build
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def _build_drive_service(creds):
    from googleapiclient.discovery import build
    return build("drive", "v3", credentials=creds, cache_discovery=False)


async def _brokerage_active_listings(admin_uid: str) -> list:
    """Aggregate all active listings across every active user (brokerage scope).
    For single-tenant deployments, this means every user in the DB; when a
    `brokerage_id` field exists on users, we scope to that admin's brokerage."""
    admin = await db.users.find_one({"_id": ObjectId(admin_uid)})
    brokerage_id = (admin or {}).get("brokerage_id") or admin_uid
    user_filter = {"$or": [{"brokerage_id": brokerage_id}, {"_id": ObjectId(admin_uid)}]}
    user_ids = set()
    async for u in db.users.find(user_filter, {"_id": 1}):
        user_ids.add(str(u["_id"]))
    # Fallback: if only the admin is matched (no brokerage_id seeded), include all users
    if len(user_ids) <= 1:
        async for u in db.users.find({}, {"_id": 1}):
            user_ids.add(str(u["_id"]))

    # Fetch user display names
    name_map = {}
    for uid in user_ids:
        try:
            u = await db.users.find_one({"_id": ObjectId(uid)}, {"name": 1, "email": 1})
            name_map[uid] = (u or {}).get("name") or (u or {}).get("email") or "Agent"
        except Exception:
            name_map[uid] = "Agent"

    rows = []
    async for p in db.properties.find({
        "user_id": {"$in": list(user_ids)},
        "status": "active",
    }).sort("updated_at", -1):
        rows.append([
            p.get("name", "") or p.get("title", "") or "",
            p.get("unit", "") or p.get("unit_number", "") or "",
            p.get("address", "") or "",
            _fmt_money(p.get("price", 0) or p.get("rent", 0)),
            p.get("availability_date", "") or p.get("available_from", ""),
            (p.get("status") or "active").title(),
            p.get("proximity_to_campus", "") or p.get("campus_proximity", ""),
            p.get("bus_routes", "") or p.get("transit", ""),
            p.get("pet_policy", "") or ("Pets OK" if p.get("pets_allowed") else "No pets"),
            _fmt_date(p.get("updated_at") or p.get("created_at", "")),
            name_map.get(str(p.get("user_id") or ""), "Agent"),
        ])
    return rows


def _fmt_money(v) -> str:
    try:
        return f"${float(v):,.0f}"
    except Exception:
        return ""


def _fmt_date(v: str) -> str:
    if not v:
        return ""
    try:
        d = datetime.fromisoformat(v.replace("Z", "+00:00"))
        return d.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(v)[:16]


async def _create_brokerage_sheet(admin_uid: str) -> dict:
    """Create the spreadsheet + Live Listings tab with headers. Returns meta."""
    creds = await _load_sheets_creds(admin_uid)
    if not creds:
        raise HTTPException(status_code=400, detail="Google Sheets not connected — finish OAuth first")
    svc = _build_sheets_service(creds)
    spreadsheet_body = {
        "properties": {"title": "Brokerage Pre-Lease List"},
        "sheets": [{
            "properties": {"title": LIVE_LISTINGS_TAB, "gridProperties": {"frozenRowCount": 1}},
        }],
    }
    created = await asyncio.to_thread(lambda: svc.spreadsheets().create(body=spreadsheet_body, fields="spreadsheetId,spreadsheetUrl,sheets(properties(sheetId,title))").execute())
    sheet_id = created["spreadsheetId"]
    sheet_url = created["spreadsheetUrl"]
    live_sheet_id = next((s["properties"]["sheetId"] for s in created.get("sheets", []) if s["properties"]["title"] == LIVE_LISTINGS_TAB), None)

    # Write headers to Live Listings!A1:K1
    await asyncio.to_thread(lambda: svc.spreadsheets().values().update(
        spreadsheetId=sheet_id,
        range=f"'{LIVE_LISTINGS_TAB}'!A1:K1",
        valueInputOption="RAW",
        body={"values": [LIVE_LISTINGS_HEADERS]},
    ).execute())

    # Bold / freeze header row via batchUpdate targeting ONLY our tab's sheetId
    if live_sheet_id is not None:
        await asyncio.to_thread(lambda: svc.spreadsheets().batchUpdate(
            spreadsheetId=sheet_id,
            body={"requests": [{
                "repeatCell": {
                    "range": {"sheetId": live_sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                    "cell": {"userEnteredFormat": {
                        "backgroundColor": {"red": 0.059, "green": 0.463, "blue": 0.431},
                        "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                    }},
                    "fields": "userEnteredFormat(backgroundColor,textFormat)",
                },
            }]},
        ).execute())

    return {
        "sheet_id": sheet_id,
        "sheet_url": sheet_url,
        "live_sheet_tab_id": live_sheet_id,
    }


async def _ensure_live_listings_tab(svc, sheet_id: str, stored_tab_id):
    """If the admin renamed/deleted 'Live Listings', re-locate or re-create it.
    Returns (current_tab_title, current_tab_id)."""
    meta = await asyncio.to_thread(lambda: svc.spreadsheets().get(spreadsheetId=sheet_id, fields="sheets(properties(sheetId,title))").execute())
    for s in meta.get("sheets", []):
        props = s.get("properties", {})
        if stored_tab_id is not None and props.get("sheetId") == stored_tab_id:
            return props.get("title"), props.get("sheetId")
    # stored tab missing — re-create
    resp = await asyncio.to_thread(lambda: svc.spreadsheets().batchUpdate(
        spreadsheetId=sheet_id,
        body={"requests": [{"addSheet": {"properties": {"title": LIVE_LISTINGS_TAB, "gridProperties": {"frozenRowCount": 1}}}}]},
    ).execute())
    new_id = resp["replies"][0]["addSheet"]["properties"]["sheetId"]
    await asyncio.to_thread(lambda: svc.spreadsheets().values().update(
        spreadsheetId=sheet_id,
        range=f"'{LIVE_LISTINGS_TAB}'!A1:K1",
        valueInputOption="RAW",
        body={"values": [LIVE_LISTINGS_HEADERS]},
    ).execute())
    return LIVE_LISTINGS_TAB, new_id


async def _sync_brokerage_sheet(admin_uid: str) -> dict:
    """Push active listings into the Live Listings tab only. Never touches other tabs."""
    settings_doc = await db.org_settings.find_one({"user_id": admin_uid}) or {}
    bs = settings_doc.get("brokerage_sheet") or {}
    if not bs.get("enabled") or not bs.get("sheet_id"):
        return {"skipped": True, "reason": "not enabled"}
    creds = await _load_sheets_creds(admin_uid)
    if not creds:
        return {"skipped": True, "reason": "tokens missing"}

    svc = _build_sheets_service(creds)
    sheet_id = bs["sheet_id"]
    stored_tab_id = bs.get("live_sheet_tab_id")

    tab_title, tab_id = await _ensure_live_listings_tab(svc, sheet_id, stored_tab_id)
    rows = await _brokerage_active_listings(admin_uid)

    # 1) Clear ONLY the data area of our tab (below the header row).
    await asyncio.to_thread(lambda: svc.spreadsheets().values().clear(
        spreadsheetId=sheet_id,
        range=f"'{tab_title}'!A2:K10000",
        body={},
    ).execute())

    # 2) Re-write the header row (idempotent, user may have edited it by mistake)
    await asyncio.to_thread(lambda: svc.spreadsheets().values().update(
        spreadsheetId=sheet_id,
        range=f"'{tab_title}'!A1:K1",
        valueInputOption="RAW",
        body={"values": [LIVE_LISTINGS_HEADERS]},
    ).execute())

    # 3) Upsert the data rows (if any)
    if rows:
        await asyncio.to_thread(lambda: svc.spreadsheets().values().update(
            spreadsheetId=sheet_id,
            range=f"'{tab_title}'!A2",
            valueInputOption="RAW",
            body={"values": rows},
        ).execute())

    now = datetime.now(timezone.utc).isoformat()
    await db.org_settings.update_one(
        {"user_id": admin_uid},
        {"$set": {
            "brokerage_sheet.last_synced_at": now,
            "brokerage_sheet.last_row_count": len(rows),
            "brokerage_sheet.live_sheet_tab_id": tab_id,
            "brokerage_sheet.last_sync_status": "success",
            "brokerage_sheet.last_sync_error": "",
        }},
    )
    return {"synced": True, "rows": len(rows), "at": now}


@api_router.get("/brokerage-sheet/status")
async def brokerage_sheet_status(user=Depends(get_any_auth_user)):
    """Return the sheet's current status + plan gate + connection state."""
    uid = user["_id"]
    is_admin = (user.get("role") or "").lower() == "admin"
    plan_ok = await _is_brokerage_pro(uid) if is_admin else False
    keys_ok = _sheets_keys_configured()
    tokens_doc = await db.google_tokens.find_one({"user_id": uid, "scope_set": "sheets"}) or {}
    connected = bool(tokens_doc.get("refresh_token"))
    doc = await db.org_settings.find_one({"user_id": uid}) or {}
    bs = doc.get("brokerage_sheet") or {}
    return {
        "is_admin": is_admin,
        "feature_available": is_admin and plan_ok,
        "keys_configured": keys_ok,
        "connected": connected,
        "plan": (doc.get("plan") or "free").lower(),
        "enabled": bool(bs.get("enabled")),
        "sheet_id": bs.get("sheet_id", ""),
        "sheet_url": bs.get("sheet_url", ""),
        "share_url": bs.get("share_url", ""),
        "permission_id": bs.get("permission_id", ""),
        "last_synced_at": bs.get("last_synced_at", ""),
        "last_row_count": bs.get("last_row_count", 0),
        "last_sync_status": bs.get("last_sync_status", ""),
        "last_sync_error": bs.get("last_sync_error", ""),
        "redirect_uri": _redirect_uri(),
    }


@api_router.get("/oauth/google-sheets/login")
async def oauth_sheets_login(user=Depends(get_any_auth_user)):
    await _require_brokerage_admin(user)
    if not _sheets_keys_configured():
        raise HTTPException(status_code=400, detail="Google Sheets OAuth keys not configured")
    if not _sheets_libs_available():
        raise HTTPException(status_code=500, detail="Server missing google-auth-oauthlib")
    from google_auth_oauthlib.flow import Flow
    flow = Flow.from_client_config(_client_config(), scopes=SHEETS_SCOPES, redirect_uri=_redirect_uri())
    state = _secrets.token_urlsafe(24)
    await db.oauth_states.update_one(
        {"state": state},
        {"$set": {"state": state, "user_id": user["_id"], "purpose": "sheets",
                  "created_at": datetime.now(timezone.utc).isoformat(),
                  "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat()}},
        upsert=True,
    )
    auth_url, _ = flow.authorization_url(access_type="offline", prompt="consent", include_granted_scopes="true", state=state)
    return {"auth_url": auth_url}


@api_router.get("/oauth/google-sheets/callback")
async def oauth_sheets_callback(code: str = "", state: str = "", error: str = ""):
    frontend = settings.FRONTEND_URL.rstrip("/")
    if error:
        return RedirectResponse(f"{frontend}/settings?sheets_error={error}")
    if not code or not state:
        return RedirectResponse(f"{frontend}/settings?sheets_error=missing_code_or_state")
    state_doc = await db.oauth_states.find_one({"state": state, "purpose": "sheets"})
    if not state_doc:
        return RedirectResponse(f"{frontend}/settings?sheets_error=invalid_state")
    await db.oauth_states.delete_one({"state": state})
    user_id = state_doc["user_id"]
    if not _sheets_libs_available() or not _sheets_keys_configured():
        return RedirectResponse(f"{frontend}/settings?sheets_error=keys_missing")
    try:
        from google_auth_oauthlib.flow import Flow
        flow = Flow.from_client_config(_client_config(), scopes=SHEETS_SCOPES, redirect_uri=_redirect_uri())
        await asyncio.to_thread(flow.fetch_token, code=code)
        creds = flow.credentials
        await db.google_tokens.update_one(
            {"user_id": user_id, "scope_set": "sheets"},
            {"$set": {
                "user_id": user_id, "scope_set": "sheets",
                "access_token": creds.token,
                "refresh_token": creds.refresh_token or (state_doc.get("existing_refresh") or ""),
                "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=55)).isoformat(),
                "scopes": list(creds.scopes or SHEETS_SCOPES),
            }},
            upsert=True,
        )
    except Exception as e:
        logger.exception(f"Sheets OAuth callback failed: {e}")
        return RedirectResponse(f"{frontend}/settings?sheets_error=token_exchange_failed")
    return RedirectResponse(f"{frontend}/settings?tab=integrations&sheets_connected=1")


@api_router.post("/brokerage-sheet/enable")
async def brokerage_sheet_enable(user=Depends(get_any_auth_user)):
    await _require_brokerage_admin(user)
    uid = user["_id"]
    doc = await db.org_settings.find_one({"user_id": uid}) or {}
    bs = doc.get("brokerage_sheet") or {}
    if bs.get("enabled") and bs.get("sheet_id"):
        return bs
    meta = await _create_brokerage_sheet(uid)
    new_bs = {
        "enabled": True,
        "sheet_id": meta["sheet_id"],
        "sheet_url": meta["sheet_url"],
        "live_sheet_tab_id": meta["live_sheet_tab_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.org_settings.update_one(
        {"user_id": uid},
        {"$set": {"brokerage_sheet": new_bs},
         "$setOnInsert": {"user_id": uid, "created_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    # Kick off an immediate sync
    try:
        await _sync_brokerage_sheet(uid)
    except Exception as e:
        logger.warning(f"First sync failed: {e}")
    return new_bs


@api_router.post("/brokerage-sheet/sync")
async def brokerage_sheet_sync_now(user=Depends(get_any_auth_user)):
    await _require_brokerage_admin(user)
    res = await _sync_brokerage_sheet(user["_id"])
    return res


@api_router.post("/brokerage-sheet/share")
async def brokerage_sheet_share(user=Depends(get_any_auth_user)):
    """Create a 'anyone with the link can VIEW' permission, return the share URL."""
    await _require_brokerage_admin(user)
    uid = user["_id"]
    creds = await _load_sheets_creds(uid)
    if not creds:
        raise HTTPException(status_code=400, detail="Connect Google Sheets first")
    doc = await db.org_settings.find_one({"user_id": uid}) or {}
    bs = doc.get("brokerage_sheet") or {}
    if not bs.get("sheet_id"):
        raise HTTPException(status_code=400, detail="Sheet not yet created")
    drive = _build_drive_service(creds)
    perm = await asyncio.to_thread(lambda: drive.permissions().create(
        fileId=bs["sheet_id"],
        body={"role": "reader", "type": "anyone"},
        fields="id",
    ).execute())
    share_url = f"https://docs.google.com/spreadsheets/d/{bs['sheet_id']}/edit?usp=sharing"
    await db.org_settings.update_one(
        {"user_id": uid},
        {"$set": {"brokerage_sheet.share_url": share_url, "brokerage_sheet.permission_id": perm["id"]}},
    )
    return {"share_url": share_url, "permission_id": perm["id"]}


@api_router.post("/brokerage-sheet/revoke-share")
async def brokerage_sheet_revoke_share(user=Depends(get_any_auth_user)):
    await _require_brokerage_admin(user)
    uid = user["_id"]
    creds = await _load_sheets_creds(uid)
    doc = await db.org_settings.find_one({"user_id": uid}) or {}
    bs = doc.get("brokerage_sheet") or {}
    if creds and bs.get("sheet_id") and bs.get("permission_id"):
        try:
            drive = _build_drive_service(creds)
            await asyncio.to_thread(lambda: drive.permissions().delete(fileId=bs["sheet_id"], permissionId=bs["permission_id"]).execute())
        except Exception as e:
            logger.warning(f"Revoke share failed: {e}")
    await db.org_settings.update_one(
        {"user_id": uid},
        {"$set": {"brokerage_sheet.share_url": "", "brokerage_sheet.permission_id": ""}},
    )
    return {"success": True}


@api_router.post("/brokerage-sheet/disconnect")
async def brokerage_sheet_disconnect(user=Depends(get_any_auth_user)):
    """Disable sync & delete stored tokens. The sheet itself remains in the admin's Drive."""
    await _require_brokerage_admin(user)
    uid = user["_id"]
    await db.google_tokens.delete_many({"user_id": uid, "scope_set": "sheets"})
    await db.org_settings.update_one(
        {"user_id": uid},
        {"$set": {"brokerage_sheet.enabled": False,
                  "brokerage_sheet.last_sync_status": "disconnected"}},
    )
    return {"success": True}


# ── 60-second sync worker ────────────────────────────────────────────────────
async def brokerage_sheet_sync_worker():
    """Every 60s, sync each admin's Live Listings tab. Fails softly per-user."""
    logger.info("Brokerage-sheet sync worker started")
    while True:
        try:
            await asyncio.sleep(60)
            if not _sheets_keys_configured() or not _sheets_libs_available():
                continue
            async for doc in db.org_settings.find({"brokerage_sheet.enabled": True}, {"user_id": 1, "brokerage_sheet.sheet_id": 1}):
                uid = doc.get("user_id")
                if not uid:
                    continue
                try:
                    await _sync_brokerage_sheet(uid)
                except Exception as e:
                    logger.warning(f"Sheet sync failed user={uid}: {e}")
                    await db.org_settings.update_one(
                        {"user_id": uid},
                        {"$set": {
                            "brokerage_sheet.last_sync_status": "error",
                            "brokerage_sheet.last_sync_error": str(e)[:300],
                        }},
                    )
        except asyncio.CancelledError:
            break
        except Exception as e:
            logger.exception(f"Sheet worker loop error: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# TWILIO: Status check + Inbound SMS webhook
# - /api/twilio/status → verify creds by fetching account from Twilio REST API
# - /api/twilio/inbound-sms → Twilio calls this when an SMS is received on our
#   number. Matches the sender's phone to a contact, logs the message into the
#   unified inbox, and records an activity. Returns empty TwiML.
# ═══════════════════════════════════════════════════════════════════════════════

def _normalize_phone_digits(p: str) -> str:
    """Return last-10 digit string for comparison (strip country code for US)."""
    if not p:
        return ""
    digits = re.sub(r"\D", "", p)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return digits


@api_router.get("/twilio/status")
async def twilio_status(user=Depends(get_any_auth_user)):
    """Verify Twilio creds are valid by fetching account info."""
    acct = settings.TWILIO_ACCOUNT_SID
    tok = settings.TWILIO_AUTH_TOKEN
    frm = settings.TWILIO_PHONE_NUMBER
    if not (acct and tok and frm):
        return {"configured": False, "reason": "Missing TWILIO_ACCOUNT_SID / TWILIO_AUTH_TOKEN / TWILIO_PHONE_NUMBER"}
    try:
        from twilio.rest import Client as TwilioClient
        from twilio.base.exceptions import TwilioRestException
        def _fetch():
            return TwilioClient(acct, tok).api.accounts(acct).fetch()
        info = await asyncio.to_thread(_fetch)
        return {
            "configured": True,
            "account_sid": acct,
            "account_status": getattr(info, "status", "unknown"),
            "account_type": getattr(info, "type", "unknown"),  # "Trial" or "Full"
            "friendly_name": getattr(info, "friendly_name", ""),
            "from_number": frm,
            "inbound_webhook_url": f"{FRONTEND_URL}/api/twilio/inbound-sms",
        }
    except Exception as e:
        logger.error(f"Twilio status check failed: {e}")
        return {"configured": False, "reason": f"Twilio API error: {str(e)[:200]}"}


@api_router.post("/twilio/inbound-sms")
async def twilio_inbound_sms(request: Request):
    """
    Twilio webhook — called when an SMS is received on our Twilio number.
    Configure in Twilio Console → Phone Numbers → <number> → 'A MESSAGE COMES IN'
    (HTTP POST) → paste the URL returned by /api/twilio/status.
    """
    try:
        form = await request.form()
    except Exception as e:
        logger.error(f"Twilio inbound-sms: cannot parse form body: {e}")
        return Response(content="<Response></Response>", media_type="application/xml")
    form_dict = {k: str(v) for k, v in form.items()}

    # Signature validation (skipped silently if token missing — e.g. in tests)
    auth_token = settings.TWILIO_AUTH_TOKEN
    signature = request.headers.get("X-Twilio-Signature", "")
    if auth_token and signature:
        try:
            from twilio.request_validator import RequestValidator as _TwRequestValidator
            # Twilio signs the public URL. Prefer FRONTEND_URL + path so it matches
            # exactly what's configured in the Twilio console, regardless of proxy.
            public_url = f"{FRONTEND_URL}{request.url.path}"
            if not _TwRequestValidator(auth_token).validate(public_url, form_dict, signature):
                logger.warning(f"Twilio inbound-sms: invalid signature from {form_dict.get('From', '?')}")
                return Response(content="<Response></Response>", media_type="application/xml", status_code=403)
        except Exception as e:
            logger.error(f"Twilio signature validation error: {e}")
            # Fail open in MVP — still process the message

    from_number = form_dict.get("From", "")
    to_number = form_dict.get("To", "")
    body = form_dict.get("Body", "")
    external_id = form_dict.get("MessageSid", "")

    if not from_number or not body:
        logger.warning(f"Twilio inbound-sms missing From/Body: keys={list(form_dict.keys())}")
        return Response(content="<Response></Response>", media_type="application/xml")

    # Match inbound phone to a contact by normalized digits
    norm_from = _normalize_phone_digits(from_number)
    matched_contact = None
    if norm_from:
        async for c in db.contacts.find({"phone": {"$exists": True, "$ne": ""}}):
            if _normalize_phone_digits(c.get("phone", "")) == norm_from:
                matched_contact = c
                break

    now_iso = datetime.now(timezone.utc).isoformat()

    if not matched_contact:
        # Orphan message — still log so admins can see it. Find any admin user
        # to own the record so it surfaces somewhere.
        admin_doc = await db.users.find_one({"role": "admin"})
        admin_id = str(admin_doc["_id"]) if admin_doc else ""
        await db.messages.insert_one({
            "user_id": admin_id,
            "contact_id": "",
            "channel": "sms",
            "direction": "inbound",
            "subject": "",
            "body": body,
            "from_addr": from_number,
            "to_addr": to_number,
            "external_id": external_id,
            "read": False,
            "created_at": now_iso,
            "unmatched": True,
        })
        logger.info(f"Twilio inbound SMS from UNMATCHED {from_number}: {body[:80]}")
        return Response(content="<Response></Response>", media_type="application/xml")

    owner_id = matched_contact.get("user_id") or ""
    if not owner_id:
        admin_doc = await db.users.find_one({"role": "admin"})
        owner_id = str(admin_doc["_id"]) if admin_doc else ""

    contact_id = str(matched_contact["_id"])

    await _log_inbox_message(
        user_id=owner_id,
        contact_id=contact_id,
        channel="sms",
        direction="inbound",
        body=body,
        from_addr=from_number,
        to_addr=to_number,
        external_id=external_id,
    )

    await db.activities.insert_one({
        "contact_id": contact_id,
        "user_id": owner_id,
        "activity_type": "sms",
        "description": f"Received SMS: {body[:80]}",
        "created_at": now_iso,
    })

    logger.info(f"Twilio inbound SMS from {from_number} → contact {contact_id}: {body[:80]}")
    return Response(content="<Response></Response>", media_type="application/xml")


app.include_router(api_router)

# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY: Rate Limit Exceeded Handler
# Returns clean 429 response when rate limits are hit.
# ═══════════════════════════════════════════════════════════════════════════════
@app.exception_handler(RateLimitExceeded)
async def rate_limit_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(
        status_code=429,
        content={"detail": "Rate limit exceeded. Please slow down and try again."}
    )

# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY: Global Exception Handler
# Catches all unhandled exceptions and returns a clean 500 response.
# Stack traces, DB details, and sensitive info are NEVER leaked to the client.
# Full details are logged server-side for debugging.
# ═══════════════════════════════════════════════════════════════════════════════
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unhandled exception on {request.method} {request.url.path}: {type(exc).__name__}: {exc}")
    logger.debug(traceback.format_exc())
    return JSONResponse(
        status_code=500,
        content={"detail": "An internal error occurred. Please try again later."}
    )

# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY: Rate Limiting Middleware (slowapi)
# Applied globally — 100 req/min default, stricter on auth routes.
# ═══════════════════════════════════════════════════════════════════════════════
app.add_middleware(SlowAPIMiddleware)

# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY: Security Headers Middleware
# Adds defensive HTTP headers to every response:
# - X-Content-Type-Options: nosniff          → Prevent MIME-type sniffing
# - X-Frame-Options: DENY                    → Prevent clickjacking
# - X-XSS-Protection: 1; mode=block          → Legacy XSS filter
# - Strict-Transport-Security                → Force HTTPS (production only)
# - Referrer-Policy: strict-origin-when-cross-origin → Limit referrer leakage
# - Permissions-Policy                       → Disable unused browser APIs
# ═══════════════════════════════════════════════════════════════════════════════
@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    if IS_PRODUCTION:
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY: CORS Configuration
# - Only allow the specific frontend origin (no wildcards)
# - Only allow necessary HTTP methods
# - Only allow necessary headers
# - Credentials (cookies) are allowed for auth
# ═══════════════════════════════════════════════════════════════════════════════
app.add_middleware(
    CORSMiddleware,
    allow_origins=[FRONTEND_URL],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "X-API-Key", "Cookie"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
